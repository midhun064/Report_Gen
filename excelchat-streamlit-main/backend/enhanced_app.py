"""
Enhanced Flask App with Full LLM-Based Agentic System
Implements complete Excel processing with full DataFrame loading and intelligent analysis
"""

# Load environment variables from .env file
from dotenv import load_dotenv
load_dotenv()

import os
import sys
import uuid
import time
import logging
import json
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional

from flask import Flask, request, jsonify, send_file, Response, stream_with_context
from flask_cors import CORS
import google.generativeai as genai
import pandas as pd

# Import configuration
from config.settings import settings, error_messages

# Import our enhanced services
from services.full_data_agent import FullDataAgent
from services.simple_chat import SimpleChatSystem

app = Flask(__name__)
CORS(app, origins=["http://localhost:5173", "http://127.0.0.1:5173"])

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def convert_numpy_types(obj):
    """Recursively convert numpy types to Python native types for JSON serialization"""
    if isinstance(obj, dict):
        return {key: convert_numpy_types(value) for key, value in obj.items()}
    elif isinstance(obj, list):
        return [convert_numpy_types(item) for item in obj]
    elif isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    else:
        return obj

def _extract_response_text(formatted_response):
    """Extract text from formatted response (handles both string and structured responses)"""
    if isinstance(formatted_response, dict):
        return formatted_response.get("text", "")
    return formatted_response

def _format_analysis_response(result: Dict[str, Any], agent=None):
    """
    Format the analysis result for display
    
    Args:
        result: Dictionary containing analysis results
        agent: Optional FullDataAgent instance for session info
    
    Returns:
        Formatted response string
    """
    try:
        # Handle Google Sheets operation results
        if result.get("type") == "google_sheet_operation":
            sheet_url = result.get("sheet_url", "")
            message = result.get("message", "")
            operations = result.get("operations_applied", [])
            
            # Create structured response with separate URL for frontend to render as hyperlink
            response_parts = [message]
            if operations:
                response_parts.append(f"\n**Operations applied**: {', '.join(operations)}")
            
            # Instead of including URL in text, return structured data
            # Frontend should handle URL rendering as hyperlink
            response_text = "\n".join(response_parts)
            
            # Return structured response that frontend can use to render hyperlink
            return {
                "text": response_text,
                "sheet_url": sheet_url,
                "has_link": bool(sheet_url),
                "link_text": "View updated sheet"
            }
        
        # Handle Excel operation results
        if result.get("operation_type") == "excel_file":
            exec_result = result.get("result", {})
            output = exec_result.get("execution_output", "")
            
            # Clean up the output - remove code blocks and technical details
            lines = output.split('\n')
            
            # Simple heuristic: if output is very short and simple, return as-is
            is_simple_response = (len(output.split('\n')) <= settings.SIMPLE_RESPONSE_MAX_LINES and 
                                 len(output) < settings.SIMPLE_RESPONSE_MAX_CHARS)
            
            if is_simple_response:
                # For simple queries, show clean, direct response
                response_parts = []
                
                if output:
                    # Look for natural LLM responses - no hardcoded patterns
                    lines = output.split('\n')
                    clean_lines = [line.strip() for line in lines if line.strip() and 
                                  not line.startswith('===') and 
                                  not line.startswith('DataFrame') and
                                  not line.startswith('Error:') and
                                  not 'print(' in line and
                                  not line.startswith('Chart successfully saved to') and
                                  not line.startswith('![Chart') and
                                  not line.startswith('Download:') and
                                  len(line.strip()) > 10]  # Meaningful content only
                    
                    if clean_lines:
                        # Use the most meaningful line from LLM output
                        response_parts.append(clean_lines[-1])
                
                # Charts are handled separately by the frontend through the charts array
                # No need to add chart URLs to the text response
                
                if response_parts:
                    return "\n".join(response_parts)
                
                # If there are charts, don't show dataset info (chart speaks for itself)
                if exec_result.get("charts"):
                    return ""  # Empty response - chart will be displayed
                
                # Fallback for simple queries - include multi-file info
                data_shape = result.get("data_shape", "")
                if data_shape:
                    # Check if this is a multi-file session
                    session_id = result.get("session_id")
                    if agent and session_id and session_id in agent.session_data:
                        session_data = agent.session_data[session_id]
                        files = session_data.get("files", {})
                        
                        if len(files) > 1:
                            # Multi-file response
                            file_info = []
                            for file_id, file_data in files.items():
                                df = file_data["df"]
                                file_info.append(f"**{file_data['filename']}**: {len(df)} rows, {len(df.columns)} columns")
                            
                            combined_info = f"**Combined Dataset**: {data_shape}"
                            return "\n".join(file_info + [combined_info])
                        else:
                            return f"**Dataset Information**: {data_shape}"
                    else:
                        return f"**Dataset Information**: {data_shape}"
                return "Query completed successfully."
            
            else:
                # For complex queries, let LLM generate natural responses
                message_parts = []
                if output:
                    # Filter out unwanted technical messages
                    lines = output.split('\n')
                    clean_lines = [line for line in lines if line.strip() and 
                                  not 'Chart successfully saved to' in line and
                                  not line.strip().startswith('![Chart') and
                                  not line.strip().startswith('Download:')]
                    
                    if clean_lines:
                        message_parts.append('\n'.join(clean_lines))
                
                # Charts are handled separately by the frontend through the charts array
                # No need to add chart URLs to the text response
                if not message_parts:
                    return "## ✅ Analysis Complete\nThe analysis has been completed successfully. The results have been processed and are ready for review."
                
                return "\n".join(message_parts)
        
        # Handle regular analysis results
        exec_result = result.get("result", {})
        output = exec_result.get("execution_output", "")
        
        # Check if this is a simple query response (short, direct output)
        is_simple_response = (len(output.split('\n')) <= settings.SIMPLE_RESPONSE_MAX_LINES and 
                             len(output) < settings.SIMPLE_RESPONSE_MAX_CHARS)
        
        if is_simple_response:
            # For simple queries, show clean, direct response
            response_parts = []
            
            if output:
                # Look for natural LLM responses - no hardcoded patterns
                lines = output.split('\n')
                clean_lines = [line.strip() for line in lines if line.strip() and 
                              not line.startswith('===') and 
                              not line.startswith('DataFrame') and
                              not line.startswith('Error:') and
                              not 'print(' in line and
                              not line.startswith('Chart successfully saved to') and
                              not line.startswith('![Chart') and
                              not line.startswith('Download:') and
                              len(line.strip()) > 10]  # Meaningful content only
                
                if clean_lines:
                    # Use the most meaningful line from LLM output
                    response_parts.append(clean_lines[-1])
            
            # Charts are handled separately by the frontend through the charts array
            # No need to add chart URLs to the text response
            
            if response_parts:
                return "\n".join(response_parts)
            
            # If there are charts, don't show dataset info (chart speaks for itself)
            if exec_result.get("charts"):
                return ""  # Empty response - chart will be displayed
            
            # Fallback for simple queries - include multi-file info
            data_shape = result.get("data_shape", "")
            if data_shape:
                # Check if this is a multi-file session
                session_id = result.get("session_id")
                if agent and session_id and session_id in agent.session_data:
                    session_data = agent.session_data[session_id]
                    files = session_data.get("files", {})
                    
                    if len(files) > 1:
                        # Multi-file response
                        file_info = []
                        for file_id, file_data in files.items():
                            df = file_data["df"]
                            file_info.append(f"**{file_data['filename']}**: {len(df)} rows, {len(df.columns)} columns")
                        
                        combined_info = f"**Combined Dataset**: {data_shape}"
                        return "\n".join(file_info + [combined_info])
                    else:
                        return f"**Dataset Information**: {data_shape}"
                else:
                    return f"**Dataset Information**: {data_shape}"
            return "Query completed successfully."
        
        else:
            # For complex queries, let LLM generate natural responses
            message_parts = []
            if output:
                # Filter out unwanted technical messages
                lines = output.split('\n')
                clean_lines = [line for line in lines if line.strip() and 
                              not 'Chart successfully saved to' in line and
                              not line.strip().startswith('![Chart') and
                              not line.strip().startswith('Download:')]
                
                if clean_lines:
                    message_parts.append('\n'.join(clean_lines))
            
            # Charts are handled separately by the frontend through the charts array
            # No need to add chart URLs to the text response
            if not message_parts:
                return "## ✅ Analysis Complete\nThe analysis has been completed successfully. The results have been processed and are ready for review."
            
            return "\n".join(message_parts)
    
    except Exception as e:
        logger.error(f"Error formatting response: {e}")
        return "Analysis completed successfully."

def create_enhanced_app():
    """Create enhanced Flask application with full data processing"""
    
    app = Flask(__name__)
    
    # Enable CORS for all routes
    CORS(app, origins=settings.get_cors_origins())
    
    # Load API key from environment (REQUIRED - no default for security)
    api_key = os.getenv('GEMINI_API_KEY')
    if not api_key:
        raise ValueError(error_messages.api_key_missing())
    
    # Initialize systems
    try:
        # Full data agent for Excel processing
        full_data_agent = FullDataAgent(api_key)
        logger.info("✅ Full Data Agent initialized successfully")
        
        # Simple chat system for basic conversations
        simple_chat = SimpleChatSystem(api_key)
        logger.info("✅ Simple Chat System initialized successfully")
        
    except Exception as e:
        logger.error(f"❌ Failed to initialize systems: {e}", exc_info=True)
        raise
    
    @app.route('/health', methods=['GET'])
    def health_check():
        return jsonify({
            "status": "healthy", 
            "version": "3.0.0-enhanced",
            "features": ["full_data_processing", "intelligent_profiling", "code_generation"]
        })
    
    @app.route('/api/upload-full', methods=['POST'])
    def upload_full_excel():
        """
        Enhanced upload endpoint that loads entire Excel file into memory
        
        Flow:
        1. File validation and optimization
        2. Full DataFrame loading
        3. Comprehensive data profiling
        4. Memory storage for instant access
        """
        try:
            from werkzeug.utils import secure_filename
            
            if 'file' not in request.files:
                return jsonify({"success": False, "error": "No file provided"}), 400

            file = request.files['file']
            if file.filename == '':
                return jsonify({"success": False, "error": "Empty filename"}), 400

            # Ensure uploads directory exists
            uploads_dir = Path(__file__).parent / "uploads"
            uploads_dir.mkdir(exist_ok=True)

            # Save file
            original_filename = file.filename
            filename = secure_filename(file.filename)
            file_path = uploads_dir / filename
            file.save(file_path)
            
            # Generate session ID
            session_id = request.form.get('session_id') or str(uuid.uuid4())
            
            logger.info(f"📁 Processing full Excel upload: {filename} for session {session_id}")
            
            # Generate unique file ID
            file_id = request.form.get('file_id') or f"file_{filename}_{int(time.time())}"
            
            # Load entire file using Full Data Agent with multi-file support
            result = full_data_agent.load_excel_file(str(file_path), session_id, file_id)
            
            if not result["success"]:
                return jsonify(result), 400
            
            # Check if this is a multi-sheet Excel file
            if result.get("is_multi_sheet", False):
                # Multi-sheet Excel file
                sheets = result.get("sheets", [])
                total_rows = result.get("total_rows", 0)
                
                response_data = {
                    "success": True,
                    "session_id": session_id,
                    "file_id": result["file_id"],
                    "filename": filename,
                    "original_filename": original_filename,
                    "is_multi_sheet": True,
                    "total_files": len(sheets),
                    "sheets": sheets,
                    "data_info": {
                        "rows": total_rows,
                        "columns": sum(s["columns"] for s in sheets),
                        "memory_usage_mb": 0,  # Will be calculated from sheets
                        "file_size_mb": 0,
                        "memory_saved_mb": 0
                    },
                    "message": result.get("message", f"✅ Loaded {len(sheets)} sheets with {total_rows:,} total rows")
                }
                
                logger.info(f"✅ Multi-sheet Excel upload successful: {filename} ({len(sheets)} sheets, {total_rows} rows)")
                return jsonify(response_data)
            
            # Single file upload (regular behavior)
            response_data = {
                "success": True,
                "session_id": session_id,
                "file_id": result["file_id"],
                "filename": filename,
                "original_filename": original_filename,
                "total_files": int(result["total_files"]),
                "data_info": {
                    "rows": int(result["rows"]),
                    "columns": int(result["columns"]),
                    "memory_usage_mb": round(float(result["memory_usage_mb"]), 2),
                    "file_size_mb": round(float(result["file_size_mb"]), 2),
                    "memory_saved_mb": round(float(result["memory_saved_mb"]), 2)
                },
                "profile_summary": {
                    "column_count": len(result["profile"]["basic_info"]["column_names"]),
                    "missing_values": int(sum(result["profile"]["data_quality"]["missing_values"].values())),
                    "duplicate_rows": int(result["profile"]["data_quality"]["duplicate_rows"]),
                    "numerical_columns": len(result["profile"]["numerical_analysis"]),
                    "categorical_columns": len(result["profile"]["categorical_analysis"])
                },
                "sample_data": result["profile"]["sample_data"]["head_5"][:3],  # First 3 rows for preview
                "message": f"✅ Loaded {int(result['rows']):,} rows × {int(result['columns'])} columns into memory. Ready for analysis!"
            }
            
            logger.info(f"✅ Full Excel upload successful: {filename} ({result['rows']} rows)")
            return jsonify(response_data)
            
        except Exception as e:
            logger.error(f"❌ Error in full Excel upload: {e}", exc_info=True)
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/upload-google-sheet', methods=['POST'])
    def upload_google_sheet():
        """
        Load Google Sheet(s) by URL - supports single or multiple sheets
        
        Request JSON:
        {
            "spreadsheet_url": "https://docs.google.com/spreadsheets/d/.../edit",
            "session_id": "optional-session-id",
            "sheet_name": "optional-sheet-name",
            "load_all_sheets": false  // Set to true to load all sheets/tabs
        }
        
        OR for multiple Google Sheets:
        {
            "spreadsheet_urls": [
                "https://docs.google.com/spreadsheets/d/.../edit",
                "https://docs.google.com/spreadsheets/d/.../edit"
            ],
            "session_id": "optional-session-id",
            "load_all_sheets": false
        }
        """
        try:
            data = request.get_json()
            
            # Check if it's multiple URLs or single URL
            spreadsheet_urls = data.get('spreadsheet_urls', [])
            single_url = data.get('spreadsheet_url')
            
            # Handle comma-separated URLs in single_url field as fallback
            if single_url and ',' in single_url:
                # Split comma-separated URLs and clean them
                urls = [url.strip() for url in single_url.split(',') if url.strip()]
                if len(urls) > 1:
                    spreadsheet_urls = urls
                    single_url = None  # Clear single_url to use multiple URLs path
                    logger.info(f"📊 Detected {len(urls)} comma-separated URLs, treating as multiple Google Sheets")
            
            if not data or (not single_url and not spreadsheet_urls):
                return jsonify({
                    "success": False, 
                    "error": "spreadsheet_url or spreadsheet_urls required"
                }), 400
            
            session_id = data.get('session_id') or str(uuid.uuid4())
            sheet_name = data.get('sheet_name')
            load_all_sheets = data.get('load_all_sheets', False)
            
            # Handle multiple Google Sheets
            if spreadsheet_urls:
                logger.info(f"📊 Loading {len(spreadsheet_urls)} Google Sheets for session {session_id}")
                
                all_results = []
                total_sheets = 0
                total_rows = 0
                
                for idx, url in enumerate(spreadsheet_urls):
                    logger.info(f"📊 Loading Google Sheet {idx+1}/{len(spreadsheet_urls)}: {url[:50]}...")
                    
                    # Generate unique file_id for each sheet
                    file_id = f"gsheet_{idx+1}_{int(time.time())}"
                    
                    result = full_data_agent.load_google_sheet(
                        url,
                        session_id,
                        sheet_name=sheet_name,
                        file_id=file_id
                    )
                    
                    if result["success"]:
                        all_results.append(result)
                        if result.get("is_multi_sheet", False):
                            total_sheets += len(result["sheets"])
                            total_rows += sum(sheet["rows"] for sheet in result["sheets"])
                        else:
                            total_sheets += 1
                            total_rows += result["rows"]
                    else:
                        logger.warning(f"⚠️ Failed to load Google Sheet {idx+1}: {result.get('error', 'Unknown error')}")
                
                if not all_results:
                    return jsonify({
                        "success": False,
                        "error": "Failed to load any Google Sheets"
                    }), 400
                
                # Match Excel's exact response format for consistency
                # Flatten all sheets from all results into a single sheets array (like Excel multi-sheet)
                all_sheets = []
                for result in all_results:
                    if result.get("is_multi_sheet", False):
                        all_sheets.extend(result.get("sheets", []))
                    else:
                        # Single sheet result - convert to sheet format
                        all_sheets.append({
                            "sheet_name": result.get("filename", "GoogleSheet"),
                            "file_id": result.get("file_id"),
                            "rows": result.get("rows", 0),
                            "columns": result.get("columns", 0)
                        })
                
                # Calculate total columns like Excel does
                total_columns = max(sheet["columns"] for sheet in all_sheets) if all_sheets else 0
                
                response_data = {
                    "success": True,
                    "session_id": session_id,
                    "file_id": f"google_sheets_merged_{len(all_results)}",  # Combined file_id like Excel
                    "filename": f"Multiple Google Sheets ({len(all_results)} files)",
                    "is_multi_sheet": True,  # Always true for multiple files
                    "total_files": len(all_sheets),  # Match Excel's total_files
                    "sheets": all_sheets,  # Flattened sheets array like Excel
                    "data_info": {
                        "rows": int(total_rows),
                        "columns": total_columns,
                        "memory_usage_mb": 0,  # Google Sheets don't have file size
                        "file_size_mb": 0,
                        "memory_saved_mb": 0
                    },
                    "message": f"✅ Loaded {len(all_results)} Google Sheets with {int(total_sheets)} total sheets and {int(total_rows):,} total rows"
                }
                
                # Convert all numpy types to Python types for JSON serialization
                response_data = convert_numpy_types(response_data)
                
                logger.info(f"✅ Multiple Google Sheets loaded: {len(all_results)} sheets, {total_rows} rows")
                return jsonify(response_data)
            
            # Handle single Google Sheet (existing logic)
            spreadsheet_url = single_url
            
            logger.info(f"📊 Loading Google Sheet from URL: {spreadsheet_url[:50]}... (load_all={load_all_sheets})")
            
            # Load single sheet (the load_google_sheet method now handles multi-sheet automatically)
            result = full_data_agent.load_google_sheet(
                spreadsheet_url,
                session_id,
                sheet_name=sheet_name
            )
            
            if not result["success"]:
                return jsonify(result), 400
            
            # Handle both single-sheet and multi-sheet responses
            if result.get("is_multi_sheet", False):
                # Multi-sheet response
                total_rows = sum(sheet["rows"] for sheet in result["sheets"])
                total_columns = max(sheet["columns"] for sheet in result["sheets"]) if result["sheets"] else 0
                
                response_data = {
                    "success": True,
                    "session_id": session_id,
                    "file_id": result["file_id"],
                    "filename": "Multi-sheet Google Sheet",
                    "is_multi_sheet": True,
                    "sheets": result["sheets"],
                    "total_files": len(result["sheets"]),
                    "data_info": {
                        "total_rows": total_rows,
                        "max_columns": total_columns,
                        "sheets_count": len(result["sheets"])
                    },
                    "message": result["message"]
                }
                
                logger.info(f"✅ Multi-sheet Google Sheet loaded: {len(result['sheets'])} sheets with {total_rows:,} total rows")
            else:
                # Single-sheet response
                response_data = {
                    "success": True,
                    "session_id": session_id,
                    "file_id": result["file_id"],
                    "filename": result["filename"],
                    "total_files": result["total_files"],
                    "data_info": {
                        "rows": result["rows"],
                        "columns": result["columns"],
                        "memory_usage_mb": round(result["memory_usage_mb"], 2)
                    },
                    "message": ""  # Remove predefined message to avoid duplication in chat
                }
                
                logger.info(f"✅ Google Sheet loaded successfully: {result['filename']}")
            
            return jsonify(response_data)
            
        except Exception as e:
            logger.error(f"❌ Error loading Google Sheet: {e}", exc_info=True)
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/upload', methods=['POST'])
    def upload_file():
        """Regular upload endpoint (redirects to full upload for compatibility)"""
        try:
            from werkzeug.utils import secure_filename
            
            if 'file' not in request.files:
                return jsonify({"success": False, "error": "No file provided"}), 400

            file = request.files['file']
            if file.filename == '':
                return jsonify({"success": False, "error": "Empty filename"}), 400

            # Ensure uploads directory exists
            uploads_dir = Path(__file__).parent / "uploads"
            uploads_dir.mkdir(exist_ok=True)

            # Save file
            filename = secure_filename(file.filename)
            file_path = uploads_dir / filename
            file.save(file_path)
            
            # Generate session ID
            session_id = request.form.get('agent_id') or str(uuid.uuid4())
            
            logger.info(f"📁 Processing upload (compatibility mode): {filename} for session {session_id}")
            
            # Load using Full Data Agent
            result = full_data_agent.load_excel_file(str(file_path), session_id)
            
            if not result["success"]:
                return jsonify(result), 400
            
            # Return in format expected by frontend
            return jsonify({
                "success": True,
                "agent_id": session_id,
                "rows": result["rows"],
                "filename": filename
            })
            
        except Exception as e:
            logger.error(f"❌ Error in upload: {e}", exc_info=True)
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/session-files/<session_id>', methods=['GET'])
    def get_session_files(session_id):
        """Get all files for a session"""
        try:
            files = full_data_agent.get_session_files(session_id)
            
            files_info = []
            for file_id, file_data in files.items():
                files_info.append({
                    "file_id": file_id,
                    "filename": file_data["filename"],
                    "rows": len(file_data["df"]),
                    "columns": len(file_data["df"].columns),
                    "memory_usage_mb": round(file_data["memory_usage_mb"], 2),
                    "file_size_mb": round(file_data["file_size_mb"], 2)
                })
            
            return jsonify({
                "success": True,
                "session_id": session_id,
                "total_files": len(files_info),
                "files": files_info
            })
            
        except Exception as e:
            logger.error(f"❌ Error getting session files: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/set-active-file', methods=['POST'])
    def set_active_file():
        """Set the active file for analysis"""
        try:
            data = request.get_json()
            session_id = data.get('session_id')
            file_id = data.get('file_id')
            
            if not session_id or not file_id:
                return jsonify({"success": False, "error": "session_id and file_id required"}), 400
            
            success = full_data_agent.set_active_file(session_id, file_id)
            
            if success:
                return jsonify({
                    "success": True,
                    "message": f"Active file set to {file_id}",
                    "session_id": session_id,
                    "active_file": file_id
                })
            else:
                return jsonify({"success": False, "error": "Invalid session or file ID"}), 400
                
        except Exception as e:
            logger.error(f"❌ Error setting active file: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/chat-full', methods=['POST'])
    def chat_with_full_data():
        """
        Enhanced chat endpoint that uses full DataFrame for analysis
        
        Flow:
        1. Receive user query
        2. Access full DataFrame from memory
        3. LLM generates Python code based on complete data understanding
        4. Execute code on full DataFrame
        5. Return results (data, plots, insights)
        """
        try:
            data = request.json
            if not data:
                return jsonify({"success": False, "error": "No JSON data provided"}), 400
            
            session_id = data.get('session_id')
            message = data.get('message')
            
            if not session_id or not message:
                return jsonify({"success": False, "error": "Missing session_id or message"}), 400
            
            logger.info(f"🤖 Full data chat request: session={session_id}, query='{message[:50]}...'")
            
            # Process query using Full Data Agent
            result = full_data_agent.process_query(session_id, message)
            
            if not result["success"]:
                return jsonify(result), 400
            
            # Format response for frontend
            formatted_response = _format_analysis_response(result, full_data_agent)
            
            response_data = {
                "success": True,
                "query": result["query"],
                "data_shape": result["data_shape"],
                "generated_code": result["generated_code"],
                "explanation": result["explanation"],
                "execution_result": result["result"]
            }
            
            # Handle structured response with URL for hyperlink
            if isinstance(formatted_response, dict):
                response_data["message"] = formatted_response["text"]
                response_data["sheet_url"] = formatted_response.get("sheet_url", "")
                response_data["has_link"] = formatted_response.get("has_link", False)
                response_data["link_text"] = formatted_response.get("link_text", "")
            else:
                response_data["message"] = formatted_response
            
            # Add chart information if available
            if result["result"].get("charts"):
                response_data["charts"] = result["result"]["charts"]
                # Frontend will display charts automatically, no need to add message
            
            # Add DataFrame results if available
            if result["result"].get("dataframes"):
                df_info = []
                for df_name, df_data in result["result"]["dataframes"].items():
                    df_info.append(f"**{df_name}**: {df_data['shape'][0]} rows × {df_data['shape'][1]} columns")
                if df_info:
                    response_data["message"] += f"\n\n📋 Results:\n" + "\n".join(df_info)
            
            logger.info(f"✅ Full data chat successful: session={session_id}")
            return jsonify(response_data)
            
        except Exception as e:
            logger.error(f"❌ Error in full data chat: {e}", exc_info=True)
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/chat', methods=['POST'])
    def chat():
        """Regular chat endpoint (uses full data agent for compatibility)"""
        try:
            data = request.json
            if not data:
                return jsonify({"success": False, "error": "No JSON data provided"}), 400
            
            # Accept both agent_id (from frontend) and session_id
            session_id = data.get('agent_id') or data.get('session_id')
            message = data.get('message')
            
            if not session_id or not message:
                return jsonify({"success": False, "error": "Missing agent_id or message"}), 400
            
            logger.info(f"💬 Chat request: session={session_id}, query='{message[:50]}...'")
            
            # Check if session has data loaded
            if session_id in full_data_agent.session_data:
                # Use full data agent for data analysis
                result = full_data_agent.process_query(session_id, message)
                
                if result["success"]:
                    formatted_response = _format_analysis_response(result, full_data_agent)
                    response_data = {
                        "success": True,
                        "data_shape": result["data_shape"]
                    }
                    
                    # Handle structured response with URL for hyperlink
                    if isinstance(formatted_response, dict):
                        response_data["message"] = formatted_response["text"]
                        response_data["sheet_url"] = formatted_response.get("sheet_url", "")
                        response_data["has_link"] = formatted_response.get("has_link", False)
                        response_data["link_text"] = formatted_response.get("link_text", "")
                    else:
                        response_data["message"] = formatted_response
                    
                    return jsonify(response_data)
                else:
                    return jsonify(result), 400
            else:
                # Use simple chat for basic conversations
                result = simple_chat.process_chat_message(session_id, message)
                return jsonify(result)
            
        except Exception as e:
            logger.error(f"❌ Error in chat: {e}", exc_info=True)
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/chat-stream', methods=['POST'])
    def chat_stream():
        """Stream chat response chunks (SSE) - Enhanced version with error handling"""
        try:
            data = request.json
            if not data:
                return jsonify({"success": False, "error": "No JSON data provided"}), 400
            
            session_id = data.get('agent_id') or data.get('session_id')
            message = data.get('message')
            
            if not session_id or not message:
                return jsonify({"success": False, "error": "Missing agent_id or message"}), 400

            def generate():
                """Generator with comprehensive error handling and timeout protection"""
                try:
                    logger.info(f"/api/chat-stream ▶ session={session_id} msg_len={len(message)} preview={message[:60]!r}")
                    
                    # Send initial keep-alive to establish connection
                    yield ": keep-alive\n\n"
                    
                    # Check if session has data loaded in full data agent
                    if session_id in full_data_agent.session_data:
                        logger.info(f"/api/chat-stream 🤖 Using full data agent for session={session_id}")
                        
                        try:
                            # Execute full data processing with comprehensive error handling
                            import signal
                            from contextlib import contextmanager
                            
                            @contextmanager
                            def timeout_handler(seconds):
                                """Timeout context manager"""
                                def timeout_signal(signum, frame):
                                    raise TimeoutError(f"Operation timed out after {seconds} seconds")
                                
                                # Set timeout only on Unix systems
                                if hasattr(signal, 'SIGALRM'):
                                    old_handler = signal.signal(signal.SIGALRM, timeout_signal)
                                    signal.alarm(seconds)
                                    try:
                                        yield
                                    finally:
                                        signal.alarm(0)
                                        signal.signal(signal.SIGALRM, old_handler)
                                else:
                                    # Windows - no timeout, just execute
                                    yield
                            
                            # Execute with timeout protection (60 seconds)
                            try:
                                with timeout_handler(60):
                                    result = full_data_agent.process_query(session_id, message)
                            except TimeoutError as te:
                                logger.error(f"/api/chat-stream ⏱️ Timeout: {te}")
                                yield f"data: ⏱️ Request timed out. Please try a simpler query or smaller dataset.\n\n"
                                yield "data: [DONE]\n\n"
                                return
                            
                            if not result["success"]:
                                error_msg = result.get('error', 'Unknown error occurred')
                                logger.error(f"/api/chat-stream ❌ Query failed: {error_msg}")
                                yield f"data: ❌ Error: {error_msg}\n\n"
                                yield "data: [DONE]\n\n"
                                return
                            
                            # Handle different result structures
                            result_data = result.get("result", result)  # For backward compatibility
                            
                            # Send chart data first if available (as JSON event)
                            if isinstance(result_data, dict) and result_data.get("charts"):
                                chart_data = json.dumps({"charts": result_data["charts"]})
                                yield f"data: [CHARTS]{chart_data}\n\n"
                                logger.info(f"/api/chat-stream 📊 Sent chart data: {len(result_data['charts'])} chart(s)")
                            
                            # Send Excel file URLs if available (as JSON event)
                            if isinstance(result_data, dict) and result_data.get("excel_files"):
                                excel_data = json.dumps({"excel_files": result_data["excel_files"]})
                                yield f"data: [EXCEL]{excel_data}\n\n"
                                logger.info(f"/api/chat-stream 📊 Sent Excel URLs: {result_data['excel_files']}")
                            
                            # Format the response
                            formatted_response = _format_analysis_response(result, full_data_agent)
                            
                            # Extract text for streaming (handle structured response)
                            response_text = _extract_response_text(formatted_response)
                            
                            # Send URL data if available (as JSON event)
                            if isinstance(formatted_response, dict) and formatted_response.get("has_link"):
                                url_data = json.dumps({
                                    "sheet_url": formatted_response["sheet_url"],
                                    "link_text": formatted_response["link_text"]
                                })
                                yield f"data: [URL]{url_data}\n\n"
                                logger.info(f"/api/chat-stream 🔗 Sent URL data: {formatted_response['sheet_url']}")
                            
                            # Stream word by word with error handling
                            words = response_text.split()
                            for i, word in enumerate(words):
                                try:
                                    if word.strip():
                                        token_to_send = word if i == 0 else " " + word
                                        yield f"data: {token_to_send}\n\n"
                                        time.sleep(0.05)  # Reduced delay for faster streaming
                                        
                                        # Send keep-alive every 50 words
                                        if i % 50 == 0 and i > 0:
                                            yield ": keep-alive\n\n"
                                except GeneratorExit:
                                    logger.warning(f"/api/chat-stream ⚠️ Client disconnected during streaming")
                                    return
                                except Exception as stream_err:
                                    logger.error(f"/api/chat-stream ❌ Streaming error: {stream_err}")
                                    break
                            
                            yield "data: [DONE]\n\n"
                            logger.info(f"/api/chat-stream ✅ Full data response streamed")
                            return
                            
                        except TimeoutError as te:
                            logger.error(f"/api/chat-stream ⏱️ Timeout error: {te}")
                            yield f"data: ⏱️ Request timed out. Please try again with a simpler query.\n\n"
                            yield "data: [DONE]\n\n"
                            return
                        except MemoryError as me:
                            logger.error(f"/api/chat-stream 💾 Memory error: {me}")
                            yield f"data: 💾 Out of memory. Please try with a smaller dataset.\n\n"
                            yield "data: [DONE]\n\n"
                            return
                        except Exception as data_error:
                            logger.error(f"/api/chat-stream ❌ Data processing error: {data_error}", exc_info=True)
                            error_message = f"❌ Sorry, I encountered an error: {str(data_error)[:200]}"
                            yield f"data: {error_message}\n\n"
                            yield "data: [DONE]\n\n"
                            return
                    
                    else:
                        # Use simple chat system for basic conversations
                        logger.info(f"/api/chat-stream 💬 Using simple chat for session={session_id}")
                        
                        # Use simple chat service properly (non-streaming for now to avoid duplicates)
                        try:
                            # Ensure session exists before processing message
                            if session_id not in simple_chat.sessions:
                                # Create session manually with the specific session_id
                                simple_chat.sessions[session_id] = {
                                    "session_id": session_id,
                                    "user_id": None,
                                    "created_at": datetime.now().isoformat(),
                                    "conversation_history": [],
                                    "memory_summary": "",
                                    "preferences": {},
                                }
                                simple_chat._persist_session_to_disk(session_id)
                                logger.info(f"Created new session: {session_id}")
                            
                            chat_response = simple_chat.process_chat_message(session_id, message)
                            
                            if chat_response["success"]:
                                ai_message = chat_response["message"]
                                
                                # Stream the response word by word
                                words = ai_message.split()
                                total_words_sent = 0
                                
                                for word in words:
                                    if word.strip():
                                        token_to_send = word if total_words_sent == 0 else " " + word
                                        total_words_sent += 1
                                        yield f"data: {token_to_send}\n\n"
                                        time.sleep(0.05)  # Reduced delay
                                        
                                        # Keep-alive every 50 words
                                        if total_words_sent % 50 == 0:
                                            yield ": keep-alive\n\n"
                            else:
                                yield f"data: Error: {chat_response.get('error', 'Unknown error')}\n\n"
                            
                        except Exception as e:
                            # Fallback error handling
                            logger.error(f"/api/chat-stream ❌ Simple chat failed: {e}")
                            yield f"data: I apologize, but I'm experiencing technical difficulties. Please try again.\n\n"
                        
                        yield "data: [DONE]\n\n"
                        logger.info(f"/api/chat-stream ✅ Simple chat response streamed")
                        
                except GeneratorExit:
                    logger.warning(f"/api/chat-stream ⚠️ Generator closed by client for session={session_id}")
                    return
                except Exception as eg:
                    logger.error(f"/api/chat-stream ❌ Critical error session={session_id}: {eg}", exc_info=True)
                    try:
                        yield f"data: ❌ Critical error: {str(eg)[:100]}\n\n"
                        yield "data: [DONE]\n\n"
                    except:
                        pass  # Connection already closed
                    yield f"data: {json.dumps({'error': str(eg)})}\n\n"

            headers = {
                'Content-Type': 'text/event-stream',
                'Cache-Control': 'no-cache, no-transform',
                'Connection': 'keep-alive',
                'X-Accel-Buffering': 'no',
            }
            return Response(stream_with_context(generate()), headers=headers)
            
        except Exception as e:
            logger.error(f"❌ Error streaming chat: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/session-info/<session_id>', methods=['GET'])
    def get_session_info(session_id):
        """Get detailed information about a loaded session"""
        try:
            result = full_data_agent.get_session_info(session_id)
            return jsonify(result)
        except Exception as e:
            logger.error(f"❌ Error getting session info: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/clear-session/<session_id>', methods=['DELETE'])
    def clear_session(session_id):
        """Clear session data from memory"""
        try:
            success = full_data_agent.clear_session(session_id)
            if success:
                return jsonify({"success": True, "message": f"Session {session_id} cleared"})
            else:
                return jsonify({"success": False, "error": "Failed to clear session"}), 500
        except Exception as e:
            logger.error(f"❌ Error clearing session: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/chat-simple', methods=['POST'])
    def chat_simple():
        """Simple chat endpoint for basic conversations without data"""
        try:
            data = request.json
            if not data:
                return jsonify({"success": False, "error": "No JSON data provided"}), 400
            
            session_id = data.get('session_id') or str(uuid.uuid4())
            message = data.get('message')
            
            if not message:
                return jsonify({"success": False, "error": "Missing message"}), 400
            
            # Use simple chat system
            result = simple_chat.process_chat_message(session_id, message)
            return jsonify(result)
            
        except Exception as e:
            logger.error(f"❌ Error in simple chat: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/download/<filename>', methods=['GET'])
    def download_file(filename):
        """Download generated files (charts, Excel files)"""
        try:
            # Security: prevent directory traversal
            if '..' in filename or '/' in filename or '\\' in filename:
                return jsonify({"success": False, "error": "Invalid filename"}), 400
            
            file_path = full_data_agent.outputs_dir / filename
            
            if not file_path.exists():
                logger.warning(f"❌ File not found: {file_path}")
                return jsonify({"success": False, "error": f"File not found: {filename}"}), 404
            
            # Determine mimetype and clean download name
            if filename.endswith('.png'):
                mimetype = 'image/png'
                download_name = filename
            elif filename.endswith('.jpg') or filename.endswith('.jpeg'):
                mimetype = 'image/jpeg'
                download_name = filename
            elif filename.endswith('.xlsx'):
                mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                # Extract clean name without timestamp for download
                # e.g., app_1762407862.xlsx -> app.xlsx
                import re
                timestamp_match = re.match(r'^(.+)_(\d{10,})\.xlsx$', filename)
                if timestamp_match:
                    download_name = f"{timestamp_match.group(1)}.xlsx"
                else:
                    download_name = filename
            elif filename.endswith('.csv'):
                mimetype = 'text/csv'
                download_name = filename
            else:
                mimetype = 'application/octet-stream'
                download_name = filename
            
            logger.info(f"📥 Serving download: {filename} as {download_name}")
            return send_file(file_path, mimetype=mimetype, as_attachment=filename.endswith('.xlsx'), download_name=download_name)
            
        except Exception as e:
            logger.error(f"❌ Error downloading file: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/session', methods=['POST'])
    def create_session():
        """Create a new session (compatible with frontend)"""
        try:
            # Generate new session ID
            session_id = str(uuid.uuid4())
            
            # Create session in simple chat system
            simple_chat.create_session(session_id)
            
            logger.info(f"✅ Created new session: {session_id}")
            
            return jsonify({
                "success": True,
                "session_id": session_id,
                "agent_id": session_id,  # For compatibility
                "created_at": datetime.now().isoformat()
            })
            
        except Exception as e:
            logger.error(f"❌ Error creating session: {e}", exc_info=True)
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/sessions', methods=['GET'])
    def list_sessions():
        """List all sessions (compatible with frontend)"""
        try:
            sessions = []
            
            # Get sessions from full data agent
            for session_id in full_data_agent.session_data.keys():
                session_data = full_data_agent.session_data[session_id]
                conversation_count = len(session_data.get("conversation_history", []))
                sessions.append({
                    "id": session_id,
                    "created_at": datetime.now().isoformat(),  # Placeholder
                    "message_count": conversation_count,
                    "filename": f"Data Session {session_id[:8]}"
                })
            
            # Sort by created_at descending (newest first)
            sessions.sort(key=lambda x: x.get("created_at", ""), reverse=True)
            
            return jsonify({
                "success": True,
                "sessions": sessions
            })
            
        except Exception as e:
            logger.error(f"❌ Error listing sessions: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/session/<session_id>/history', methods=['GET'])
    def get_session_history(session_id):
        """Get session history (compatible with frontend)"""
        try:
            if session_id not in full_data_agent.session_data:
                return jsonify({"success": False, "error": "Session not found"}), 404
            
            # Get conversation history
            session_data = full_data_agent.session_data[session_id]
            history = session_data.get("conversation_history", [])
            
            # Format for frontend compatibility
            session_data = {
                "session_id": session_id,
                "created_at": datetime.now().isoformat(),
                "conversation_history": history,
                "data_info": {
                    "total_files": len(session_data["files"]),
                    "active_file": session_data.get("active_file"),
                    "files": list(session_data["files"].keys())
                }
            }
            
            return jsonify({"success": True, "session": session_data})
            
        except Exception as e:
            logger.error(f"❌ Error getting session history: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/new-chat', methods=['POST'])
    def new_chat():
        """Create new chat session (compatible with frontend)"""
        try:
            data = request.get_json(silent=True) or {}
            session_id = data.get('session_id') or str(uuid.uuid4())
            
            # Clear existing session if it exists
            if session_id in full_data_agent.session_data:
                full_data_agent.clear_session(session_id)
            
            return jsonify({
                "success": True, 
                "session_id": session_id,
                "message": "New chat session created"
            })
            
        except Exception as e:
            logger.error(f"❌ Error creating new chat: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/health', methods=['GET'])
    def api_health_check():
        """Health check endpoint"""
        try:
            return jsonify({
                "success": True,
                "status": "healthy",
                "timestamp": datetime.now().isoformat(),
                "version": "1.0.0",
                "services": {
                    "full_data_agent": "running",
                    "simple_chat": "running"
                }
            })
        except Exception as e:
            return jsonify({
                "success": False,
                "status": "unhealthy",
                "error": str(e)
            }), 500
    
    @app.route('/api/memory-stats', methods=['GET'])
    def memory_stats():
        """Get memory usage statistics"""
        try:
            stats = {
                "loaded_sessions": len(full_data_agent.session_data),
                "total_memory_mb": 0,
                "sessions": {}
            }
            
            total_memory = 0
            for session_id, session_data in full_data_agent.session_data.items():
                session_memory = 0
                files_info = []
                
                for file_id, file_data in session_data["files"].items():
                    df = file_data["df"]
                    file_memory = df.memory_usage(deep=True).sum() / 1024 / 1024
                    session_memory += file_memory
                    files_info.append({
                        "file_id": file_id,
                        "filename": file_data["filename"],
                        "shape": f"{len(df)} × {len(df.columns)}",
                        "memory_mb": file_memory
                    })
                
                total_memory += session_memory
                stats["sessions"][session_id] = {
                    "total_files": len(session_data["files"]),
                    "memory_mb": session_memory,
                    "conversation_count": len(session_data.get("conversation_history", [])),
                    "files": files_info
                }
            
            stats["total_memory_mb"] = total_memory
            
            return jsonify({"success": True, "stats": stats})
            
        except Exception as e:
            logger.error(f"❌ Error getting memory stats: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/list-excel-files', methods=['GET'])
    def list_excel_files():
        """List all generated Excel files"""
        try:
            excel_files = []
            outputs_dir = full_data_agent.outputs_dir
            
            if outputs_dir.exists():
                for file_path in outputs_dir.glob("*.xlsx"):
                    stat = file_path.stat()
                    excel_files.append({
                        "name": file_path.name,
                        "size": stat.st_size,
                        "modified": stat.st_mtime
                    })
                
                # Sort by modification time (newest first)
                excel_files.sort(key=lambda x: x['modified'], reverse=True)
            
            return jsonify({"success": True, "files": excel_files})
            
        except Exception as e:
            logger.error(f"❌ Error listing Excel files: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/downloads', methods=['GET'])
    def downloads_page():
        """Serve the downloads page"""
        try:
            from flask import render_template
            return render_template('downloads.html')
        except Exception as e:
            logger.error(f"❌ Error serving downloads page: {e}")
            return f"Error: {e}", 500
    
    @app.route('/api/preview/<filename>', methods=['GET'])
    def preview_excel(filename):
        """Preview Excel file content without downloading"""
        try:
            logger.info(f"📋 Preview request for: {filename}")
            
            # Try multiple directories to find the file
            outputs_dir = full_data_agent.outputs_dir
            uploads_dir = Path(__file__).parent / "uploads"
            temp_dir = Path(__file__).parent / "temp"
            
            possible_paths = [
                outputs_dir / filename,
                uploads_dir / filename,
                temp_dir / filename,
            ]
            
            file_path = None
            for path in possible_paths:
                logger.info(f"🔍 Checking path: {path}")
                if path.exists():
                    file_path = path
                    logger.info(f"✅ Found file at: {path}")
                    break
            
            if file_path is None:
                logger.warning(f"❌ Preview file not found in any directory: {filename}")
                logger.warning(f"📁 Checked directories:")
                for i, path in enumerate(possible_paths):
                    logger.warning(f"   {i+1}. {path.parent} (exists: {path.parent.exists()})")
                
                # List files in outputs directory for debugging
                try:
                    if outputs_dir.exists():
                        files = list(outputs_dir.glob("*.xlsx"))
                        logger.warning(f"📄 Files in outputs directory: {[f.name for f in files]}")
                except Exception as e:
                    logger.warning(f"Error listing files: {e}")
                
                return jsonify({"success": False, "error": "File not found"}), 404
            
            if not file_path.suffix.lower() in ['.xlsx', '.xls']:
                logger.warning(f"❌ Invalid file type for preview: {filename}")
                return jsonify({"success": False, "error": "Only Excel files can be previewed"}), 400
            
            # Read Excel file and convert to JSON
            # Read all sheets with pandas for data
            excel_data = pd.read_excel(file_path, sheet_name=None)  # Read all rows for complete preview
            
            # Check if any sheet is very large and warn user
            max_preview_rows = 1000  # Reasonable limit for web display
            large_sheets = []
            for sheet_name, df in excel_data.items():
                if len(df) > max_preview_rows:
                    large_sheets.append(f"{sheet_name} ({len(df)} rows)")
            
            if large_sheets:
                logger.info(f"⚠️ Large sheets detected for preview: {', '.join(large_sheets)}")
                logger.info(f"📊 Showing all rows but performance may be affected for very large datasets")
            
            # Also read with openpyxl to get formatting information
            from openpyxl import load_workbook
            try:
                wb = load_workbook(file_path, data_only=False)
                formatting_data = {}
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    sheet_formatting = {}
                    # Read cell formatting for all rows
                    for row_idx in range(1, ws.max_row + 1):
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                                cell_key = f"{row_idx}_{col_idx}"
                                rgb = cell.fill.start_color.rgb
                                if rgb and rgb != "00000000":  # Not default/transparent
                                    sheet_formatting[cell_key] = {
                                        "background_color": f"#{rgb[2:8]}" if len(rgb) == 8 else f"#{rgb}"
                                    }
                    formatting_data[sheet_name] = sheet_formatting
            except Exception as e:
                logger.warning(f"Could not read formatting data: {e}")
                formatting_data = {}
            
            preview_data = {
                "filename": filename,
                "sheets": []
            }
            
            for sheet_name, df in excel_data.items():
                # Clean the DataFrame - replace NaN with None (which becomes null in JSON)
                df_clean = df.fillna('')  # Replace NaN with empty string for better display - show all rows
                
                # Convert DataFrame to records (list of dicts)
                records = df_clean.to_dict('records')
                
                # Clean the records to ensure JSON serialization and add formatting
                clean_records = []
                sheet_formatting = formatting_data.get(sheet_name, {})
                
                for row_idx, record in enumerate(records):
                    clean_record = {}
                    for col_idx, (key, value) in enumerate(record.items()):
                        # Handle various problematic values
                        if pd.isna(value) or value is None:
                            clean_record[key] = None
                        elif isinstance(value, (float, int)) and (pd.isna(value) or value != value):  # NaN check
                            clean_record[key] = None
                        else:
                            # Convert to string to ensure JSON serialization
                            clean_record[key] = str(value) if value is not None else None
                        
                        # Add formatting information if available
                        # Note: row_idx + 2 because openpyxl is 1-indexed and we skip header row
                        cell_key = f"{row_idx + 2}_{col_idx + 1}"
                        if cell_key in sheet_formatting:
                            if f"{key}_formatting" not in clean_record:
                                clean_record[f"{key}_formatting"] = sheet_formatting[cell_key]
                    
                    clean_records.append(clean_record)
                
                # Get column info
                columns = [{"name": col, "type": str(df[col].dtype)} for col in df.columns]
                
                sheet_info = {
                    "name": sheet_name,
                    "rows": len(df),
                    "columns": len(df.columns),
                    "preview_rows": len(clean_records),
                    "column_info": columns,
                    "data": clean_records
                }
                
                preview_data["sheets"].append(sheet_info)
            
            logger.info(f"✅ Preview generated for {filename}: {len(preview_data['sheets'])} sheets")
            return jsonify({"success": True, "preview": preview_data})
            
        except Exception as e:
            logger.error(f"❌ Error previewing Excel file {filename}: {e}", exc_info=True)
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route('/api/debug/files', methods=['GET'])
    def debug_files():
        """Debug endpoint to list all Excel files in various directories"""
        try:
            result = {
                "directories": {},
                "requested_file": request.args.get('filename', 'Not specified')
            }
            
            # Check multiple directories
            directories_to_check = {
                "outputs": full_data_agent.outputs_dir,
                "uploads": Path(__file__).parent / "uploads",
                "temp": Path(__file__).parent / "temp",
                "backend": Path(__file__).parent,
            }
            
            for dir_name, dir_path in directories_to_check.items():
                result["directories"][dir_name] = {
                    "path": str(dir_path),
                    "exists": dir_path.exists(),
                    "files": []
                }
                
                if dir_path.exists():
                    try:
                        excel_files = list(dir_path.glob("*.xlsx")) + list(dir_path.glob("*.xls"))
                        result["directories"][dir_name]["files"] = [f.name for f in excel_files]
                    except Exception as e:
                        result["directories"][dir_name]["error"] = str(e)
            
            return jsonify(result)
            
        except Exception as e:
            logger.error(f"❌ Error in debug files: {e}")
            return jsonify({"error": str(e)}), 500
    
    return app

if __name__ == '__main__':
    import sys
    if sys.platform == 'win32':
        sys.stdout.reconfigure(encoding='utf-8')
    
    try:
        app = create_enhanced_app()
        logger.info("🌐 Server running on http://127.0.0.1:9000")
        logger.info("🔧 Debug mode: Disabled (for Streamlit compatibility)")
        logger.info("📡 CORS: Enabled for all origins")
        logger.info("🚀 Ready to accept connections!")
        
        app.run(
            host='127.0.0.1', 
            port=9000, 
            debug=False,  # Disable debug mode to avoid signal conflicts
            threaded=True,
            use_reloader=False,  # Disable reloader to avoid signal conflicts
        )
    except KeyboardInterrupt:
        logger.info("🛑 Server stopped by user")
    except Exception as e:
        logger.error(f"❌ Server startup failed: {e}")
        logger.error("🔍 Please check:")
        logger.error("   - Port 9000 is not in use by another application")
        logger.error("   - Python dependencies are installed correctly")
        logger.error("   - API key is configured in .env file")
        raise
