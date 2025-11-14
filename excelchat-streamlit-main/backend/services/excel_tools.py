"""
Excel Tools for Agentic Workflow
Provides tools for chart generation, Excel modification, and data analysis
"""

import os
import json
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import base64
from io import BytesIO

logger = logging.getLogger(__name__)

# Set style for better-looking charts
sns.set_style("whitegrid")
plt.rcParams['figure.figsize'] = (10, 6)
plt.rcParams['font.size'] = 10


class ExcelTools:
    """Tools for Excel data analysis, visualization, and modification"""
    
    def __init__(self, uploads_dir: str = "backend/uploads", outputs_dir: str = "backend/outputs"):
        # Resolve paths relative to the backend directory
        backend_dir = Path(__file__).resolve().parent.parent
        
        # If paths are relative, make them relative to backend dir
        if not Path(uploads_dir).is_absolute():
            self.uploads_dir = backend_dir / uploads_dir.replace("backend/", "").replace("backend\\", "")
        else:
            self.uploads_dir = Path(uploads_dir)
            
        if not Path(outputs_dir).is_absolute():
            self.outputs_dir = backend_dir / outputs_dir.replace("backend/", "").replace("backend\\", "")
        else:
            self.outputs_dir = Path(outputs_dir)
        
        # Create outputs directory if it doesn't exist
        self.outputs_dir.mkdir(exist_ok=True, parents=True)
        
        logger.info(f"📁 ExcelTools initialized - uploads: {self.uploads_dir}, outputs: {self.outputs_dir}")
        
    def get_tool_definitions(self) -> List[Dict[str, Any]]:
        """Return tool definitions for Gemini function calling"""
        return [
            {
                "name": "generate_chart",
                "description": "Generate a chart/visualization from Excel data. Supports bar, line, pie, scatter, histogram charts.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "filename": {
                            "type": "string",
                            "description": "Name of the Excel file to analyze"
                        },
                        "chart_type": {
                            "type": "string",
                            "enum": ["bar", "line", "pie", "scatter", "histogram", "box"],
                            "description": "Type of chart to generate"
                        },
                        "x_column": {
                            "type": "string",
                            "description": "Column name for X-axis (or categories for pie chart)"
                        },
                        "y_column": {
                            "type": "string",
                            "description": "Column name for Y-axis (optional for some chart types)"
                        },
                        "title": {
                            "type": "string",
                            "description": "Title for the chart"
                        },
                        "aggregate": {
                            "type": "string",
                            "enum": ["count", "sum", "mean", "median", "min", "max"],
                            "description": "Aggregation method if needed (default: count)"
                        }
                    },
                    "required": ["filename", "chart_type", "x_column"]
                }
            },
            {
                "name": "modify_excel",
                "description": "Modify an Excel file by highlighting rows, changing cell colors, or formatting data based on conditions.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "filename": {
                            "type": "string",
                            "description": "Name of the Excel file to modify"
                        },
                        "operation": {
                            "type": "string",
                            "enum": ["highlight_rows", "highlight_cells", "filter_data"],
                            "description": "Type of modification to perform"
                        },
                        "column": {
                            "type": "string",
                            "description": "Column name to check for condition"
                        },
                        "condition_value": {
                            "type": "string",
                            "description": "Value to match for highlighting/filtering"
                        },
                        "color": {
                            "type": "string",
                            "enum": ["green", "red", "yellow", "blue", "orange"],
                            "description": "Color for highlighting (default: green)"
                        },
                        "output_filename": {
                            "type": "string",
                            "description": "Name for the output file"
                        }
                    },
                    "required": ["filename", "operation", "column", "condition_value"]
                }
            },
            {
                "name": "analyze_data",
                "description": "Perform data analysis on Excel data: statistics, filtering, grouping, sorting.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "filename": {
                            "type": "string",
                            "description": "Name of the Excel file to analyze"
                        },
                        "operation": {
                            "type": "string",
                            "enum": ["statistics", "filter", "group_by", "sort", "unique_values"],
                            "description": "Type of analysis to perform"
                        },
                        "column": {
                            "type": "string",
                            "description": "Column name for the operation"
                        },
                        "filter_value": {
                            "type": "string",
                            "description": "Value to filter by (if operation is filter)"
                        },
                        "aggregate_function": {
                            "type": "string",
                            "enum": ["count", "sum", "mean", "median", "min", "max"],
                            "description": "Aggregation function for group_by"
                        }
                    },
                    "required": ["filename", "operation"]
                }
            }
        ]
    
    def generate_chart(self, filename: str, chart_type: str, x_column: str, 
                      y_column: Optional[str] = None, title: Optional[str] = None,
                      aggregate: str = "count") -> Dict[str, Any]:
        """Generate a chart from Excel data"""
        try:
            # Load the Excel file
            file_path = self.uploads_dir / filename
            logger.info(f"📂 Looking for file at: {file_path}")
            
            if not file_path.exists():
                # List available files for debugging
                try:
                    available_files = list(self.uploads_dir.glob("*.xlsx")) + list(self.uploads_dir.glob("*.xls"))
                    logger.warning(f"❌ File not found: {file_path}")
                    logger.info(f"📋 Available files in {self.uploads_dir}: {[f.name for f in available_files]}")
                except Exception as e:
                    logger.error(f"Error listing files: {e}")
                
                return {"success": False, "error": f"File {filename} not found at {file_path}"}
            
            df = pd.read_excel(file_path)
            
            # Validate columns
            if x_column not in df.columns:
                return {"success": False, "error": f"Column '{x_column}' not found in file"}
            
            if y_column and y_column not in df.columns:
                return {"success": False, "error": f"Column '{y_column}' not found in file"}
            
            # Create figure
            plt.figure(figsize=(10, 6))
            
            # Generate chart based on type
            if chart_type == "bar":
                if y_column:
                    # Direct bar chart
                    df_plot = df.groupby(x_column)[y_column].agg(aggregate)
                else:
                    # Count occurrences
                    df_plot = df[x_column].value_counts()
                df_plot.plot(kind='bar', color='steelblue')
                plt.ylabel(y_column if y_column else 'Count')
                
            elif chart_type == "line":
                if y_column:
                    df.plot(x=x_column, y=y_column, kind='line', marker='o')
                else:
                    df[x_column].value_counts().sort_index().plot(kind='line', marker='o')
                    
            elif chart_type == "pie":
                if y_column:
                    df_plot = df.groupby(x_column)[y_column].sum()
                else:
                    df_plot = df[x_column].value_counts()
                
                # Improved pie chart with better label handling
                fig, ax = plt.subplots(figsize=(10, 8))
                
                # Calculate percentages
                total = df_plot.sum()
                percentages = [(v/total)*100 for v in df_plot.values]
                
                # Use colors from seaborn palette
                colors = sns.color_palette("husl", len(df_plot))
                
                # Create pie chart with improved settings
                wedges, texts, autotexts = ax.pie(
                    df_plot.values,
                    labels=df_plot.index,
                    autopct='%1.1f%%',
                    startangle=90,
                    colors=colors,
                    textprops={'fontsize': 10},
                    pctdistance=0.85,  # Distance of percentage labels from center
                    labeldistance=1.1,  # Distance of text labels from center
                )
                
                # Improve text visibility
                for autotext in autotexts:
                    autotext.set_color('white')
                    autotext.set_fontweight('bold')
                    autotext.set_fontsize(9)
                
                # If too many categories or small slices, use legend instead
                if len(df_plot) > 8 or any(p < 2.0 for p in percentages):
                    # Remove labels from pie and use legend
                    for text in texts:
                        text.set_text('')
                    
                    # Create legend outside the pie
                    ax.legend(
                        wedges,
                        [f'{label}: {p:.1f}%' for label, p in zip(df_plot.index, percentages)],
                        title=x_column,
                        loc="center left",
                        bbox_to_anchor=(1, 0, 0.5, 1),
                        fontsize=9
                    )
                else:
                    # Improve label positioning for readability
                    for i, (text, percentage) in enumerate(zip(texts, percentages)):
                        if percentage < 3.0:  # Small slices - adjust label position
                            # Move labels for small slices further out
                            text.set_fontsize(9)
                            # Use leader lines for very small slices
                            if percentage < 1.5:
                                text.set_visible(False)  # Hide if too small
                                # Add annotation instead
                                angle = (wedges[i].theta2 + wedges[i].theta1) / 2
                                x = 1.2 * np.cos(np.radians(angle))
                                y = 1.2 * np.sin(np.radians(angle))
                                ax.annotate(
                                    f'{df_plot.index[i]}\n{percentage:.1f}%',
                                    xy=(np.cos(np.radians(angle)), np.sin(np.radians(angle))),
                                    xytext=(x, y),
                                    arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0'),
                                    fontsize=8,
                                    ha='center'
                                )
                
                ax.set_aspect('equal')
                
                # Set title for pie chart
                if title:
                    ax.set_title(title, fontsize=14, fontweight='bold', pad=20)
                else:
                    ax.set_title(f'{chart_type.capitalize()} Chart: {x_column}', fontsize=14, fontweight='bold', pad=20)
                
                # Adjust layout for pie chart
                plt.tight_layout()
                
            elif chart_type == "scatter":
                if not y_column:
                    return {"success": False, "error": "Scatter plot requires both x_column and y_column"}
                plt.scatter(df[x_column], df[y_column], alpha=0.6, c='steelblue')
                plt.xlabel(x_column)
                plt.ylabel(y_column)
                
            elif chart_type == "histogram":
                df[x_column].hist(bins=20, color='steelblue', edgecolor='black')
                plt.ylabel('Frequency')
                
            elif chart_type == "box":
                if y_column:
                    df.boxplot(column=y_column, by=x_column)
                else:
                    df.boxplot(column=x_column)
                plt.suptitle('')
            
            # Set title and labels (skip for pie charts - handled separately)
            if chart_type != "pie":
                if title:
                    plt.title(title, fontsize=14, fontweight='bold')
                else:
                    plt.title(f'{chart_type.capitalize()} Chart: {x_column}', fontsize=14, fontweight='bold')
                
                plt.xlabel(x_column)
                plt.xticks(rotation=45, ha='right')
                plt.tight_layout()
            
            # Save to file
            output_filename = f"chart_{filename.rsplit('.', 1)[0]}_{chart_type}.png"
            output_path = self.outputs_dir / output_filename
            plt.savefig(output_path, dpi=150, bbox_inches='tight')
            plt.close()
            
            logger.info(f"✅ Generated {chart_type} chart: {output_filename}")
            
            return {
                "success": True,
                "message": f"Chart generated successfully: {chart_type} chart for {x_column}",
                "filename": output_filename,
                "download_url": f"/api/download/{output_filename}"
            }
            
        except Exception as e:
            logger.error(f"Error generating chart: {e}", exc_info=True)
            return {"success": False, "error": str(e)}
    
    def modify_excel(self, filename: str, operation: str, column: str, 
                    condition_value: str, color: str = "green",
                    output_filename: Optional[str] = None) -> Dict[str, Any]:
        """Modify Excel file with highlighting or filtering"""
        try:
            # Load the Excel file
            file_path = self.uploads_dir / filename
            if not file_path.exists():
                return {"success": False, "error": f"File {filename} not found"}
            
            df = pd.read_excel(file_path)
            
            # Validate column
            if column not in df.columns:
                return {"success": False, "error": f"Column '{column}' not found in file"}
            
            # Color mapping
            color_map = {
                "green": "00FF00",
                "red": "FF0000",
                "yellow": "FFFF00",
                "blue": "0000FF",
                "orange": "FFA500"
            }
            hex_color = color_map.get(color.lower(), "00FF00")
            
            # Generate output filename
            if not output_filename:
                output_filename = f"modified_{filename}"
            output_path = self.outputs_dir / output_filename
            
            if operation == "highlight_rows":
                # Load workbook with openpyxl
                wb = load_workbook(file_path)
                ws = wb.active
                
                # Find column index
                headers = [cell.value for cell in ws[1]]
                try:
                    col_idx = headers.index(column) + 1
                except ValueError:
                    return {"success": False, "error": f"Column '{column}' not found"}
                
                # Highlight matching rows
                fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                matched_count = 0
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    cell_value = str(row[col_idx - 1].value).strip().lower()
                    condition = str(condition_value).strip().lower()
                    
                    if condition in cell_value or cell_value == condition:
                        for cell in row:
                            cell.fill = fill
                        matched_count += 1
                
                wb.save(output_path)
                logger.info(f"✅ Highlighted {matched_count} rows in {output_filename}")
                
                return {
                    "success": True,
                    "message": f"Highlighted {matched_count} rows where {column} matches '{condition_value}' in {color}",
                    "filename": output_filename,
                    "download_url": f"/api/download/{output_filename}",
                    "matched_count": matched_count
                }
                
            elif operation == "filter_data":
                # Filter data and save to new Excel
                condition = str(condition_value).strip().lower()
                filtered_df = df[df[column].astype(str).str.lower().str.contains(condition, na=False)]
                
                filtered_df.to_excel(output_path, index=False)
                logger.info(f"✅ Filtered {len(filtered_df)} rows to {output_filename}")
                
                return {
                    "success": True,
                    "message": f"Filtered {len(filtered_df)} rows where {column} contains '{condition_value}'",
                    "filename": output_filename,
                    "download_url": f"/api/download/{output_filename}",
                    "row_count": len(filtered_df)
                }
            
            elif operation == "highlight_cells":
                # Highlight specific cells
                wb = load_workbook(file_path)
                ws = wb.active
                
                headers = [cell.value for cell in ws[1]]
                try:
                    col_idx = headers.index(column) + 1
                except ValueError:
                    return {"success": False, "error": f"Column '{column}' not found"}
                
                fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                matched_count = 0
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    cell = row[col_idx - 1]
                    cell_value = str(cell.value).strip().lower()
                    condition = str(condition_value).strip().lower()
                    
                    if condition in cell_value or cell_value == condition:
                        cell.fill = fill
                        matched_count += 1
                
                wb.save(output_path)
                logger.info(f"✅ Highlighted {matched_count} cells in {output_filename}")
                
                return {
                    "success": True,
                    "message": f"Highlighted {matched_count} cells in column {column} matching '{condition_value}' in {color}",
                    "filename": output_filename,
                    "download_url": f"/api/download/{output_filename}",
                    "matched_count": matched_count
                }
            
        except Exception as e:
            logger.error(f"Error modifying Excel: {e}", exc_info=True)
            return {"success": False, "error": str(e)}
    
    def analyze_data(self, filename: str, operation: str, column: Optional[str] = None,
                    filter_value: Optional[str] = None, 
                    aggregate_function: str = "count") -> Dict[str, Any]:
        """Perform data analysis operations"""
        try:
            # Load the Excel file
            file_path = self.uploads_dir / filename
            if not file_path.exists():
                return {"success": False, "error": f"File {filename} not found"}
            
            df = pd.read_excel(file_path)
            
            if operation == "statistics":
                if column and column in df.columns:
                    stats = df[column].describe().to_dict()
                    value_counts = df[column].value_counts().head(10).to_dict()
                    
                    return {
                        "success": True,
                        "operation": "statistics",
                        "column": column,
                        "statistics": stats,
                        "top_values": value_counts,
                        "message": f"Statistics for column '{column}'"
                    }
                else:
                    stats = df.describe(include='all').to_dict()
                    return {
                        "success": True,
                        "operation": "statistics",
                        "statistics": stats,
                        "message": "Overall statistics for all columns"
                    }
            
            elif operation == "unique_values":
                if not column:
                    return {"success": False, "error": "Column required for unique_values operation"}
                if column not in df.columns:
                    return {"success": False, "error": f"Column '{column}' not found"}
                
                unique_vals = df[column].unique().tolist()
                value_counts = df[column].value_counts().to_dict()
                
                return {
                    "success": True,
                    "operation": "unique_values",
                    "column": column,
                    "unique_values": unique_vals[:50],  # Limit to 50
                    "value_counts": value_counts,
                    "total_unique": len(unique_vals),
                    "message": f"Found {len(unique_vals)} unique values in '{column}'"
                }
            
            elif operation == "filter":
                if not column or not filter_value:
                    return {"success": False, "error": "Column and filter_value required"}
                if column not in df.columns:
                    return {"success": False, "error": f"Column '{column}' not found"}
                
                filtered_df = df[df[column].astype(str).str.contains(filter_value, case=False, na=False)]
                
                return {
                    "success": True,
                    "operation": "filter",
                    "column": column,
                    "filter_value": filter_value,
                    "matched_rows": len(filtered_df),
                    "sample_data": filtered_df.head(5).to_dict('records'),
                    "message": f"Found {len(filtered_df)} rows where '{column}' contains '{filter_value}'"
                }
            
            elif operation == "group_by":
                if not column:
                    return {"success": False, "error": "Column required for group_by operation"}
                if column not in df.columns:
                    return {"success": False, "error": f"Column '{column}' not found"}
                
                grouped = df.groupby(column).size().to_dict()
                
                return {
                    "success": True,
                    "operation": "group_by",
                    "column": column,
                    "groups": grouped,
                    "message": f"Grouped data by '{column}'"
                }
            
            elif operation == "sort":
                if not column:
                    return {"success": False, "error": "Column required for sort operation"}
                if column not in df.columns:
                    return {"success": False, "error": f"Column '{column}' not found"}
                
                sorted_df = df.sort_values(by=column)
                
                return {
                    "success": True,
                    "operation": "sort",
                    "column": column,
                    "top_rows": sorted_df.head(10).to_dict('records'),
                    "message": f"Sorted data by '{column}'"
                }
            
        except Exception as e:
            logger.error(f"Error analyzing data: {e}", exc_info=True)
            return {"success": False, "error": str(e)}

