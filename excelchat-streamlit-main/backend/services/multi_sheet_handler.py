"""
Multi-Sheet Excel Operation Handler
Handles operations that need to be applied to all sheets while preserving the multi-sheet structure
"""

import pandas as pd
from pathlib import Path
from typing import Dict, Any, List
import logging
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


def handle_multi_sheet_operation(
    session_data: Dict,
    session_id: str,
    operations: Dict[str, Any],
    excel_ops,
    outputs_dir: Path
) -> Dict[str, Any]:
    """
    Apply operations to all sheets in a multi-sheet Excel file
    Preserves the multi-sheet structure in the output
    
    Args:
        session_data: Session data dictionary
        session_id: Session identifier
        operations: Operations to apply
        excel_ops: ExcelOperations instance
        outputs_dir: Output directory path
    
    Returns:
        Result dictionary with success status and file info
    """
    try:
        # Get all sheets from session
        files = session_data[session_id]["files"]
        
        # Find all sheets (files with is_sheet=True)
        sheets_data = []
        for file_id, file_data in files.items():
            if file_data.get("is_sheet", False):
                sheets_data.append({
                    "sheet_name": file_data["sheet_name"],
                    "df": file_data["df"].copy(),
                    "sheet_index": file_data["sheet_index"]
                })
        
        if not sheets_data:
            return {
                "success": False,
                "error": "No multi-sheet Excel file found in session"
            }
        
        # Sort by sheet index to maintain order
        sheets_data.sort(key=lambda x: x["sheet_index"])
        
        logger.info(f"📊 Applying operations to {len(sheets_data)} sheets")
        
        # Check if this is a target_sheet operation
        target_sheet = operations.get('target_sheet')
        apply_to_all_sheets = operations.get('apply_to_all_sheets', False)
        
        # Apply operations to sheets based on scope
        processed_sheets = []
        total_rows_before = sum(len(s["df"]) for s in sheets_data)
        
        for sheet_info in sheets_data:
            sheet_name = sheet_info["sheet_name"]
            df = sheet_info["df"]
            
            # Determine if this sheet should be processed
            should_process = False
            if apply_to_all_sheets:
                should_process = True
                logger.info(f"🔄 Processing sheet '{sheet_name}' (apply_to_all_sheets): {len(df)} rows")
            elif target_sheet and sheet_name.lower() == target_sheet.lower():
                should_process = True
                logger.info(f"🎯 Processing TARGET sheet '{sheet_name}': {len(df)} rows")
            elif not target_sheet and not apply_to_all_sheets:
                # No specific target - apply to all (backward compatibility)
                should_process = True
                logger.info(f"🔄 Processing sheet '{sheet_name}' (no target specified): {len(df)} rows")
            else:
                logger.info(f"⏭️ Skipping sheet '{sheet_name}' (not target): {len(df)} rows")
            
            if should_process:
                # Apply operations to this sheet's DataFrame
                df_processed = _apply_operations_to_dataframe(df, operations)
                logger.info(f"✅ Applied operations to sheet '{sheet_name}': {len(df_processed)} rows")
            else:
                # Keep original DataFrame unchanged
                df_processed = df
                logger.info(f"📋 Preserved original sheet '{sheet_name}': {len(df_processed)} rows")
            
            processed_sheets.append({
                "sheet_name": sheet_name,
                "df": df_processed
            })
        
        total_rows_after = sum(len(s["df"]) for s in processed_sheets)
        
        # Create multi-sheet Excel file
        import time
        timestamp = int(time.time())
        
        # Check for custom filename
        if 'custom_filename' in operations and operations['custom_filename']:
            custom_name = operations['custom_filename']
            custom_name_clean = ''.join(c for c in custom_name if c.isalnum() or c in ('_', '-')).strip()
            filename = f"{custom_name_clean}_{timestamp}.xlsx" if custom_name_clean else f"excel_multi_sheet_{timestamp}.xlsx"
        else:
            filename = f"excel_multi_sheet_{timestamp}.xlsx"
        
        output_path = outputs_dir / filename
        
        # Create workbook with multiple sheets
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_info in processed_sheets:
                sheet_info["df"].to_excel(
                    writer,
                    sheet_name=sheet_info["sheet_name"],
                    index=False
                )
        
        # Apply Excel formatting if needed (freeze panes, auto filters, etc.)
        if 'freeze_panes' in operations:
            _apply_freeze_panes_to_all_sheets(output_path, operations['freeze_panes'], processed_sheets)
        
        # Apply auto filters to all sheets (default: True)
        if operations.get('auto_filter', True):
            _apply_auto_filters_to_all_sheets(output_path, processed_sheets)
        
        # Apply subtotals if specified
        if 'subtotals' in operations:
            _apply_subtotals_to_sheet(output_path, operations['subtotals'], processed_sheets)
        
        # Apply highlighting if specified
        if 'highlight_rows' in operations:
            # Pass target_sheet info to highlighting function
            highlight_spec = operations['highlight_rows']
            if target_sheet and not isinstance(highlight_spec, list):
                # Add target_sheet to single highlight spec
                highlight_spec = dict(highlight_spec)
                highlight_spec['target_sheet'] = target_sheet
            elif target_sheet and isinstance(highlight_spec, list):
                # Add target_sheet to each highlight spec in list
                highlight_spec = [dict(hl, target_sheet=target_sheet) for hl in highlight_spec]
            
            _apply_highlighting_to_sheet(output_path, highlight_spec, processed_sheets)
        
        # Apply conditional formatting if specified
        if 'conditional_format' in operations:
            _apply_conditional_format_to_sheet(output_path, operations['conditional_format'], processed_sheets)
        
        logger.info(f"✅ Created multi-sheet Excel: {filename} with {len(processed_sheets)} sheets")
        
        # Generate summary
        operation_summary = _generate_multi_sheet_summary(
            operations,
            processed_sheets,
            total_rows_before,
            total_rows_after
        )
        
        download_url = f"/api/download/{filename}"
        
        return {
            "success": True,
            "session_id": session_id,
            "operation_type": "excel_file",
            "excel_file": filename,
            "download_url": download_url,
            "generated_code": "",
            "explanation": operation_summary,
            "result": {
                "success": True,
                "execution_output": operation_summary,
                "excel_files": [download_url],
                "plots": [],
                "dataframes": {},
                "variables": {}
            },
            "data_shape": f"{len(processed_sheets)} sheets, {total_rows_after} total rows"
        }
        
    except Exception as e:
        logger.error(f"❌ Error in multi-sheet operation: {e}", exc_info=True)
        return {
            "success": False,
            "error": f"Failed to process multi-sheet operation: {str(e)}"
        }


def _apply_operations_to_dataframe(df: pd.DataFrame, operations: Dict[str, Any]) -> pd.DataFrame:
    """Apply operations to a single DataFrame"""
    df_result = df.copy()
    
    # Remove last row
    if operations.get('remove_last_row'):
        if len(df_result) > 0:
            df_result = df_result.iloc[:-1]
            logger.info(f"  ✂️ Removed last row")
    
    # Delete specific rows
    if 'delete_rows' in operations:
        delete_spec = operations['delete_rows']
        
        if 'row_numbers' in delete_spec:
            # Delete by row numbers (1-indexed from user)
            row_numbers = delete_spec['row_numbers']
            # Convert to 0-indexed
            indices_to_drop = [r - 1 for r in row_numbers if 0 <= r - 1 < len(df_result)]
            df_result = df_result.drop(df_result.index[indices_to_drop])
            logger.info(f"  ✂️ Deleted {len(indices_to_drop)} specific rows")
            
        elif 'column' in delete_spec and 'condition' in delete_spec:
            # Delete by condition
            column = delete_spec['column']
            condition = delete_spec['condition']
            
            if column in df_result.columns:
                operator = condition.get('operator', '==')
                value = condition.get('value')
                
                # Create mask for rows to delete
                if operator == '==':
                    mask = df_result[column] == value
                elif operator == '!=':
                    mask = df_result[column] != value
                elif operator == 'in':
                    mask = df_result[column].isin(value if isinstance(value, list) else [value])
                elif operator == '>':
                    mask = df_result[column] > value
                elif operator == '<':
                    mask = df_result[column] < value
                elif operator == '>=':
                    mask = df_result[column] >= value
                elif operator == '<=':
                    mask = df_result[column] <= value
                else:
                    mask = pd.Series([False] * len(df_result))
                
                rows_to_delete = mask.sum()
                df_result = df_result[~mask]
                logger.info(f"  ✂️ Deleted {rows_to_delete} rows matching condition")
    
    # Filter
    if 'filter' in operations:
        filter_ops = operations['filter']
        for column, condition in filter_ops.items():
            if column in df_result.columns:
                if isinstance(condition, dict):
                    operator = condition.get('operator', '==')
                    value = condition.get('value')
                    
                    if operator == '==':
                        df_result = df_result[df_result[column] == value]
                    elif operator == '!=':
                        df_result = df_result[df_result[column] != value]
                    elif operator == 'in':
                        df_result = df_result[df_result[column].isin(value if isinstance(value, list) else [value])]
                    elif operator == '>':
                        df_result = df_result[df_result[column] > value]
                    elif operator == '<':
                        df_result = df_result[df_result[column] < value]
                    elif operator == '>=':
                        df_result = df_result[df_result[column] >= value]
                    elif operator == '<=':
                        df_result = df_result[df_result[column] <= value]
                    
                    logger.info(f"  🔍 Filtered {column} {operator} {value}")
    
    # Sort
    if 'sort' in operations:
        sort_spec = operations['sort']
        by_columns = sort_spec.get('by', [])
        ascending = sort_spec.get('ascending', True)
        
        if by_columns:
            valid_columns = [col for col in by_columns if col in df_result.columns]
            if valid_columns:
                df_result = df_result.sort_values(by=valid_columns, ascending=ascending)
                logger.info(f"  📊 Sorted by {valid_columns}")
    
    return df_result


def _apply_freeze_panes_to_all_sheets(file_path: Path, freeze_spec: Dict, sheets: List[Dict]):
    """Apply freeze panes to all sheets in the workbook"""
    try:
        wb = load_workbook(file_path)
        
        freeze_row = freeze_spec.get('row', 1)
        freeze_col = freeze_spec.get('col', 0)
        
        # Convert to 1-indexed for openpyxl
        freeze_row_idx = freeze_row + 1 if freeze_row > 0 else 1
        freeze_col_idx = freeze_col + 1 if freeze_col > 0 else 1
        
        for sheet_info in sheets:
            sheet_name = sheet_info["sheet_name"]
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Set freeze panes
                if freeze_row > 0 or freeze_col > 0:
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(freeze_col_idx)
                    freeze_cell = f"{col_letter}{freeze_row_idx}"
                    ws.freeze_panes = freeze_cell
                    logger.info(f"  ❄️ Froze panes at {freeze_cell} in sheet '{sheet_name}'")
        
        wb.save(file_path)
        wb.close()
        
    except Exception as e:
        logger.error(f"⚠️ Error applying freeze panes: {e}")


def _apply_auto_filters_to_all_sheets(file_path: Path, sheets: List[Dict]):
    """Apply auto filters to all sheets in the workbook"""
    try:
        wb = load_workbook(file_path)
        
        for sheet_info in sheets:
            sheet_name = sheet_info["sheet_name"]
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Apply auto filter to the entire data range
                if ws.max_row > 1 and ws.max_column > 0:
                    ws.auto_filter.ref = ws.dimensions
                    logger.info(f"  🔽 Applied auto filter to sheet '{sheet_name}'")
        
        wb.save(file_path)
        wb.close()
        
    except Exception as e:
        logger.error(f"⚠️ Error applying auto filters: {e}")


def _generate_multi_sheet_summary(
    operations: Dict[str, Any],
    sheets: List[Dict],
    rows_before: int,
    rows_after: int
) -> str:
    """Generate summary for multi-sheet operation"""
    
    operation_parts = []
    
    # Describe what was done
    if operations.get('remove_last_row'):
        operation_parts.append("Removed last row from each sheet")
    
    if 'delete_rows' in operations:
        operation_parts.append("Deleted rows")
    
    if 'filter' in operations:
        operation_parts.append("Filtered data")
    
    if 'sort' in operations:
        operation_parts.append("Sorted data")
    
    if 'freeze_panes' in operations:
        freeze_spec = operations['freeze_panes']
        freeze_row = freeze_spec.get('row', 1)
        freeze_col = freeze_spec.get('col', 0)
        if freeze_row > 0 or freeze_col > 0:
            operation_parts.append("Froze panes")
    
    operation_desc = ", ".join(operation_parts) if operation_parts else "Applied operations"
    
    # Build summary
    sheet_names = [s["sheet_name"] for s in sheets]
    sheet_list = ", ".join(sheet_names)
    
    summary = f"✅ {operation_desc} to {len(sheets)} sheets ({sheet_list}) - "
    summary += f"{rows_before} rows → {rows_after} rows total. "
    summary += f"Download the multi-sheet Excel file."
    
    return summary


def _apply_subtotals_to_sheet(file_path: Path, subtotal_spec: Dict, sheets: List[Dict]):
    """Apply subtotals to the specified sheet in the workbook"""
    try:
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        wb = load_workbook(file_path)
        
        # Determine target sheet
        target_sheet_name = subtotal_spec.get('target_sheet')
        if not target_sheet_name:
            # Apply to first sheet by default
            target_sheet_name = sheets[0]["sheet_name"]
        
        if target_sheet_name not in wb.sheetnames:
            logger.warning(f"⚠️ Target sheet '{target_sheet_name}' not found for subtotals")
            wb.close()
            return
        
        ws = wb[target_sheet_name]
        df = next((s["df"] for s in sheets if s["sheet_name"] == target_sheet_name), None)
        
        if df is None:
            logger.warning(f"⚠️ DataFrame not found for sheet '{target_sheet_name}'")
            wb.close()
            return
        
        group_by = subtotal_spec.get('group_by')
        aggregate_column = subtotal_spec.get('aggregate_column')
        function = subtotal_spec.get('function', 'count').upper()
        
        if not group_by or group_by not in df.columns:
            logger.warning(f"⚠️ Subtotal group_by column '{group_by}' not found")
            wb.close()
            return
        
        if not aggregate_column or aggregate_column not in df.columns:
            logger.warning(f"⚠️ Subtotal aggregate_column '{aggregate_column}' not found")
            wb.close()
            return
        
        # Get column indices (1-indexed for Excel)
        group_col_idx = df.columns.get_loc(group_by) + 1
        agg_col_idx = df.columns.get_loc(aggregate_column) + 1
        
        # Track groups with their start and end rows
        groups = []
        current_group = None
        group_start_row = 2  # Data starts at row 2 (after header)
        
        for row_idx in range(2, ws.max_row + 1):
            group_value = ws.cell(row=row_idx, column=group_col_idx).value
            
            if current_group is not None and group_value != current_group:
                groups.append({
                    'name': current_group,
                    'start_row': group_start_row,
                    'end_row': row_idx - 1,
                    'insert_at': row_idx
                })
                group_start_row = row_idx
            
            current_group = group_value
        
        # Add final group
        if current_group is not None:
            groups.append({
                'name': current_group,
                'start_row': group_start_row,
                'end_row': ws.max_row,
                'insert_at': ws.max_row + 1
            })
        
        logger.info(f"📊 Found {len(groups)} groups for subtotals in sheet '{target_sheet_name}'")
        
        # Insert subtotal rows (in reverse to maintain row numbers)
        for group in reversed(groups):
            insert_row = group['insert_at']
            group_name = group['name']
            start_row = group['start_row']
            end_row = group['end_row']
            
            # Insert new row
            ws.insert_rows(insert_row)
            
            # Add subtotal label
            ws.cell(row=insert_row, column=group_col_idx, value=f"{group_name} Count")
            
            # Calculate the value
            if function == 'COUNT':
                count_value = sum(1 for r in range(start_row, end_row + 1) 
                                 if ws.cell(row=r, column=agg_col_idx).value is not None)
                calculated_value = count_value
            elif function == 'SUM':
                sum_value = sum(float(ws.cell(row=r, column=agg_col_idx).value or 0) 
                               for r in range(start_row, end_row + 1))
                calculated_value = sum_value
            else:  # AVERAGE
                values = [float(ws.cell(row=r, column=agg_col_idx).value or 0) 
                         for r in range(start_row, end_row + 1)]
                calculated_value = sum(values) / len(values) if values else 0
            
            # Set the cell with the calculated value
            cell = ws.cell(row=insert_row, column=agg_col_idx)
            cell.value = calculated_value
            cell.number_format = '0' if function == 'COUNT' else '0.00'
            
            # Format subtotal row (bold, gray background)
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=insert_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='FFE0E0E0', end_color='FFE0E0E0', fill_type='solid')
            
            logger.info(f"  ➕ Subtotal for '{group_name}': {calculated_value} items")
        
        wb.save(file_path)
        wb.close()
        logger.info(f"✅ Applied {len(groups)} subtotals to sheet '{target_sheet_name}'")
        
    except Exception as e:
        logger.error(f"⚠️ Error applying subtotals: {e}", exc_info=True)


def _apply_highlighting_to_sheet(file_path: Path, highlight_spec, sheets: List[Dict]):
    """Apply row highlighting to the specified sheet"""
    try:
        from openpyxl.styles import PatternFill
        
        wb = load_workbook(file_path)
        
        # Handle both single and multiple highlights
        highlights = highlight_spec if isinstance(highlight_spec, list) else [highlight_spec]
        
        for hl_spec in highlights:
            target_sheet_name = hl_spec.get('target_sheet')
            if not target_sheet_name:
                target_sheet_name = sheets[0]["sheet_name"]
            
            if target_sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[target_sheet_name]
            df = next((s["df"] for s in sheets if s["sheet_name"] == target_sheet_name), None)
            
            if df is None:
                continue
            
            column = hl_spec.get('column')
            condition = hl_spec.get('condition', {})
            color = hl_spec.get('color', 'yellow')
            
            if not column or column not in df.columns:
                continue
            
            col_idx = df.columns.get_loc(column) + 1
            
            # Color mapping
            color_map = {
                'red': 'FFFF0000', 'green': 'FF00FF00', 'yellow': 'FFFFFF00',
                'blue': 'FF0000FF', 'orange': 'FFFFA500', 'purple': 'FF800080',
                'pink': 'FFFFC0CB', 'cyan': 'FF00FFFF', 'light_green': 'FF90EE90',
                'light_blue': 'FFADD8E6', 'light_yellow': 'FFFFFFE0',
                'light_red': 'FFFFCCCB', 'gray': 'FFD3D3D3'
            }
            fill_color = color_map.get(color, 'FFFFFF00')
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            # Apply highlighting
            operator = condition.get('operator', '==')
            value = condition.get('value')
            
            for row_idx in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                
                match = False
                if operator == '==':
                    match = cell_value == value
                elif operator == '!=':
                    match = cell_value != value
                elif operator == 'in':
                    match = cell_value in (value if isinstance(value, list) else [value])
                
                if match:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col).fill = fill
            
            logger.info(f"  🎨 Applied highlighting to sheet '{target_sheet_name}'")
        
        wb.save(file_path)
        wb.close()
        
    except Exception as e:
        logger.error(f"⚠️ Error applying highlighting: {e}", exc_info=True)


def _apply_conditional_format_to_sheet(file_path: Path, cf_spec, sheets: List[Dict]):
    """Apply conditional formatting to the specified sheet"""
    try:
        from openpyxl.styles import PatternFill
        
        wb = load_workbook(file_path)
        
        # Handle both single and multiple conditional formats
        formats = cf_spec if isinstance(cf_spec, list) else [cf_spec]
        
        for fmt_spec in formats:
            target_sheet_name = fmt_spec.get('target_sheet')
            if not target_sheet_name:
                target_sheet_name = sheets[0]["sheet_name"]
            
            if target_sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[target_sheet_name]
            df = next((s["df"] for s in sheets if s["sheet_name"] == target_sheet_name), None)
            
            if df is None:
                continue
            
            column = fmt_spec.get('column')
            rules = fmt_spec.get('rules', [])
            
            if not column or column not in df.columns:
                continue
            
            col_idx = df.columns.get_loc(column) + 1
            
            # Color mapping
            color_map = {
                'red': 'FFFF0000', 'green': 'FF00FF00', 'yellow': 'FFFFFF00',
                'blue': 'FF0000FF', 'orange': 'FFFFA500', 'purple': 'FF800080',
                'amber': 'FFFFBF00'
            }
            
            # Apply each rule
            for rule in rules:
                condition = rule.get('condition', {})
                color = rule.get('color', 'yellow')
                fill_color = color_map.get(color, 'FFFFFF00')
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                
                operator = condition.get('operator', '==')
                value = condition.get('value')
                
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell_value = cell.value
                    
                    match = False
                    if operator == '==':
                        match = cell_value == value
                    elif operator == '!=':
                        match = cell_value != value
                    elif operator == '>':
                        match = cell_value is not None and cell_value > value
                    elif operator == '<':
                        match = cell_value is not None and cell_value < value
                    elif operator == '>=':
                        match = cell_value is not None and cell_value >= value
                    elif operator == '<=':
                        match = cell_value is not None and cell_value <= value
                    
                    if match:
                        cell.fill = fill
            
            logger.info(f"  🎨 Applied conditional formatting to sheet '{target_sheet_name}'")
        
        wb.save(file_path)
        wb.close()
        
    except Exception as e:
        logger.error(f"⚠️ Error applying conditional formatting: {e}", exc_info=True)
