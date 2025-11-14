"""
Excel Operations Module
Handles intelligent Excel file operations based on LLM-understood user intent
Supports: highlighting, filtering, sorting, conditional formatting, and more
"""

import logging
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime
import time
import json
import re
import sys
import google.generativeai as genai

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# Import configuration
backend_path = Path(__file__).resolve().parent.parent
if str(backend_path) not in sys.path:
    sys.path.insert(0, str(backend_path))
from config.settings import settings, error_messages

logger = logging.getLogger(__name__)


class ExcelOperations:
    """
    Handles various Excel operations with intelligent formatting
    """
    
    # Predefined color schemes
    COLORS = {
        'red': 'FFFF0000',
        'green': 'FF00FF00',
        'yellow': 'FFFFFF00',
        'blue': 'FF0000FF',
        'orange': 'FFFFA500',
        'purple': 'FF800080',
        'pink': 'FFFFC0CB',
        'cyan': 'FF00FFFF',
        'light_green': 'FF90EE90',
        'light_blue': 'FFADD8E6',
        'light_yellow': 'FFFFFFE0',
        'light_red': 'FFFFCCCB',
        'gray': 'FFD3D3D3'
    }
    
    def __init__(self, outputs_dir: Path):
        self.outputs_dir = Path(outputs_dir)
        self.outputs_dir.mkdir(exist_ok=True, parents=True)
        self.google_sheets_formatter = None  # Lazy load when needed
    
    def create_excel_with_operations(
        self,
        df: pd.DataFrame,
        operations: Dict[str, Any],
        filename: str = "output.xlsx",
        source_file_path: str = None
    ) -> str:
        """
        Create Excel file with specified operations applied
        
        Args:
            df: Source DataFrame
            operations: Dict containing operation specifications
            filename: Output filename
            source_file_path: Optional path to source Excel file to preserve formatting
            
        Returns:
            Path to created Excel file
        """
        try:
            output_path = self.outputs_dir / filename
            
            logger.info(f"📝 Creating Excel with operations: {operations}")
            logger.info(f"📊 Input DataFrame shape: {df.shape}")
            
            # Apply data operations (filtering, sorting) first
            processed_df = self._apply_data_operations(df.copy(), operations)
            logger.info(f"📊 After data operations shape: {processed_df.shape}")
            
            # If source file provided, try to preserve its formatting
            if source_file_path and Path(source_file_path).exists():
                try:
                    logger.info(f"📋 Preserving formatting from source file: {source_file_path}")
                    return self._create_excel_preserving_format(
                        processed_df, operations, output_path, source_file_path
                    )
                except Exception as e:
                    logger.warning(f"⚠️ Could not preserve formatting: {e}, creating new file")
            
            # Create Excel workbook from scratch
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            
            # Write DataFrame to worksheet
            for r_idx, row in enumerate(dataframe_to_rows(processed_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Apply header formatting
                    if r_idx == 1:
                        cell.font = Font(bold=True, color='FFFFFFFF')
                        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply formatting operations (highlighting, conditional formatting)
            self._apply_formatting_operations(ws, processed_df, operations)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
            # Add filters to header row (always applied unless explicitly disabled)
            if operations.get('auto_filter', True):  # Default to True
                ws.auto_filter.ref = ws.dimensions
                logger.info("✅ Auto filter applied to header row")
            
            # Apply freeze panes if requested
            if 'freeze_panes' in operations:
                freeze_spec = operations['freeze_panes']
                freeze_row = freeze_spec.get('row', 1)
                freeze_col = freeze_spec.get('col', 0)
                # openpyxl uses 1-indexed, so add 1 to our 0-indexed values
                # freeze_panes expects the cell AFTER the freeze point
                ws.freeze_panes = ws.cell(row=freeze_row + 1, column=freeze_col + 1)
                logger.info(f"✅ Froze panes at row {freeze_row}, col {freeze_col}")
            
            # Apply subtotals if requested
            if 'subtotals' in operations:
                self._apply_subtotals(ws, processed_df, operations['subtotals'])
            
            # Set workbook to calculate formulas on load
            wb.calculation.calcMode = 'auto'
            wb.calculation.fullCalcOnLoad = True
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"✅ Excel file created with operations: {output_path}")
            
            return str(output_path)
            
        except Exception as e:
            logger.error(f"❌ Error creating Excel file: {e}", exc_info=True)
            raise
    
    def apply_operations_to_google_sheet(
        self,
        df: pd.DataFrame,
        operations: Dict[str, Any],
        original_url: str,
        session_data: Dict = None,
        session_id: str = None
    ) -> Dict[str, Any]:
        """
        Apply operations to Google Sheet and return new sheet URL
        
        Args:
            df: Source DataFrame
            operations: Dict containing operation specifications
            original_url: Original Google Sheets URL
            
        Returns:
            Dict with success status and new sheet URL
        """
        try:
            # Lazy load Google Sheets formatter
            if self.google_sheets_formatter is None:
                from .google_sheets_formatter import GoogleSheetsFormatter
                self.google_sheets_formatter = GoogleSheetsFormatter()
            
            logger.info(f"📊 Applying operations to Google Sheet...")
            
            # Apply data operations (filtering, sorting) first
            processed_df = self._apply_data_operations(df.copy(), operations)
            logger.info(f"📊 After data operations shape: {processed_df.shape}")
            
            # Reset index to ensure sequential row numbers (0, 1, 2, ...)
            # This is crucial after filtering which may leave gaps in indices
            processed_df = processed_df.reset_index(drop=True)
            logger.info(f"📊 Reset index - new shape: {processed_df.shape}")
            
            # Prepare operations for Google Sheets formatter
            gs_operations = self._prepare_google_sheets_operations(operations, df, processed_df)
            
            # Pass session data to formatter for session storage (matching Excel flow)
            if session_data and session_id:
                self.google_sheets_formatter.session_data = session_data
                self.google_sheets_formatter.session_id = session_id
            
            # Apply operations using Google Sheets formatter
            result = self.google_sheets_formatter.apply_operations_to_google_sheet(
                df=processed_df,
                operations=gs_operations,
                original_url=original_url
            )
            
            return result
            
        except Exception as e:
            logger.error(f"❌ Error applying operations to Google Sheet: {e}", exc_info=True)
            return {
                'success': False,
                'error': str(e)
            }
    
    def _prepare_google_sheets_operations(
        self,
        operations: Dict[str, Any],
        original_df: pd.DataFrame,
        processed_df: pd.DataFrame
    ) -> Dict[str, Any]:
        """Prepare operations dict for Google Sheets formatter"""
        gs_ops = {}
        
        # Reset index of processed_df to ensure sequential indices (0, 1, 2, ...)
        # This is crucial after filtering which may leave gaps in indices
        processed_df = processed_df.reset_index(drop=True)
        
        # Handle row highlighting
        if 'highlight_rows' in operations:
            highlight_specs = operations['highlight_rows']
            if not isinstance(highlight_specs, list):
                highlight_specs = [highlight_specs]
            
            for spec in highlight_specs:
                column = spec.get('column')
                condition = spec.get('condition')
                color = spec.get('color', 'yellow')
                
                # Find matching rows in processed DataFrame
                matching_rows = []
                if column and condition:
                    mask = self._check_condition_mask(processed_df[column], condition)
                    matching_rows = processed_df[mask].index.tolist()
                
                gs_ops['highlight_rows'] = {
                    'color': color,
                    'row_indices': matching_rows
                }
        
        # Handle cell highlighting
        if 'highlight_cells' in operations:
            highlight_specs = operations['highlight_cells']
            if not isinstance(highlight_specs, list):
                highlight_specs = [highlight_specs]
            
            for spec in highlight_specs:
                column = spec.get('column')
                condition = spec.get('condition')
                color = spec.get('color', 'yellow')
                
                # Find matching cells
                matching_cells = []
                if column and condition:
                    mask = self._check_condition_mask(processed_df[column], condition)
                    matching_cells = processed_df[mask].index.tolist()
                
                gs_ops['highlight_cells'] = {
                    'column': column,
                    'color': color,
                    'cell_indices': matching_cells
                }
        
        # Handle conditional formatting
        if 'conditional_format' in operations:
            format_specs = operations['conditional_format']
            if not isinstance(format_specs, list):
                format_specs = [format_specs]
            
            gs_ops['conditional_format'] = format_specs
        
        # Handle duplicate highlighting
        if 'highlight_duplicates' in operations:
            gs_ops['highlight_duplicates'] = operations['highlight_duplicates']
        
        # Handle null highlighting
        if 'highlight_nulls' in operations:
            gs_ops['highlight_nulls'] = operations['highlight_nulls']
        
        # Handle tab renaming
        if 'rename_tabs' in operations:
            gs_ops['rename_tabs'] = operations['rename_tabs']
        
        # Handle merge files operation
        if 'merge_files' in operations:
            gs_ops['merge_files'] = operations['merge_files']
            # Also pass sheet names if specified
            if 'sheet_names' in operations:
                gs_ops['sheet_names'] = operations['sheet_names']
        
        return gs_ops
    
    def _create_excel_preserving_format(
        self,
        processed_df: pd.DataFrame,
        operations: Dict[str, Any],
        output_path: Path,
        source_file_path: str
    ) -> str:
        """
        Create Excel file while preserving original formatting from source file
        """
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from copy import copy
        
        # Load the original workbook to preserve formatting
        wb = load_workbook(source_file_path)
        ws = wb.active
        
        # Find where the actual data starts (skip title rows, headers, etc.)
        data_start_row = None
        header_row = None
        
        # Look for the header row (row with column names)
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20), start=1):
            row_values = [cell.value for cell in row if cell.value]
            if row_values:
                # Check if this row matches DataFrame columns
                df_cols_lower = [str(col).lower().strip() for col in processed_df.columns]
                row_vals_lower = [str(val).lower().strip() for val in row_values if val]
                
                # If at least 50% of columns match, this is likely the header
                matches = sum(1 for val in row_vals_lower if val in df_cols_lower)
                if matches >= len(df_cols_lower) * 0.5:
                    header_row = row_idx
                    data_start_row = row_idx + 1
                    logger.info(f"📍 Found header at row {header_row}, data starts at row {data_start_row}")
                    break
        
        if not data_start_row:
            # Fallback: assume data starts after first non-empty row
            data_start_row = 2
            header_row = 1
            logger.warning(f"⚠️ Could not find header row, assuming row {header_row}")
        
        # Store original formatting from header and first data row
        original_header_formats = {}
        original_data_formats = {}
        
        if header_row:
            for col_idx, cell in enumerate(ws[header_row], start=1):
                if cell.value:
                    original_header_formats[col_idx] = {
                        'font': copy(cell.font),
                        'fill': copy(cell.fill),
                        'border': copy(cell.border),
                        'alignment': copy(cell.alignment),
                        'number_format': cell.number_format
                    }
        
        if data_start_row and data_start_row <= ws.max_row:
            for col_idx, cell in enumerate(ws[data_start_row], start=1):
                original_data_formats[col_idx] = {
                    'font': copy(cell.font),
                    'fill': copy(cell.fill),
                    'border': copy(cell.border),
                    'alignment': copy(cell.alignment),
                    'number_format': cell.number_format
                }
        
        # Delete old data rows (keep title rows and header)
        if ws.max_row >= data_start_row:
            ws.delete_rows(data_start_row, ws.max_row - data_start_row + 1)
        
        # Write new data starting from data_start_row
        for r_idx, row_data in enumerate(dataframe_to_rows(processed_df, index=False, header=False), start=data_start_row):
            for c_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Apply original data formatting if available
                if c_idx in original_data_formats:
                    fmt = original_data_formats[c_idx]
                    cell.font = copy(fmt['font'])
                    cell.fill = copy(fmt['fill'])
                    cell.border = copy(fmt['border'])
                    cell.alignment = copy(fmt['alignment'])
                    cell.number_format = fmt['number_format']
        
        # Apply new formatting operations (highlighting, etc.) on top of preserved formatting
        self._apply_formatting_operations(ws, processed_df, operations, data_start_row=data_start_row)
        
        # Apply freeze panes if requested
        if 'freeze_panes' in operations:
            freeze_spec = operations['freeze_panes']
            freeze_row = freeze_spec.get('row', 1)
            freeze_col = freeze_spec.get('col', 0)
            ws.freeze_panes = ws.cell(row=freeze_row + 1, column=freeze_col + 1)
        
        # Apply subtotals if requested
        if 'subtotals' in operations:
            self._apply_subtotals(ws, processed_df, operations['subtotals'])
        
        # Set workbook to calculate formulas on load
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True
        
        # Save the workbook
        wb.save(output_path)
        logger.info(f"✅ Excel file created with preserved formatting: {output_path}")
        
        return str(output_path)
    
    def _apply_data_operations(self, df: pd.DataFrame, operations: Dict[str, Any]) -> pd.DataFrame:
        """Apply data manipulation operations (filter, sort, etc.)"""
        
        # Remove last row (e.g., filter description row)
        if operations.get('remove_last_row'):
            if len(df) > 0:
                df = df.iloc[:-1].copy()
                logger.info(f"✅ Removed last row - new shape: {df.shape}")
        
        # Delete specific rows
        if 'delete_rows' in operations:
            delete_spec = operations['delete_rows']
            
            # Delete by row numbers
            if 'row_numbers' in delete_spec:
                row_numbers = delete_spec['row_numbers']
                # Convert to 0-indexed and filter out invalid indices
                indices_to_drop = [i - 1 for i in row_numbers if 0 < i <= len(df)]
                if indices_to_drop:
                    df = df.drop(df.index[indices_to_drop]).reset_index(drop=True)
                    logger.info(f"✅ Deleted {len(indices_to_drop)} rows - new shape: {df.shape}")
            
            # Delete by condition
            elif 'column' in delete_spec and 'condition' in delete_spec:
                column = delete_spec['column']
                condition = delete_spec['condition']
                if column in df.columns:
                    # Find rows that match the condition
                    original_len = len(df)
                    mask = self._check_condition_mask(df[column], condition)
                    # Keep rows that DON'T match (inverse of filter)
                    df = df[~mask].reset_index(drop=True)
                    deleted_count = original_len - len(df)
                    logger.info(f"✅ Deleted {deleted_count} rows matching condition - new shape: {df.shape}")
        
        # Filtering
        if 'filter' in operations:
            filter_ops = operations['filter']
            if isinstance(filter_ops, dict):
                for column, condition in filter_ops.items():
                    if column in df.columns:
                        df = self._apply_filter(df, column, condition)
        
        # Sorting
        if 'sort' in operations:
            sort_ops = operations['sort']
            if isinstance(sort_ops, dict):
                by = sort_ops.get('by', [])
                ascending = sort_ops.get('ascending', True)
                if by:
                    df = df.sort_values(by=by, ascending=ascending)
        
        # Top N rows
        if 'top_n' in operations:
            n = operations['top_n']
            df = df.head(n)
        
        # Bottom N rows
        if 'bottom_n' in operations:
            n = operations['bottom_n']
            df = df.tail(n)
        
        return df
    
    def _apply_filter(self, df: pd.DataFrame, column: str, condition: Any) -> pd.DataFrame:
        """Apply filter condition to DataFrame"""
        try:
            if isinstance(condition, dict):
                operator = condition.get('operator', '==')
                value = condition.get('value')
                
                if operator == '==':
                    return df[df[column] == value]
                elif operator == '!=':
                    return df[df[column] != value]
                elif operator == '>':
                    return df[df[column] > value]
                elif operator == '<':
                    return df[df[column] < value]
                elif operator == '>=':
                    return df[df[column] >= value]
                elif operator == '<=':
                    return df[df[column] <= value]
                elif operator == 'contains':
                    return df[df[column].astype(str).str.contains(str(value), case=False, na=False)]
                elif operator == 'in':
                    return df[df[column].isin(value)]
            else:
                # Simple equality filter
                return df[df[column] == condition]
        except Exception as e:
            logger.warning(f"⚠️ Filter operation failed for {column}: {e}")
            return df
    
    def _check_condition_mask(self, series: pd.Series, condition: Any) -> pd.Series:
        """Check condition and return boolean mask for series"""
        try:
            if isinstance(condition, dict):
                operator = condition.get('operator', '==')
                value = condition.get('value')
                
                if operator == '==':
                    return series == value
                elif operator == '!=':
                    return series != value
                elif operator == '>':
                    return series > value
                elif operator == '<':
                    return series < value
                elif operator == '>=':
                    return series >= value
                elif operator == '<=':
                    return series <= value
                elif operator == 'contains':
                    return series.astype(str).str.contains(str(value), case=False, na=False)
                elif operator == 'in':
                    return series.isin(value)
            else:
                # Simple equality
                return series == condition
        except Exception as e:
            logger.warning(f"⚠️ Condition check failed: {e}")
            return pd.Series([False] * len(series))
    
    def _apply_formatting_operations(
        self,
        ws,
        df: pd.DataFrame,
        operations: Dict[str, Any],
        data_start_row: int = 2
    ):
        """Apply formatting operations (highlighting, conditional formatting)
        
        Args:
            ws: Worksheet object
            df: DataFrame with data
            operations: Operations to apply
            data_start_row: Row number where data starts (default 2, after header)
        """
        
        # Highlighting specific rows - handle both single and multiple
        if 'highlight_rows' in operations:
            highlight_spec = operations['highlight_rows']
            if isinstance(highlight_spec, list):
                # Multiple highlights
                for hl in highlight_spec:
                    self._highlight_rows(ws, df, hl, data_start_row)
            else:
                # Single highlight
                self._highlight_rows(ws, df, highlight_spec, data_start_row)
        
        # Highlighting specific cells - handle both single and multiple
        if 'highlight_cells' in operations:
            highlight_spec = operations['highlight_cells']
            if isinstance(highlight_spec, list):
                # Multiple highlights
                for hl in highlight_spec:
                    self._highlight_cells(ws, df, hl, data_start_row)
            else:
                # Single highlight
                self._highlight_cells(ws, df, highlight_spec, data_start_row)
        
        # Conditional formatting - handle both single and multiple
        if 'conditional_format' in operations:
            format_spec = operations['conditional_format']
            if isinstance(format_spec, list):
                # Multiple conditional formats
                for cf in format_spec:
                    self._apply_conditional_formatting(ws, df, cf, data_start_row)
            else:
                # Single conditional format
                self._apply_conditional_formatting(ws, df, format_spec, data_start_row)
        
        # Highlight duplicates - handle both single and multiple
        if 'highlight_duplicates' in operations:
            dup_spec = operations['highlight_duplicates']
            if isinstance(dup_spec, list):
                # Multiple duplicate highlights
                for ds in dup_spec:
                    self._highlight_duplicates(ws, df, ds)
            else:
                # Single duplicate highlight
                self._highlight_duplicates(ws, df, dup_spec)
        
        # Highlight nulls/missing values - handle both single and multiple
        if 'highlight_nulls' in operations:
            null_spec = operations['highlight_nulls']
            if isinstance(null_spec, list):
                # Multiple null highlights
                for ns in null_spec:
                    self._highlight_nulls(ws, df, ns)
            else:
                # Single null highlight
                self._highlight_nulls(ws, df, null_spec)
    
    def _highlight_rows(self, ws, df: pd.DataFrame, highlight_spec: Dict[str, Any], data_start_row: int = 2):
        """Highlight entire rows based on conditions"""
        try:
            column = highlight_spec.get('column')
            condition = highlight_spec.get('condition')
            color = highlight_spec.get('color', 'yellow')
            
            if not column or column not in df.columns:
                return
            
            # Get color code
            fill_color = self.COLORS.get(color, self.COLORS['yellow'])
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            # Find matching rows
            col_idx = df.columns.get_loc(column) + 1
            
            # Handle topN/bottomN operators specially (case-insensitive)
            operator_str = condition.get('operator', '').upper().replace('_', '') if isinstance(condition, dict) else ''
            if operator_str in ['TOPN', 'BOTTOMN']:
                n = condition.get('value', 5)
                
                # Get top/bottom N values
                if 'TOP' in operator_str:
                    top_values = df.nlargest(n, column)[column].values
                else:  # BOTTOM
                    top_values = df.nsmallest(n, column)[column].values
                
                # Highlight rows where the column value is in top/bottom N values
                # We iterate through the DataFrame row by row and check if value matches
                for df_idx, value in enumerate(df[column]):
                    if value in top_values:
                        excel_row = df_idx + data_start_row
                        for col in range(1, len(df.columns) + 1):
                            ws.cell(row=excel_row, column=col).fill = fill
            else:
                # Regular condition checking
                for row_idx, value in enumerate(df[column], start=data_start_row):  # Start at data_start_row
                    if self._check_condition(value, condition):
                        # Highlight entire row
                        for col in range(1, len(df.columns) + 1):
                            ws.cell(row=row_idx, column=col).fill = fill
                        
        except Exception as e:
            logger.warning(f"⚠️ Row highlighting failed: {e}")
    
    def _highlight_cells(self, ws, df: pd.DataFrame, highlight_spec: Dict[str, Any], data_start_row: int = 2):
        """Highlight specific cells based on conditions"""
        try:
            column = highlight_spec.get('column')
            condition = highlight_spec.get('condition')
            color = highlight_spec.get('color', 'yellow')
            
            if not column or column not in df.columns:
                return
            
            # Get color code
            fill_color = self.COLORS.get(color, self.COLORS['yellow'])
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            # Find matching cells
            col_idx = df.columns.get_loc(column) + 1
            
            # Handle topN/bottomN operators specially (case-insensitive)
            operator_str = condition.get('operator', '').upper().replace('_', '') if isinstance(condition, dict) else ''
            if operator_str in ['TOPN', 'BOTTOMN']:
                n = condition.get('value', 5)
                
                # Get top/bottom N values
                if 'TOP' in operator_str:
                    top_values = df.nlargest(n, column)[column].values
                else:  # BOTTOM
                    top_values = df.nsmallest(n, column)[column].values
                
                # Highlight cells where the column value is in top/bottom N values
                # We iterate through the DataFrame row by row and check if value matches
                for df_idx, value in enumerate(df[column]):
                    if value in top_values:
                        excel_row = df_idx + data_start_row
                        ws.cell(row=excel_row, column=col_idx).fill = fill
            else:
                # Regular condition checking
                for row_idx, value in enumerate(df[column], start=data_start_row):  # Start at data_start_row
                    if self._check_condition(value, condition):
                        ws.cell(row=row_idx, column=col_idx).fill = fill
                        
        except Exception as e:
            logger.warning(f"⚠️ Cell highlighting failed: {e}")
    
    def _apply_conditional_formatting(self, ws, df: pd.DataFrame, format_spec: Dict[str, Any], data_start_row: int = 2):
        """Apply conditional formatting rules"""
        try:
            column = format_spec.get('column')
            rules = format_spec.get('rules', [])
            
            if not column or column not in df.columns:
                return
            
            col_idx = df.columns.get_loc(column) + 1
            
            for row_idx, value in enumerate(df[column], start=data_start_row):
                for rule in rules:
                    condition = rule.get('condition')
                    color = rule.get('color', 'yellow')
                    
                    if self._check_condition(value, condition):
                        fill_color = self.COLORS.get(color, self.COLORS['yellow'])
                        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                        ws.cell(row=row_idx, column=col_idx).fill = fill
                        break  # Apply first matching rule only
                        
        except Exception as e:
            logger.warning(f"⚠️ Conditional formatting failed: {e}")
    
    def _highlight_duplicates(self, ws, df: pd.DataFrame, dup_spec: Dict[str, Any]):
        """Highlight duplicate values in specified columns"""
        try:
            columns = dup_spec.get('columns', [])
            color = dup_spec.get('color', 'light_red')
            
            if not columns:
                columns = df.columns.tolist()
            
            fill_color = self.COLORS.get(color, self.COLORS['light_red'])
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            for column in columns:
                if column not in df.columns:
                    continue
                
                col_idx = df.columns.get_loc(column) + 1
                duplicates = df[column].duplicated(keep=False)
                
                for row_idx, is_dup in enumerate(duplicates, start=2):
                    if is_dup:
                        ws.cell(row=row_idx, column=col_idx).fill = fill
                        
        except Exception as e:
            logger.warning(f"⚠️ Duplicate highlighting failed: {e}")
    
    def _highlight_nulls(self, ws, df: pd.DataFrame, null_spec: Dict[str, Any]):
        """Highlight null/missing values"""
        try:
            columns = null_spec.get('columns', [])
            color = null_spec.get('color', 'gray')
            
            if not columns:
                columns = df.columns.tolist()
            
            fill_color = self.COLORS.get(color, self.COLORS['gray'])
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            for column in columns:
                if column not in df.columns:
                    continue
                
                col_idx = df.columns.get_loc(column) + 1
                
                for row_idx, value in enumerate(df[column], start=2):
                    if pd.isna(value):
                        ws.cell(row=row_idx, column=col_idx).fill = fill
                        
        except Exception as e:
            logger.warning(f"⚠️ Null highlighting failed: {e}")
    
    def _check_condition(self, value, condition: Any) -> bool:
        """Check if value meets condition"""
        try:
            if isinstance(condition, dict):
                operator = condition.get('operator', '==')
                target = condition.get('value')
                
                # Handle NaN values
                if pd.isna(value):
                    return operator == 'is_null'
                
                if operator == '==':
                    return value == target
                elif operator == '!=':
                    return value != target
                elif operator == '>':
                    return float(value) > float(target)
                elif operator == '<':
                    return float(value) < float(target)
                elif operator == '>=':
                    return float(value) >= float(target)
                elif operator == '<=':
                    return float(value) <= float(target)
                elif operator == 'contains':
                    return str(target).lower() in str(value).lower()
                elif operator == 'in':
                    return value in target
                elif operator == 'is_null':
                    return pd.isna(value)
                elif operator == 'date_in_month':
                    # Check if date falls within a specific month
                    # target should be in format "YYYY-MM" or "MM" or month name
                    return self._check_date_in_month(value, target)
                elif operator == 'date_in_previous_month':
                    # Check if date is in the previous month from a reference date
                    # target can be a reference date or None (uses current date)
                    return self._check_date_in_previous_month(value, target)
            else:
                # Simple equality check
                return value == condition
        except Exception:
            return False
    
    def _check_date_in_month(self, value, month_spec: str) -> bool:
        """Check if a date value falls within a specific month"""
        try:
            from datetime import datetime
            
            # Try to parse the value as a date
            if isinstance(value, str):
                # Try common date formats
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d', '%d-%m-%Y']:
                    try:
                        date_val = datetime.strptime(value, fmt)
                        break
                    except:
                        continue
                else:
                    # If no format worked, try string matching
                    return str(month_spec) in str(value)
            elif hasattr(value, 'month'):  # datetime object
                date_val = value
            else:
                return False
            
            # Check if month matches
            if '-' in str(month_spec):  # Format: "YYYY-MM" or "MM"
                parts = str(month_spec).split('-')
                if len(parts) == 2:  # YYYY-MM
                    return date_val.year == int(parts[0]) and date_val.month == int(parts[1])
                else:  # MM
                    return date_val.month == int(month_spec)
            else:
                # Just month number or name
                try:
                    month_num = int(month_spec)
                    return date_val.month == month_num
                except:
                    # Month name matching
                    month_names = ['january', 'february', 'march', 'april', 'may', 'june',
                                   'july', 'august', 'september', 'october', 'november', 'december']
                    month_spec_lower = str(month_spec).lower()
                    if month_spec_lower in month_names:
                        return date_val.month == month_names.index(month_spec_lower) + 1
                    return False
        except Exception:
            return False
    
    def _check_date_in_previous_month(self, value, reference_date=None) -> bool:
        """Check if a date is in the previous month relative to a reference date"""
        try:
            from datetime import datetime
            from dateutil.relativedelta import relativedelta
            
            # Parse the value as a date
            if isinstance(value, str):
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d', '%d-%m-%Y']:
                    try:
                        date_val = datetime.strptime(value, fmt)
                        break
                    except:
                        continue
                else:
                    return False
            elif hasattr(value, 'month'):
                date_val = value
            else:
                return False
            
            # Get reference date (default to current date)
            if reference_date:
                if isinstance(reference_date, str):
                    ref_date = datetime.strptime(reference_date, '%Y-%m-%d')
                else:
                    ref_date = reference_date
            else:
                ref_date = datetime.now()
            
            # Calculate previous month
            prev_month_date = ref_date - relativedelta(months=1)
            
            # Check if date_val is in the previous month
            return (date_val.year == prev_month_date.year and 
                    date_val.month == prev_month_date.month)
        except Exception:
            return False
    
    def _apply_subtotals(self, ws, df: pd.DataFrame, subtotal_spec: Dict[str, Any]):
        """Apply subtotals with formulas at group changes
        
        How it works (Excel-style):
        1. Data MUST be sorted by group_by column first
        2. Subtotal row inserted at each change in group_by value
        3. SUBTOTAL formula counts/sums the aggregate_column for that group
        4. Subtotal rows are formatted (bold, gray background)
        """
        try:
            group_by = subtotal_spec.get('group_by')
            aggregate_column = subtotal_spec.get('aggregate_column')
            function = subtotal_spec.get('function', 'count').upper()
            
            if not group_by or group_by not in df.columns:
                logger.warning(f"⚠️ Subtotal group_by column '{group_by}' not found")
                return
            
            if not aggregate_column or aggregate_column not in df.columns:
                logger.warning(f"⚠️ Subtotal aggregate_column '{aggregate_column}' not found")
                return
            
            # Get column indices (1-indexed for Excel)
            group_col_idx = df.columns.get_loc(group_by) + 1
            agg_col_idx = df.columns.get_loc(aggregate_column) + 1
            
            # CRITICAL: Track groups with their start and end rows
            groups = []
            current_group = None
            group_start_row = 2  # Data starts at row 2 (after header)
            
            for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (after header)
                group_value = ws.cell(row=row_idx, column=group_col_idx).value
                
                # Check if group changed
                if current_group is not None and group_value != current_group:
                    # Save the previous group's range
                    groups.append({
                        'name': current_group,
                        'start_row': group_start_row,
                        'end_row': row_idx - 1,
                        'insert_at': row_idx  # Insert subtotal at this row
                    })
                    group_start_row = row_idx
                
                current_group = group_value
            
            # Add final group
            if current_group is not None:
                groups.append({
                    'name': current_group,
                    'start_row': group_start_row,
                    'end_row': ws.max_row,
                    'insert_at': ws.max_row + 1
                })
            
            logger.info(f"📊 Found {len(groups)} groups for subtotals")
            
            # Insert subtotal rows (in reverse to maintain row numbers)
            for group in reversed(groups):
                insert_row = group['insert_at']
                group_name = group['name']
                start_row = group['start_row']
                end_row = group['end_row']
                
                # Insert new row
                ws.insert_rows(insert_row)
                
                # Add subtotal label in the group_by column
                ws.cell(row=insert_row, column=group_col_idx, value=f"{group_name} Subtotal")
                
                # SUBTOTAL function: 3=COUNTA (count non-empty), 9=SUM, 1=AVERAGE
                func_code = 3 if function == 'COUNT' else (9 if function == 'SUM' else 1)
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(agg_col_idx)
                
                # Calculate the actual value from the worksheet for immediate display
                # This ensures the count shows immediately without needing Excel to calculate
                if function == 'COUNT':
                    # Count non-empty values in the range
                    count_value = sum(1 for r in range(start_row, end_row + 1) 
                                     if ws.cell(row=r, column=agg_col_idx).value is not None)
                    calculated_value = count_value
                elif function == 'SUM':
                    # Sum numeric values in the range
                    sum_value = sum(float(ws.cell(row=r, column=agg_col_idx).value or 0) 
                                   for r in range(start_row, end_row + 1))
                    calculated_value = sum_value
                else:  # AVERAGE
                    # Average numeric values in the range
                    values = [float(ws.cell(row=r, column=agg_col_idx).value or 0) 
                             for r in range(start_row, end_row + 1)]
                    calculated_value = sum(values) / len(values) if values else 0
                
                # Set the cell with the calculated numeric value
                # This displays the count/sum immediately (like 17, 1, 5, 3 in the image)
                cell = ws.cell(row=insert_row, column=agg_col_idx)
                cell.value = calculated_value
                
                # Format the cell as a number
                cell.number_format = '0' if function == 'COUNT' else '0.00'
                
                # Format subtotal row (bold, gray background)
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=insert_row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='FFE0E0E0', end_color='FFE0E0E0', fill_type='solid')
                
                logger.info(f"  ➕ Subtotal for '{group_name}': {calculated_value} items (rows {start_row}-{end_row})")
            
            logger.info(f"✅ Applied {len(groups)} subtotals")
            
        except Exception as e:
            logger.error(f"⚠️ Subtotal application failed: {e}", exc_info=True)
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + settings.EXCEL_COLUMN_WIDTH_PADDING, settings.EXCEL_MAX_COLUMN_WIDTH)
                ws.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.warning(f"⚠️ Column auto-adjustment failed: {e}")


class ExcelOperationIntentParser:
    """
    Parses user intent and generates Excel operation specifications
    """
    
    def __init__(self, llm_model):
        self.model = llm_model
    
    def parse_intent(self, user_query: str, df: pd.DataFrame, conversation_history: list = None, last_generated_file: dict = None) -> Tuple[bool, Optional[Dict[str, Any]]]:
        """
        Parse user query to determine if it requires Excel file operations
        Uses pure LLM agentic approach - no hardcoded keywords
        
        Args:
            user_query: The user's current query
            df: The DataFrame to operate on
            conversation_history: Previous conversation turns for context (optional)
            last_generated_file: Info about last generated file for sequential operations (optional)
        
        Returns:
            Tuple of (is_excel_operation, operation_spec)
        """
        
        # Use LLM to intelligently determine if this is an Excel operation request
        is_excel_operation = self._detect_excel_intent_with_llm(user_query, df, conversation_history, last_generated_file)
        
        if is_excel_operation:
            # Use LLM to parse the specific operations
            operations = self._parse_operations_with_llm(user_query, df, conversation_history, last_generated_file)
            
            if operations:
                return True, operations
        
        return False, None
    
    def _detect_excel_intent_with_llm(self, user_query: str, df: pd.DataFrame, conversation_history: list = None, last_generated_file: dict = None) -> bool:
        """
        Use LLM to intelligently detect if user wants an Excel file operation
        Pure agentic approach - no hardcoded keywords
        """
        
        # Build conversation context if available
        history_context = ""
        if conversation_history and len(conversation_history) > 0:
            history_context = "\n**Recent Conversation History**:\n"
            # Include last 3 conversations for context
            recent = conversation_history[-3:] if len(conversation_history) > 3 else conversation_history
            for idx, conv in enumerate(recent, 1):
                user_q = conv.get("user_query", "")
                history_context += f"{idx}. User asked: \"{user_q}\"\n"
            history_context += "\n**IMPORTANT**: If the conversation history shows the user was discussing Excel operations (delete rows, filter, etc.) and now asks for \"the file\" or \"give me the file\", they are requesting an EXCEL FILE OPERATION.\n"
        
        intent_detection_prompt = f"""
You are an intelligent intent classifier for a data analysis system. Your task is to determine if the user wants an EXCEL FILE OPERATION or just a TEXT ANALYSIS.

**User Query**: "{user_query}"

**Available Data Context**:
- DataFrame with {df.shape[0]} rows and {df.shape[1]} columns
- Columns: {list(df.columns)}
{history_context}

## Understanding User Intent:

### EXCEL FILE OPERATION (return "YES")
User wants a downloadable Excel file when they:
- Request a "separate file" or "new file" with specific data
- Ask to "download", "export", or "save" data
- Want to "highlight", "color code", or "format" data visually
- Request filtered/sorted data as a file (not just to see the results)
- Say things like "give me a file", "create an Excel", "can you give the file"
- Ask for data "where [condition]" and imply they want it as a file
- Request operations that make sense in Excel (highlighting, conditional formatting, etc.)
- Want to "merge", "combine", "join" multiple files into one Excel file
- Ask for "single file with multiple tabs/sheets"
- Request to "consolidate" or "put files together"
- Want to "rename" the file or specify a custom filename
- Say "rename the file to X", "call it X", "name it X"

Examples of EXCEL FILE requests:
- "Give me a separate file where status is Active"
- "Can you give the file with only Sales department"
- "Highlight rows where Age > 30 in yellow"
- "Download employees with Salary > 60000"
- "Create an Excel file with filtered data"
- "Export the Active records"
- "I need a file showing only IT department"
- "Show me the Invited status records" (when context suggests they want a file)
- "Merge all files into one Excel with different tabs"
- "Give me a single file with all data in separate sheets"
- "Combine the uploaded files into one spreadsheet"
- "Put all files together in one Excel file with tabs"
- "Rename the file into app" or "Rename the file to app"
- "Call it Sales_Report" or "Name the file Dashboard"
- "now give me the correct file" (follow-up request for Excel file)
- "give me the excel file" (explicit request for Excel file)
- "where is the file" (asking for Excel file)
- "can you give me the file" (requesting Excel file)
- "I need the file" (requesting Excel file)
- "generate the file" (requesting Excel file)

**IMPORTANT FOR FOLLOW-UP REQUESTS:**
If the user is asking for "the file", "correct file", "excel file", or similar, they are requesting an EXCEL FILE OPERATION, even if the query is short or vague. These are continuation requests where the user expects an Excel file to be generated.

### TEXT ANALYSIS (return "NO")
User wants analysis, insights, or information when they:
- Ask "how many", "what is", "tell me about"
- Request statistics, summaries, or insights
- Want to understand patterns or trends
- Ask questions that need explanations
- Request calculations or aggregations without mentioning file/download
- Request charts, graphs, or visualizations (these are displayed in chat, not Excel files)

Examples of TEXT ANALYSIS requests:
- "How many rows are there?"
- "What is the average salary?"
- "Tell me about the data distribution"
- "Analyze the trends in sales"
- "What insights can you provide?"
- "Show me statistics" (just wants to see numbers, not a file)
- "I need a line chart" or "show me a pie chart" (chart requests for display)
- "Create a graph" or "make a visualization" (chart requests for display)
- "I need the line graph" or "give me a bar chart" (chart requests for display)

## Your Task:
Analyze the user's query and determine their TRUE INTENT. Consider:
1. Does the phrasing suggest they want a FILE or just INFORMATION?
2. Are they asking for data manipulation that requires Excel (highlighting, formatting)?
3. Do they use language that implies downloading/exporting?
4. Would the answer be better as a file or as text?

**IMPORTANT**: Be intelligent and context-aware. Sometimes "show me" means "give me a file" and sometimes it means "tell me about". Use your understanding of natural language to determine the true intent.

Respond with ONLY one word: "YES" (for Excel file operation) or "NO" (for text analysis)
"""
        
        try:
            import google.generativeai as genai
            response = self.model.generate_content(
                intent_detection_prompt,
                generation_config=genai.types.GenerationConfig(
                    temperature=0.1,  # Low temperature for consistent classification
                    top_p=0.8,
                    top_k=10
                )
            )
            
            result = response.text.strip().upper()
            
            if "YES" in result:
                logger.info(f"🎯 LLM detected EXCEL FILE operation for query: '{user_query[:50]}...'")
                return True
            else:
                logger.info(f"📝 LLM detected TEXT ANALYSIS for query: '{user_query[:50]}...'")
                return False
                
        except Exception as e:
            logger.error(f"❌ Error in LLM intent detection: {e}")
            # Conservative fallback - assume text analysis
            return False
    
    def _parse_operations_with_llm(self, user_query: str, df: pd.DataFrame, conversation_history: list = None, last_generated_file: dict = None) -> Optional[Dict[str, Any]]:
        """Use LLM to understand user intent and generate operation specifications"""
        
        # Build conversation context if available
        history_context = ""
        previous_operations = None
        if conversation_history and len(conversation_history) > 0:
            history_context = "\n**Recent Conversation History**:\n"
            # Include last 3 conversations for context
            recent = conversation_history[-3:] if len(conversation_history) > 3 else conversation_history
            for idx, conv in enumerate(recent, 1):
                user_q = conv.get("user_query", "")
                ops = conv.get("operations", {})
                history_context += f"{idx}. User asked: \"{user_q}\"\n"
                if ops:
                    history_context += f"   Operations applied: {ops}\n"
                    previous_operations = ops  # Store last operations
            
            history_context += "\n**IMPORTANT INSTRUCTIONS FOR MERGING OPERATIONS**:\n"
            history_context += "- If user says \"add\", \"also\", \"and\", \"additionally\", \"add another\", \"add on that\" - this means MERGE with previous operations\n"
            history_context += "- When merging highlights: Combine ALL previous highlight_cells/highlight_rows into an array with the new one\n"
            history_context += "- Example: If previous had green highlighting, and user says 'add red', output should have BOTH green AND red\n"
            history_context += "- CRITICAL: Look at 'Operations already applied' in Last Generated File Context to see what to merge with\n"
        
        # Add last generated file context
        file_context = ""
        if last_generated_file:
            file_context = f"\n**Last Generated File Context**:\n"
            file_context += f"- Filename: {last_generated_file.get('filename')}\n"
            file_context += f"- Type: {last_generated_file.get('type')}\n"
            file_context += f"- Operations already applied: {last_generated_file.get('operations', {})}\n"
            if last_generated_file.get('sheets'):
                file_context += f"- Sheets: {last_generated_file.get('sheets')}\n"
            file_context += f"- Generated: {last_generated_file.get('timestamp')}\n"
            file_context += f"\n**CRITICAL - SEQUENTIAL OPERATIONS**:\n"
            file_context += f"If user says \"take the sorted file\", \"in the sorted file\", \"the highlighted file\", \"the file\", or \"the merged file\",\n"
            file_context += f"they are referring to this last generated file. This is a SEQUENTIAL OPERATION on the existing file.\n"
            file_context += f"\n**IMPORTANT**: When user refers to \"the sorted file\" or \"the filtered file\", they mean:\n"
            file_context += f"- The data is ALREADY sorted/filtered in the last generated file\n"
            file_context += f"- You should NOT add sort/filter operations again UNLESS user explicitly requests a NEW sort/filter\n"
            file_context += f"- Only parse the NEW operations the user is requesting (e.g., highlighting, formatting)\n"
            file_context += f"- The DataFrame you receive already has the previous operations applied\n"
            file_context += f"\nDefault behavior: Work on the LAST GENERATED FILE unless user explicitly mentions original file names.\n"
        
        prompt = f"""
You are an Excel operations parser. Analyze the user's request and generate a JSON specification for Excel file operations.

**CRITICAL INSTRUCTION - READ CAREFULLY**: 
You are a highly intelligent LLM that understands complex multi-step Excel operations. Your task is to:

1. **READ THE ENTIRE USER QUERY** - Don't stop at the first operation
2. **COUNT THE OPERATIONS** - Identify how many distinct operations the user is requesting
3. **PARSE EACH OPERATION INDIVIDUALLY** - Analyze what each operation is asking for
4. **SEPARATE OPERATIONS** - Each operation should be called individually, not combined
5. **USE ARRAY FORMAT** - When multiple operations of the SAME type (e.g., multiple highlights), use array format
6. **PRESERVE ORDER** - Operations should be in the same order as user's request

**OPERATION SEPARATION RULES:**
- Multiple highlights on DIFFERENT columns → Use array: {{"highlight_rows": [...]}}
- Multiple highlights with DIFFERENT conditions → Use array: {{"highlight_rows": [...]}}
- Multiple highlights with DIFFERENT colors → Use array: {{"highlight_rows": [...]}}
- Filtering + Highlighting → Separate operations: {{"filter": ..., "highlight_rows": ...}}
- Sorting + Highlighting → Separate operations: {{"sort": ..., "highlight_rows": ...}}

**EXAMPLES OF OPERATION COUNTING:**
- "Highlight Active in green and Pending in yellow" → 2 highlight operations (use array)
- "Filter Status=Active and highlight in green" → 2 operations (filter + highlight)
- "Sort by Sales, filter Age>30, highlight top 5" → 3 operations (sort + filter + highlight)
- "Highlight duplicates in Email in red and nulls in Name in gray" → 2 operations (use arrays)

**CRITICAL - HIGHLIGHTING vs SORTING:**
- If user says "highlight top 5" or "highlight rows for top 5" → ONLY add highlight_rows with topN operator
- DO NOT add sorting unless user explicitly says "sort" or "order by"
- The topN/bottomN operator in highlighting will find the top/bottom values WITHOUT changing row order
- Example: "highlight top 5 by Sales" → ONLY {{"highlight_rows": ...}}, NO {{"sort": ...}}

**CRITICAL - MERGE DETECTION:**
- If step 1 says "files are moved onto one spreadsheet" or "combine files" or "merge files" → This is a MERGE operation
- Even if worded passively (e.g., "Output_Learners and Output_Applicants are moved onto one Spreadsheet"), recognize it as merge_files: true
- Don't skip the merge operation - include it as the FIRST operation in the operations array

**NO HARDCODED LIMITS**: User may provide 3 steps, 8 steps, 14 steps, or 20 steps. Parse ALL of them.
**NO ASSUMPTIONS**: Don't assume a certain number of operations. Parse what the user actually requests.
**INTELLIGENT UNDERSTANDING**: Understand the intent behind each step, even if worded differently.

**User Query**: "{user_query}"

**Available DataFrame Columns**: {list(df.columns)}
**DataFrame Shape**: {df.shape[0]} rows × {df.shape[1]} columns
**Sample Data**:
{df.head(3).to_string()}
{history_context}
{file_context}

## IMPORTANT: Multi-Operation Format (CRITICAL - READ CAREFULLY)

**AGENTIC APPROACH**: Generate operations in the FINAL format that will be used directly.

### Format Rules:

**1. Multiple operations of the SAME type → Use ARRAY format:**
```json
{{
  "highlight_rows": [
    {{"column": "Progress", "condition": {{"operator": "==", "value": 0.0}}, "color": "red"}},
    {{"column": "Progress", "condition": {{"operator": "==", "value": 1.0}}, "color": "green"}}
  ]
}}
```

**2. Multiple operations of DIFFERENT types → Use SEPARATE keys:**
```json
{{
  "highlight_rows": [{{"column": "Status", "condition": {{...}}, "color": "yellow"}}],
  "freeze_panes": {{"row": 1, "col": 1}},
  "sort": {{"by": ["Status"], "ascending": true}}
}}
```

**3. Single operation → Use DIRECT format:**
```json
{{
  "highlight_rows": {{"column": "Status", "condition": {{...}}, "color": "yellow"}}
}}
```

**4. Multi-sheet operations → Use "operations" array ONLY for sheet-specific operations:**
```json
{{
  "operations": [
    {{"apply_to_all_sheets": true, "remove_last_row": true}},
    {{"apply_to_all_sheets": true, "freeze_panes": {{"row": 1, "col": 1}}}},
    {{"target_sheet": "Applicants", "sort": {{"by": ["Status"], "ascending": true}}}},
    {{"target_sheet": "Applicants", "delete_rows": {{"column": "Status", "condition": {{...}}}}}}
  ]
}}
```

**CRITICAL**: Use "operations" array ONLY when operations target different sheets or need apply_to_all_sheets flag.
For single-sheet operations with multiple highlights/formats, use array format within the operation key.

## Supported Operations:

### 1. Filtering
```json
{{"filter": {{"column_name": {{"operator": "==|!=|>|<|>=|<=|contains|in|date_in_month|date_in_previous_month", "value": "target_value"}}}}}}
```

**Date Operators:**
- `date_in_month`: Check if date is in a specific month. Value can be "YYYY-MM", "MM", or month name (e.g., "September", "09", "2025-09")
- `date_in_previous_month`: Check if date is in the previous month. Value can be a reference date or null (uses current date)

**Examples:**
- Highlight dates in September 2025: `{{"operator": "date_in_month", "value": "2025-09"}}`
- Highlight dates in previous month: `{{"operator": "date_in_previous_month", "value": null}}`
- Highlight dates in month 9: `{{"operator": "date_in_month", "value": "09"}}`

### 2. Sorting
```json
{{"sort": {{"by": ["column1", "column2"], "ascending": true}}}}
```

### 3. Highlighting Rows (Single or Multiple)
**Single highlight:**
```json
{{"highlight_rows": {{"column": "column_name", "condition": {{"operator": "==", "value": "target"}}, "color": "yellow|red|green|blue|orange"}}}}
```
**Multiple highlights (different conditions/colors):**
```json
{{"highlight_rows": [
  {{"column": "Status", "condition": {{"operator": "==", "value": "Active"}}, "color": "yellow"}},
  {{"column": "Status", "condition": {{"operator": "==", "value": "Withdrawn"}}, "color": "red"}}
]}}
```

**IMPORTANT - Percentage Values:**
- If user mentions percentages (e.g., "0%", "100%"), convert to decimal format
- Examples:
  - "highlight progress is 0%" → value: 0.0 or 0
  - "highlight progress is 100%" → value: 1.0 or 1
  - "highlight progress is 50%" → value: 0.5
- Check the DataFrame schema to see if the column stores percentages as decimals (0.0-1.0) or as numbers (0-100)

### 4. Highlighting Cells (Single or Multiple)
**Single highlight:**
```json
{{"highlight_cells": {{"column": "column_name", "condition": {{"operator": ">", "value": 100}}, "color": "yellow"}}}}
```
**Multiple highlights:**
```json
{{"highlight_cells": [
  {{"column": "Score", "condition": {{"operator": ">", "value": 90}}, "color": "green"}},
  {{"column": "Score", "condition": {{"operator": "<", "value": 50}}, "color": "red"}}
]}}
```

### 5. Conditional Formatting (Single or Multiple)
**Single column:**
```json
{{"conditional_format": {{"column": "column_name", "rules": [{{"condition": {{"operator": ">", "value": 50}}, "color": "green"}}, {{"condition": {{"operator": "<", "value": 30}}, "color": "red"}}]}}}}
```
**Multiple columns:**
```json
{{"conditional_format": [
  {{"column": "Sales", "rules": [{{"condition": {{"operator": ">", "value": 1000}}, "color": "green"}}]}},
  {{"column": "Status", "rules": [{{"condition": {{"operator": "==", "value": "Pending"}}, "color": "yellow"}}]}}
]}}
```

### 6. Highlight Duplicates (Single or Multiple)
**Single specification:**
```json
{{"highlight_duplicates": {{"columns": ["column1", "column2"], "color": "light_red"}}}}
```
**Multiple specifications:**
```json
{{"highlight_duplicates": [
  {{"columns": ["Email"], "color": "red"}},
  {{"columns": ["Phone"], "color": "yellow"}}
]}}
```

### 7. Highlight Nulls (Single or Multiple)
**Single specification:**
```json
{{"highlight_nulls": {{"columns": ["column1"], "color": "gray"}}}}
```
**Multiple specifications:**
```json
{{"highlight_nulls": [
  {{"columns": ["Email"], "color": "red"}},
  {{"columns": ["Phone", "Address"], "color": "yellow"}}
]}}
```

### 8. Top/Bottom N (Data Extraction)
```json
{{"top_n": 10}}  or  {{"bottom_n": 10}}
```
**IMPORTANT**: This operation EXTRACTS only the top/bottom N rows from the data.
**Use this ONLY when user wants to "get", "show", "give me", "extract" the top/bottom N rows.**

**For HIGHLIGHTING top/bottom N rows, use the topN/bottomN operator in highlight_rows:**
```json
{{"highlight_rows": {{"column": "Sales", "condition": {{"operator": "topN", "value": 5}}, "color": "green"}}}}
```

**CRITICAL DISTINCTION:**
- "Give me top 5 rows" → Use {{"top_n": 5}} (extracts data)
- "Highlight top 5 rows" → Use {{"highlight_rows": {{"condition": {{"operator": "topN", "value": 5}}}}}} (highlights in place)
- "Highlight top 5 by Sales" → Use {{"highlight_rows": {{"column": "Sales", "condition": {{"operator": "topN", "value": 5}}, "color": "green"}}}}

**DO NOT add sorting** when user only asks to highlight. The topN/bottomN operator will automatically find the top/bottom values WITHOUT changing row order.

### 9. Merge Multiple Files (Multi-Sheet Excel)
```json
{{"merge_files": true, "sheet_names": ["SheetName1", "SheetName2"]}}
```
**Use this when user wants to combine multiple uploaded files into one Excel file with separate tabs/sheets.**
**If user specifies custom tab/sheet names, include them in the sheet_names array in the order of uploaded files.**
**If no custom names specified, use sheet_names: null or omit it.**

**IMPORTANT - MERGE DETECTION:**
Recognize merge operations even when worded as descriptions or passive voice:
- "Files are moved onto one spreadsheet" → merge_files: true
- "Output_Learners and Output_Applicants are moved onto one spreadsheet" → merge_files: true
- "Combine files into one workbook" → merge_files: true
- "Put both files together" → merge_files: true
- "Merge these files" → merge_files: true

**Examples:**
- "Merge these files into one" → {{"merge_files": true}}
- "Combine into single Excel with separate sheets" → {{"merge_files": true}}
- "Put both files in one workbook" → {{"merge_files": true}}
- "Output_Learners and Output_Applicants are moved onto one Spreadsheet with two different tabs" → {{"merge_files": true, "sheet_names": ["Learners", "Applicants"]}}

### 10. Target Specific Sheet (Multi-Sheet Excel)
```json
{{"target_sheet": "SheetName"}}
```
**Use this when user wants to work with a specific sheet from a multi-sheet Excel file.**
**Examples:**
- "Delete rows from the Learners sheet" → {{"target_sheet": "Learners", "delete_rows": ...}}
- "Filter the Applicants tab" → {{"target_sheet": "Applicants", "filter": ...}}
- "Highlight rows in the second sheet" → {{"target_sheet": "Sheet2", "highlight_rows": ...}}
**Can be combined with any other operation to apply it to a specific sheet.**

### 11. Apply to All Sheets (Multi-Sheet Excel)
```json
{{"apply_to_all_sheets": true}}
```
**Use this when user wants to apply an operation to ALL sheets in a multi-sheet Excel file.**
**Examples:**
- "Delete last row from both sheets" → {{"apply_to_all_sheets": true, "remove_last_row": true}}
- "Filter both spreadsheets where Status is Active" → {{"apply_to_all_sheets": true, "filter": ...}}
- "Highlight rows in all sheets" → {{"apply_to_all_sheets": true, "highlight_rows": ...}}
**This preserves the multi-sheet structure and applies the operation to each sheet individually.**
**Output will be a multi-sheet Excel file with the operation applied to all sheets.**

### 12. Custom Filename
```json
{{"custom_filename": "desired_name"}}
```
**Use this when user wants to rename the file or specify a custom filename.**
**Extract just the name without extension (e.g., "app", "Sales_Report", "Dashboard").**
**Can be combined with any other operation.**

### 12. Remove Last Row (Delete Filter Description)
```json
{{"remove_last_row": true}}
```
**Use this when user wants to:**
- Delete the last row
- Remove filter description row
- Remove description that was added to the bottom
- Clean up the final row containing filter information
**This removes the last row from the DataFrame before creating the Excel file.**

### 12. Delete Specific Rows
```json
{{"delete_rows": {{"row_numbers": [5, 10, 15]}}}}
```
**Or delete by condition:**
```json
{{"delete_rows": {{"column": "Status", "condition": {{"operator": "==", "value": "Empty"}}}}}}
```
**Use this when user wants to:**
- Delete specific row numbers (e.g., "delete row 5", "delete rows 10 to 15")
- Delete rows matching a condition (e.g., "delete rows where Status is Empty")
- Remove unwanted data rows

### 13. Freeze Panes
```json
{{"freeze_panes": {{"row": 1, "col": 0}}}}
```
**Parameters:**
- `row`: Number of rows to freeze (0-indexed, so 1 = freeze top row)
- `col`: Number of columns to freeze (0-indexed, so 1 = freeze first column)

**Examples:**
- Freeze top row only: `{{"freeze_panes": {{"row": 1, "col": 0}}}}`
- Freeze top row and first column: `{{"freeze_panes": {{"row": 1, "col": 1}}}}`
- Freeze top 2 rows: `{{"freeze_panes": {{"row": 2, "col": 0}}}}`
- Freeze top row and first 2 columns: `{{"freeze_panes": {{"row": 1, "col": 2}}}}`

**Use this when user wants to:**
- Keep headers visible while scrolling
- Always see certain columns (like names, IDs)
- Freeze panes for better navigation

**IMPORTANT - Column Freeze Detection:**
- If user mentions "always see [column name]" or "keep [column name] visible" → set col to 1 or higher
- Examples:
  - "freeze to see top row and learner name" → col: 1 (freeze first column which has learner name)
  - "freeze to see top row and ID" → col: 1 (freeze first column which has ID)
  - "freeze top row only" → col: 0 (no column freeze)
  - "freeze to see name and department" → col: 2 (freeze first 2 columns)

### 14. Subtotals (CRITICAL: Requires Sorting First!)
```json
{{"subtotals": {{"group_by": "Employer", "aggregate_column": "Trainer", "function": "count"}}}}
```
**For multi-sheet files, specify target_sheet:**
```json
{{"target_sheet": "Learners", "subtotals": {{"group_by": "Employer", "aggregate_column": "Trainer", "function": "count"}}}}
```
**Parameters:**
- `group_by`: Column to group by (subtotals inserted at each change in this column)
- `aggregate_column`: Column to apply the function to (the column being counted/summed)
- `function`: "count", "sum", or "average"
- `target_sheet`: (Optional) For multi-sheet files, specify which sheet to apply subtotals to

**How Subtotals Work (Excel-style):**
1. **Data MUST be sorted** by the group_by column first
2. **Subtotal row inserted** at each change in the group_by value
3. **SUBTOTAL formula** counts/sums the aggregate_column for that group
4. **Formatted rows** with bold text and gray background

**CRITICAL - AUTOMATIC SORTING:**
When user requests subtotals, you MUST include sorting:
```json
{{
  "sort": {{"by": ["Employer"], "ascending": true}},
  "subtotals": {{"group_by": "Employer", "aggregate_column": "Trainer", "function": "count"}}
}}
```

**Examples:**
- "Add subtotal for each employer counting trainers" → MUST include sort by Employer first
  ```json
  {{
    "sort": {{"by": ["Employer"], "ascending": true}},
    "subtotals": {{"group_by": "Employer", "aggregate_column": "Trainer", "function": "count"}}
  }}
  ```
- "Subtotal at each change in employer, count function on Trainer" → MUST include sort
  ```json
  {{
    "sort": {{"by": ["Employer"], "ascending": true}},
    "subtotals": {{"group_by": "Employer", "aggregate_column": "Trainer", "function": "count"}}
  }}
  ```
- "Sum sales per region with subtotals" → MUST include sort
  ```json
  {{
    "sort": {{"by": ["Region"], "ascending": true}},
    "subtotals": {{"group_by": "Region", "aggregate_column": "Sales", "function": "sum"}}
  }}
  ```

**User Query Patterns:**
- "A subtotal is input at each change in 'employer'" → Sort by Employer + Subtotals
- "Add subtotals for each department" → Sort by Department + Subtotals
- "Count function is added to Trainer" → This is the aggregate_column
- "Group by employer and count trainers" → Sort by Employer + Subtotals

**IMPORTANT:**
- **ALWAYS add sorting** when subtotals are requested
- Sort by the SAME column as group_by
- Subtotals won't work correctly without sorting
- Each subtotal row shows: "[Group Name] Subtotal" in the group_by column
- SUBTOTAL formula in the aggregate_column (e.g., =SUBTOTAL(3,C2:C10) for count)

## Examples:

**Query**: "Highlight rows where Status is 'Active'"
**Output**: 
```json
{{
  "highlight_rows": {{
    "column": "Status",
    "condition": {{"operator": "==", "value": "Active"}},
    "color": "yellow"
  }}
}}
```

**Query**: "Filter data where Age > 25 and sort by Name"
**Output**:
```json
{{
  "filter": {{
    "Age": {{"operator": ">", "value": 25}}
  }},
  "sort": {{
    "by": ["Name"],
    "ascending": true
  }}
}}
```

**Query**: "Show me top 10 records with Sales > 1000 highlighted in green"
**Output**:
```json
{{
  "filter": {{
    "Sales": {{"operator": ">", "value": 1000}}
  }},
  "top_n": 10,
  "highlight_cells": {{
    "column": "Sales",
    "condition": {{"operator": ">", "value": 1000}},
    "color": "green"
  }}
}}
```

**Query**: "Highlight Active status in yellow and Withdrawn status in red"
**Output**:
```json
{{
  "highlight_rows": [
    {{"column": "Status", "condition": {{"operator": "==", "value": "Active"}}, "color": "yellow"}},
    {{"column": "Status", "condition": {{"operator": "==", "value": "Withdrawn"}}, "color": "red"}}
  ]
}}
```

**Query**: "Give me a separate file where the status is Invited"
**Output**:
```json
{{
  "filter": {{
    "Status": {{"operator": "==", "value": "Invited"}}
  }}
}}
```

**Query**: "Can you give the separate file where Status is Active"
**Output**:
```json
{{
  "filter": {{
    "Status": {{"operator": "==", "value": "Active"}}
  }}
}}
```

**Query**: "I need only Sales department employees"
**Output**:
```json
{{
  "filter": {{
    "Department": {{"operator": "==", "value": "Sales"}}
  }}
}}
```

**Query**: "Merge all files into one Excel with different tabs"
**Output**:
```json
{{
  "merge_files": true
}}
```

**Query**: "Give me a single file with all uploaded data in separate sheets"
**Output**:
```json
{{
  "merge_files": true
}}
```

**Query**: "Highlight top 5 rows by Sales in green"
**Output**:
```json
{{
  "highlight_rows": {{
    "column": "Sales",
    "condition": {{"operator": "topN", "value": 5}},
    "color": "green"
  }}
}}
```
**NOTE**: NO sorting added - highlights top 5 values in their original positions

**Query**: "Highlight rows in green colour for the top 5 Number of Units Sold"
**Output**:
```json
{{
  "highlight_rows": {{
    "column": "Number of Units Sold",
    "condition": {{"operator": "topN", "value": 5}},
    "color": "green"
  }}
}}
```
**NOTE**: NO sorting - only highlighting the top 5 values wherever they appear
**OPERATION COUNT**: 1 operation (highlight only)

**Query**: "Highlight Active status in green, Pending in yellow, and Withdrawn in red"
**Output**:
```json
{{
  "highlight_rows": [
    {{"column": "Status", "condition": {{"operator": "==", "value": "Active"}}, "color": "green"}},
    {{"column": "Status", "condition": {{"operator": "==", "value": "Pending"}}, "color": "yellow"}},
    {{"column": "Status", "condition": {{"operator": "==", "value": "Withdrawn"}}, "color": "red"}}
  ]
}}
```
**NOTE**: 3 highlight operations on same column, different conditions - use array format
**OPERATION COUNT**: 3 highlight operations (called individually via array)

**Query**: "Filter Sales department, sort by Salary descending, and highlight top 5 in green"
**Output**:
```json
{{
  "filter": {{
    "Department": {{"operator": "==", "value": "Sales"}}
  }},
  "sort": {{
    "by": ["Salary"],
    "ascending": false
  }},
  "highlight_rows": {{
    "column": "Salary",
    "condition": {{"operator": "topN", "value": 5}},
    "color": "green"
  }}
}}
```
**NOTE**: 3 separate operations - filter, sort, highlight (each called individually)
**OPERATION COUNT**: 3 operations (filter + sort + highlight)

**Query**: "Highlight duplicates in Email in red and highlight nulls in Phone in yellow"
**Output**:
```json
{{
  "highlight_duplicates": {{
    "columns": ["Email"],
    "color": "red"
  }},
  "highlight_nulls": {{
    "columns": ["Phone"],
    "color": "yellow"
  }}
}}
```
**NOTE**: 2 different operation types - each called individually
**OPERATION COUNT**: 2 operations (duplicates + nulls)

**Query**: "Apply conditional formatting to Sales (green if >1000, red if <500) and Status (yellow if Pending)"
**Output**:
```json
{{
  "conditional_format": [
    {{"column": "Sales", "rules": [{{"condition": {{"operator": ">", "value": 1000}}, "color": "green"}}, {{"condition": {{"operator": "<", "value": 500}}, "color": "red"}}]}},
    {{"column": "Status", "rules": [{{"condition": {{"operator": "==", "value": "Pending"}}, "color": "yellow"}}]}}
  ]
}}
```
**NOTE**: 2 conditional format operations on different columns - use array format
**OPERATION COUNT**: 2 conditional format operations (called individually via array)

**Query**: "Combine the files into one spreadsheet"
**Output**:
```json
{{
  "merge_files": true
}}
```

**Query**: "A subtotal is input at each change in 'employer' a 'count' function is added to 'Trainer'"
**Output**:
```json
{{
  "sort": {{
    "by": ["Employer"],
    "ascending": true
  }},
  "subtotals": {{
    "group_by": "Employer",
    "aggregate_column": "Trainer",
    "function": "count"
  }}
}}
```
**NOTE**: MUST include sorting by Employer first, then subtotals
**OPERATION COUNT**: 2 operations (sort + subtotals)

**Query**: "Output Learners and Output Applicants are moved onto one Spreadsheet with two different tabs. Tabs are named Learners and Applicants"
**Output**:
```json
{{
  "merge_files": true,
  "sheet_names": ["Learners", "Applicants"]
}}
```

**Query**: "Merge the files with tabs called Sales and Marketing"
**Output**:
```json
{{
  "merge_files": true,
  "sheet_names": ["Sales", "Marketing"]
}}
```

**Query**: "rename the file into app"
**Output**:
```json
{{
  "custom_filename": "app"
}}
```

**Query**: "Filter Status is Active and name the file ActiveUsers"
**Output**:
```json
{{
  "filter": {{
    "Status": {{"operator": "==", "value": "Active"}}
  }},
  "custom_filename": "ActiveUsers"
}}
```

**Query**: "Give me a separate file where status is Invited, call it InvitedList"
**Output**:
```json
{{
  "filter": {{
    "Status": {{"operator": "==", "value": "Invited"}}
  }},
  "custom_filename": "InvitedList"
}}
```

**Query**: "Delete Signed Up rows from the Learners sheet"
**Output**:
```json
{{
  "target_sheet": "Learners",
  "delete_rows": {{"column": "Status", "condition": {{"operator": "==", "value": "Signed Up"}}}}
}}
```

**Query**: "Filter Active status in the Applicants tab"
**Output**:
```json
{{
  "target_sheet": "Applicants",
  "filter": {{
    "Status": {{"operator": "==", "value": "Active"}}
  }}
}}
```

**Query**: "Highlight rows where Age > 30 in the Learners sheet"
**Output**:
```json
{{
  "target_sheet": "Learners",
  "highlight_rows": {{"column": "Age", "condition": {{"operator": ">", "value": 30}}, "color": "yellow"}}
}}
```

**Query**: "Apply conditional formatting to column V with red, amber, and green colors"
**Output**:
```json
{{
  "conditional_format": {{
    "column": "V",
    "rules": [
      {{"condition": {{"operator": "<", "value": 50}}, "color": "red"}},
      {{"condition": {{"operator": ">=", "value": 50}}, "color": "orange"}},
      {{"condition": {{"operator": ">=", "value": 80}}, "color": "green"}}
    ]
  }}
}}
```
**Note**: When user doesn't specify exact conditions for red/amber/green, use reasonable defaults like <50=red, 50-79=amber/orange, >=80=green.

**Query**: "Highlight learners with a start date in the previous month in green"
**Output**:
```json
{{
  "highlight_rows": {{
    "column": "Start Date",
    "condition": {{"operator": "date_in_previous_month", "value": null}},
    "color": "green"
  }}
}}
```

**Query**: "Highlight new starters from September in green"
**Output**:
```json
{{
  "highlight_rows": {{
    "column": "Start Date",
    "condition": {{"operator": "date_in_month", "value": "09"}},
    "color": "green"
  }}
}}
```

**Query**: "Highlight dates in the reporting month (September 2025)"
**Output**:
```json
{{
  "highlight_rows": {{
    "column": "Start Date",
    "condition": {{"operator": "date_in_month", "value": "2025-09"}},
    "color": "green"
  }}
}}
```

**Query**: "Highlight cost price more than 100 in green"
**Previous Operations**: None
**Output**:
```json
{{
  "highlight_cells": {{
    "column": "Cost Price Per Unit (USD)",
    "condition": {{"operator": ">", "value": 100}},
    "color": "green"
  }}
}}
```

**Query**: "Add another operation highlight if it is less than 100 in red color"
**Previous Operations**: {{"highlight_cells": {{"column": "Cost Price Per Unit (USD)", "condition": {{"operator": ">", "value": 100}}, "color": "green"}}}}
**Output** (MERGED):
```json
{{
  "highlight_cells": [
    {{"column": "Cost Price Per Unit (USD)", "condition": {{"operator": ">", "value": 100}}, "color": "green"}},
    {{"column": "Cost Price Per Unit (USD)", "condition": {{"operator": "<", "value": 100}}, "color": "red"}}
  ]
}}
```
**Note**: When user says "add another operation", MERGE with previous operations by converting to array format.

**SEQUENTIAL OPERATIONS EXAMPLES:**

**Query**: "sort the file based on the product name"
**Last Generated File**: None
**Output**:
```json
{{
  "sort": {{"by": ["Product Name"], "ascending": true}}
}}
```

**Query**: "now take the sorted file highlight where the cost price per unit is more than 100 in green colour"
**Last Generated File**: excel_sorted.xlsx (with sort operation already applied)
**Output**:
```json
{{
  "highlight_cells": {{
    "column": "Cost Price Per Unit (USD)",
    "condition": {{"operator": ">", "value": 100}},
    "color": "green"
  }}
}}
```
**CRITICAL NOTE**: User said "take the sorted file", which means:
- The data is ALREADY sorted in the last generated file
- You should NOT add a sort operation again
- Only parse the NEW operation (highlighting)
- The system will automatically load the sorted file and apply highlighting on top of it

**MERGING MULTIPLE HIGHLIGHTS - COMPLETE WORKFLOW:**

**Step 1**: "highlight where cost price > 100 in green"
**Last Generated File Operations**: None
**Output**:
```json
{{
  "highlight_cells": {{
    "column": "Cost Price Per Unit (USD)",
    "condition": {{"operator": ">", "value": 100}},
    "color": "green"
  }}
}}
```

**Step 2**: "add on that highlight in red where the cost price per unit is less than 100"
**Last Generated File Operations**: {{"highlight_cells": {{"column": "Cost Price Per Unit (USD)", "condition": {{"operator": ">", "value": 100}}, "color": "green"}}}}
**Output** (MERGED - includes previous green AND new red):
```json
{{
  "highlight_cells": [
    {{"column": "Cost Price Per Unit (USD)", "condition": {{"operator": ">", "value": 100}}, "color": "green"}},
    {{"column": "Cost Price Per Unit (USD)", "condition": {{"operator": "<", "value": 100}}, "color": "red"}}
  ]
}}
```

**Step 3**: "if the cost price per unit is exactly 100 highlight in yellow"
**Last Generated File Operations**: {{"highlight_cells": [{{"column": "Cost Price Per Unit (USD)", "condition": {{"operator": ">", "value": 100}}, "color": "green"}}, {{"column": "Cost Price Per Unit (USD)", "condition": {{"operator": "<", "value": 100}}, "color": "red"}}]}}
**Output** (MERGED - includes previous green, red, AND new yellow):
```json
{{
  "highlight_cells": [
    {{"column": "Cost Price Per Unit (USD)", "condition": {{"operator": ">", "value": 100}}, "color": "green"}},
    {{"column": "Cost Price Per Unit (USD)", "condition": {{"operator": "<", "value": 100}}, "color": "red"}},
    {{"column": "Cost Price Per Unit (USD)", "condition": {{"operator": "==", "value": 100}}, "color": "yellow"}}
  ]
}}
```
**CRITICAL**: When user says "add", "add on that", "also", you MUST include ALL previous highlight operations in the array!
**Query**: "The filters description which is displayed as the final row is deleted in both spreadsheets"
**Output**:
```json
{{
  "apply_to_all_sheets": true,
  "remove_last_row": true
}}
```

**Query**: "Delete last row from both sheets"
**Output**:
```json
{{
  "apply_to_all_sheets": true,
  "remove_last_row": true
}}
```

**Query**: "The filters description which is displayed as the final row is deleted in both spreadsheets"
**Output**:
```json
{{
  "apply_to_all_sheets": true,
  "remove_last_row": true
}}
```

**Query**: "Delete the last row from both spreadsheets"
**Output**:
```json
{{
  "apply_to_all_sheets": true,
  "remove_last_row": true
}}
```

**Query**: "Remove filter description from both sheets"
**Output**:
```json
{{
  "apply_to_all_sheets": true,
  "remove_last_row": true
}}
```

**Query**: "Tabs are named Learners and Applicants"
**Context**: Last generated file has 2 sheets
**Output**:
```json
{{
  "rename_tabs": {{
    "sheet_0": "Learners",
    "sheet_1": "Applicants"
  }}
}}
```

**Query**: "Rename first tab to Students and second tab to Teachers"
**Output**:
```json
{{
  "rename_tabs": {{
    "sheet_0": "Students",
    "sheet_1": "Teachers"
  }}
}}
```

**Query**: "Filter Active status in all sheets"
**Output**:
```json
{{
  "apply_to_all_sheets": true,
  "filter": {{
    "Status": {{"operator": "==", "value": "Active"}}
  }}
}}
```

**Query**: "1. Output Learners and Output Applicants are moved onto one Spreadsheet with two different tabs 2. Tabs are named Learners and Applicants 3. The filters description which is displayed as the final row is deleted in both spreadsheets 4. Filters are applied to both sheets 5. Columns are fitted on both sheets 6. Panes are frozen to always see the top row and learner name on both sheets 7. Order is sorted by status on the applicants tab 8. Signed Up, Completed Edited and Signed Up Edited rows are deleted from the Applicants Tab"

**STEP-BY-STEP PARSING (8 steps total):**
- Step 1: "Output Learners and Output Applicants are moved onto one Spreadsheet with two different tabs" → **MERGE OPERATION** → Operation 1: {{"merge_files": true, "sheet_names": ["Learners", "Applicants"]}}
- Step 2: "Tabs are named Learners and Applicants" → **PART OF MERGE** (already included in Operation 1 sheet_names)
- Step 3: "The filters description which is displayed as the final row is deleted in both spreadsheets" → Operation 2: {{"apply_to_all_sheets": true, "remove_last_row": true}}
- Step 4: "Filters are applied to both sheets" → SKIP (auto-handled by Excel)
- Step 5: "Columns are fitted on both sheets" → SKIP (auto-handled by Excel)
- Step 6: "Panes are frozen to always see the top row and learner name on both sheets" → Operation 3: {{"apply_to_all_sheets": true, "freeze_panes": {{"row": 1, "col": 1}}}}
- Step 7: "Order is sorted by status on the applicants tab" → Operation 4: {{"target_sheet": "Applicants", "sort": {{"by": ["Status"], "ascending": true}}}}
- Step 8: "Signed Up, Completed Edited and Signed Up Edited rows are deleted from the Applicants Tab" → Operation 5: {{"target_sheet": "Applicants", "delete_rows": {{"column": "Status", "condition": {{"operator": "in", "value": ["Signed Up", "Completed Edited", "Signed Up Edited"]}}}}}}

**Output**:
```json
{{
  "operations": [
    {{"merge_files": true, "sheet_names": ["Learners", "Applicants"]}},
    {{"apply_to_all_sheets": true, "remove_last_row": true}},
    {{"apply_to_all_sheets": true, "freeze_panes": {{"row": 1, "col": 1}}}},
    {{"target_sheet": "Applicants", "sort": {{"by": ["Status"], "ascending": true}}}},
    {{"target_sheet": "Applicants", "delete_rows": {{"column": "Status", "condition": {{"operator": "in", "value": ["Signed Up", "Completed Edited", "Signed Up Edited"]}}}}}}
  ]
}}
```
**Result**: 5 operations (8 steps minus 2 auto-handled items and 1 merge metadata step)

**Query**: "1. Delete filter description from both sheets 2. Apply filters to both sheets 3. Fit columns on both sheets 4. Freeze panes to see top row and learner name on both sheets 5. Sort Applicants by Status 6. Delete Signed Up, Completed Edited, and Signed Up Edited rows from Applicants"

**STEP-BY-STEP PARSING:**
- Step 1: "Delete filter description from both sheets" → Operation 1: {{"apply_to_all_sheets": true, "remove_last_row": true}}
- Step 2: "Apply filters to both sheets" → SKIP (auto-handled by Excel)
- Step 3: "Fit columns on both sheets" → SKIP (auto-handled by Excel)
- Step 4: "Freeze panes to see top row and learner name on both sheets" → Operation 2: {{"apply_to_all_sheets": true, "freeze_panes": {{"row": 1, "col": 1}}}}
- Step 5: "Sort Applicants by Status" → Operation 3: {{"target_sheet": "Applicants", "sort": {{"by": ["Status"], "ascending": true}}}}
- Step 6: "Delete Signed Up, Completed Edited, and Signed Up Edited rows from Applicants" → Operation 4: {{"target_sheet": "Applicants", "delete_rows": {{"column": "Status", "condition": {{"operator": "in", "value": ["Signed Up", "Completed Edited", "Signed Up Edited"]}}}}}}

**Output**:
```json
{{
  "operations": [
    {{"apply_to_all_sheets": true, "remove_last_row": true}},
    {{"apply_to_all_sheets": true, "freeze_panes": {{"row": 1, "col": 1}}}},
    {{"target_sheet": "Applicants", "sort": {{"by": ["Status"], "ascending": true}}}},
    {{"target_sheet": "Applicants", "delete_rows": {{"column": "Status", "condition": {{"operator": "in", "value": ["Signed Up", "Completed Edited", "Signed Up Edited"]}}}}}}
  ]
}}
```
**Result**: 4 operations (6 steps minus 2 auto-handled items)

**Query**: "Delete last row from both sheets and freeze panes to see top row and first column on both sheets, then sort Applicants by Status and delete Signed Up rows from Applicants"
**Output**:
```json
{{
  "operations": [
    {{"apply_to_all_sheets": true, "remove_last_row": true}},
    {{"apply_to_all_sheets": true, "freeze_panes": {{"row": 1, "col": 1}}}},
    {{"target_sheet": "Applicants", "sort": {{"by": ["Status"], "ascending": true}}}},
    {{"target_sheet": "Applicants", "delete_rows": {{"column": "Status", "condition": {{"operator": "in", "value": ["Signed Up", "Completed Edited", "Signed Up Edited"]}}}}}}
  ]
}}
```

**Query**: "Learners with a Status of On Break in column G are highlighted yellow, learners with Withdrawal Requested are highlighted in Red"
**Output**:
```json
{{
  "target_sheet": "Learners",
  "highlight_rows": [
    {{"column": "Status", "condition": {{"operator": "==", "value": "On Break"}}, "color": "yellow"}},
    {{"column": "Status", "condition": {{"operator": "==", "value": "Withdrawal Requested"}}, "color": "red"}}
  ]
}}
```

**Query**: "Remove filter description from both sheets and sort Learners by Age"
**Output**:
```json
{{
  "operations": [
    {{"apply_to_all_sheets": true, "remove_last_row": true}},
    {{"target_sheet": "Learners", "sort": {{"by": ["Age"], "ascending": true}}}}
  ]
}}
```

**Query**: "Freeze panes on all sheets and highlight Active rows in Applicants"
**Output**:
```json
{{
  "operations": [
    {{"apply_to_all_sheets": true, "freeze_panes": {{"row": 1, "col": 0}}}},
    {{"target_sheet": "Applicants", "highlight_rows": {{"column": "Status", "condition": {{"operator": "==", "value": "Active"}}, "color": "green"}}}}
  ]
}}
```

**Query**: "1. Merge files with tabs Learners and Applicants 2. Delete last row from both sheets 3. Delete Signed Up rows from Applicants 4. Sort Applicants by Employer Site 5. Sort Learners by Start Date then Employer 6. Highlight Status On Break in yellow on Learners 7. Highlight Withdrawal Requested in red on Learners 8. Highlight September start dates in green on Learners"

**STEP-BY-STEP PARSING (8 steps):**
- Step 1: Merge → Operation 1: {{"merge_files": true, "sheet_names": ["Learners", "Applicants"]}}
- Step 2: Delete last row → Operation 2: {{"apply_to_all_sheets": true, "remove_last_row": true}}
- Step 3: Delete rows → Operation 3: {{"target_sheet": "Applicants", "delete_rows": ...}}
- Step 4: Sort Applicants → Operation 4: {{"target_sheet": "Applicants", "sort": {{"by": ["Employer Site"]}}}}
- Step 5: Sort Learners → Operation 5: {{"target_sheet": "Learners", "sort": {{"by": ["Start Date", "Employer"]}}}}
- Step 6: Highlight yellow → Operation 6: {{"target_sheet": "Learners", "highlight_rows": {{"column": "Status", "condition": {{"operator": "==", "value": "On Break"}}, "color": "yellow"}}}}
- Step 7: Highlight red → Operation 7: {{"target_sheet": "Learners", "highlight_rows": {{"column": "Status", "condition": {{"operator": "==", "value": "Withdrawal Requested"}}, "color": "red"}}}}
- Step 8: Highlight green → Operation 8: {{"target_sheet": "Learners", "highlight_rows": {{"column": "Start Date", "condition": {{"operator": "contains", "value": "September"}}, "color": "green"}}}}

**Output**:
```json
{{
  "operations": [
    {{"merge_files": true, "sheet_names": ["Learners", "Applicants"]}},
    {{"apply_to_all_sheets": true, "remove_last_row": true}},
    {{"target_sheet": "Applicants", "delete_rows": {{"column": "Status", "condition": {{"operator": "==", "value": "Signed Up"}}}}}},
    {{"target_sheet": "Applicants", "sort": {{"by": ["Employer Site"], "ascending": true}}}},
    {{"target_sheet": "Learners", "sort": {{"by": ["Start Date", "Employer"], "ascending": true}}}},
    {{"target_sheet": "Learners", "highlight_rows": {{"column": "Status", "condition": {{"operator": "==", "value": "On Break"}}, "color": "yellow"}}}},
    {{"target_sheet": "Learners", "highlight_rows": {{"column": "Status", "condition": {{"operator": "==", "value": "Withdrawal Requested"}}, "color": "red"}}}},
    {{"target_sheet": "Learners", "highlight_rows": {{"column": "Start Date", "condition": {{"operator": "contains", "value": "September"}}, "color": "green"}}}}
  ]
}}
```
**Result**: 8 operations - demonstrates parsing complex multi-step requests with different operations on different sheets

**Query**: "1. Merge Output_Learners and Output_Applicants with tabs named Learners and Applicants 2. Delete last row from both sheets 3. Freeze panes on both sheets 4. Sort Applicants by Status 5. Delete Signed Up, Completed Edited, and Signed Up Edited rows from Applicants"
**Output**:
```json
{{
  "operations": [
    {{"merge_files": true, "sheet_names": ["Learners", "Applicants"]}},
    {{"apply_to_all_sheets": true, "remove_last_row": true}},
    {{"apply_to_all_sheets": true, "freeze_panes": {{"row": 1, "col": 1}}}},
    {{"target_sheet": "Applicants", "sort": {{"by": ["Status"], "ascending": true}}}},
    {{"target_sheet": "Applicants", "delete_rows": {{"column": "Status", "condition": {{"operator": "in", "value": ["Signed Up", "Completed Edited", "Signed Up Edited"]}}}}}}
  ]
}}
```
**Note**: This is a complete workflow - merge first, then apply all operations to the merged file.

**Query**: "Sort Learners by Employer and add subtotals counting trainers for each employer"
**Output**:
```json
{{
  "target_sheet": "Learners",
  "sort": {{
    "by": ["Employer"],
    "ascending": true
  }},
  "subtotals": {{
    "group_by": "Employer",
    "aggregate_column": "Trainer",
    "function": "count"
  }}
}}
```
**NOTE**: Subtotals require sorting first, and both operations target the Learners sheet

**Query**: "The Learners Tab is then sorted by Start Date, then by Employer. A subtotal is input at each change in 'employer' a 'count' function is added to 'Trainer'"
**Output**:
```json
{{
  "operations": [
    {{"target_sheet": "Learners", "sort": {{"by": ["Start Date", "Employer"], "ascending": true}}}},
    {{"target_sheet": "Learners", "subtotals": {{"group_by": "Employer", "aggregate_column": "Trainer", "function": "count"}}}}
  ]
}}
```
**NOTE**: 2 operations on Learners sheet - sort by multiple columns, then add subtotals by Employer

**Query**: "delete the final row which contains filter description"
**Output**:
```json
{{
  "remove_last_row": true
}}
```

**Query**: "remove the last row with filter description"
**Output**:
```json
{{
  "remove_last_row": true
}}
```

**Query**: "filters description which is displayed as the final row is deleted in the spreadsheets"
**Output**:
```json
{{
  "remove_last_row": true
}}
```

**Query**: "Freeze panes to always see the top row and learner name"
**Output**:
```json
{{
  "freeze_panes": {{"row": 1, "col": 1}}
}}
```

**Query**: "Keep the header row visible while scrolling"
**Output**:
```json
{{
  "freeze_panes": {{"row": 1, "col": 0}}
}}
```

**Query**: "Freeze top row and first column"
**Output**:
```json
{{
  "freeze_panes": {{"row": 1, "col": 1}}
}}
```

**Query**: "Delete row 5"
**Output**:
```json
{{
  "delete_rows": {{"row_numbers": [5]}}
}}
```

**Query**: "Delete rows 10 to 15"
**Output**:
```json
{{
  "delete_rows": {{"row_numbers": [10, 11, 12, 13, 14, 15]}}
}}
```

**Query**: "Remove rows where Status is Empty"
**Output**:
```json
{{
  "delete_rows": {{"column": "Status", "condition": {{"operator": "==", "value": "Empty"}}}}
}}
```

**Query**: "'Signed Up', 'Completed Edited' and 'Signed Up Edited' rows are deleted from the Applicants excel"
**Output**:
```json
{{
  "delete_rows": {{"column": "Status", "condition": {{"operator": "in", "value": ["Signed Up", "Completed Edited", "Signed Up Edited"]}}}}
}}
```

**Query**: "Delete all rows with Status 'Withdrawn' or 'Cancelled'"
**Output**:
```json
{{
  "delete_rows": {{"column": "Status", "condition": {{"operator": "in", "value": ["Withdrawn", "Cancelled"]}}}}
}}
```

## CRITICAL: Filter vs Delete

**FILTER** = Keep ONLY matching rows (remove everything else)
- "Show me only Active status" → filter
- "Give me a file with just Active records" → filter
- "I want only the Active ones" → filter

**DELETE** = Remove matching rows (keep everything else)
- "Delete Active status rows" → delete_rows
- "Remove all Active records" → delete_rows
- "Get rid of Active ones" → delete_rows
- "Without Active status" → delete_rows

## Your Task:
Analyze the user query and generate the appropriate JSON specification.

**CRITICAL: HANDLE ALL OPERATIONS - NO SKIPPING**
The system now supports ALL operations. You MUST parse and include EVERY operation the user requests:
- **Subtotals** - SUPPORTED, include them
- **Auto filters** - SUPPORTED, include them  
- **Conditional formatting with multiple rules** - SUPPORTED, include them
- **Complex highlighting** - SUPPORTED, include them
- **Date-based operations** - SUPPORTED, include them

**HOW TO HANDLE PREVIOUSLY "UNSUPPORTED" OPERATIONS:**
- If user requests subtotals → Include: {{"subtotals": {{"group_by": "column_name", "aggregate_column": "column_to_count", "function": "count"}}}}
- If user requests auto filters → Include: {{"auto_filter": true}}
- If user requests multiple conditional formats on one column → Create separate highlight_cells operations for each condition
- If user requests complex date logic → Use date comparison operators or contains

**STEP-BY-STEP PARSING PROCESS:**
1. **Read the ENTIRE user query carefully**
2. **Identify ALL distinct operations** (look for numbered lists, "AND", commas, separate sentences)
3. **For EACH operation, determine:**
   - What type of operation? (delete, filter, sort, freeze, highlight, etc.)
   - What target? (all sheets, specific sheet, or default)
   - What parameters? (column names, values, conditions)
4. **Create one operation object for EACH distinct action**
5. **DO NOT skip any operations** - if user lists 8 steps, generate 8 operations (excluding auto-handled items)
6. **Preserve the order** - operations should appear in the same order as user's request

**CRITICAL RULES**:

**MULTI-OPERATION REQUESTS:**
- If user requests MULTIPLE distinct operations, use "operations" array format
- Each operation in the array can have its own target (apply_to_all_sheets, target_sheet, or neither)
- **NUMBERED LISTS**: If user provides a numbered list (1. 2. 3. etc.), parse EACH item as a separate operation
- **DO NOT SKIP OPERATIONS**: Parse every single step the user mentions
- **MERGE + OTHER OPERATIONS**: If user requests merge AND other operations in the same query, use operations array with merge_files as first operation
  - Example: "1. Merge files 2. Delete last row 3. Freeze panes" → operations array with 3 operations
  - The merge_files operation MUST be first in the array
  - All other operations will be applied to the merged file
- **AUTO-HANDLED ITEMS (still include in operations)**:
  - "Apply filters" → Include as {{"auto_filter": true}} (enables Excel's filter dropdowns)
  - "Fit columns" → Auto-handled by default, no operation needed
  - "Tabs are named X" → Part of merge_files operation, use sheet_names parameter
- Examples:
  - "Delete last row from both sheets AND sort Applicants" → 2 operations
  - "Freeze panes on all sheets AND highlight rows in Learners" → 2 operations
  - "1. Delete last row 2. Freeze panes 3. Sort Applicants" → 3 operations
  - "1. Merge files 2. Delete last row 3. Sort Applicants" → 3 operations (merge first!)
  - "Just delete last row" → 1 operation (no array)

**FILTER vs DELETE - PAY ATTENTION:**
- **FILTER** (keep ONLY matching rows): "show me only X", "give me just X", "I want only X", "where Status is X"
  - Use: `filter` operation
  - Result: File contains ONLY the matching rows
- **DELETE** (remove matching rows): "delete X rows", "remove X", "without X", "get rid of X", "X rows are deleted"
  - Use: `delete_rows` operation
  - Result: File contains everything EXCEPT the matching rows

**Sheet-Specific Operations:**
- If user mentions a specific sheet/tab name (e.g., "Learners sheet", "Applicants tab", "the second sheet"), use target_sheet
- If user says "both sheets", "all sheets", "both spreadsheets", "all tabs", use apply_to_all_sheets: true
- Sheet names are case-insensitive: "learners", "Learners", "LEARNERS" all refer to the same sheet
- Common sheet references: "sheet", "tab", "worksheet", "spreadsheet"
- **CRITICAL**: If user mentions a sheet name ANYWHERE in the query (even without "sheet" or "tab"), use target_sheet
- Examples: 
  - "Delete rows from Learners" → {{"target_sheet": "Learners", "delete_rows": ...}}
  - "Highlight rows in Learners" → {{"target_sheet": "Learners", "highlight_rows": ...}}
  - "Learners with Status On Break are highlighted yellow" → {{"target_sheet": "Learners", "highlight_rows": ...}}
  - "In Applicants, sort by Status" → {{"target_sheet": "Applicants", "sort": ...}}
  - "Delete last row from both sheets" → {{"apply_to_all_sheets": true, "remove_last_row": true}}
  - "Filter both spreadsheets" → {{"apply_to_all_sheets": true, "filter": ...}}

**Other Rules:**
- If user wants to "merge", "combine", "consolidate" files or wants "single file with tabs/sheets", use merge_files: true
- If user specifies custom tab/sheet names (e.g., "tabs are named X and Y", "sheets called A and B"), extract them into sheet_names array
- If user wants to "rename", "call it", "name it", "name the file" - extract the desired name into custom_filename (without extension)
- If user wants to "delete", "remove" the "last row", "final row", "filter description", "description row" - use remove_last_row: true
- If user wants to "delete row X", "remove rows X to Y", "delete rows where..." - use delete_rows with row_numbers or condition
- If user wants to "freeze panes", "keep header visible", "always see top row", "freeze columns" - use freeze_panes with row and col parameters
  - "freeze top row" = {{"row": 1, "col": 0}}
  - "freeze top row and first column" = {{"row": 1, "col": 1}}
  - "freeze top row and learner name" (if learner name is first column) = {{"row": 1, "col": 1}}
- custom_filename can be combined with any other operation (filter, merge, highlight, etc.)
- remove_last_row, delete_rows, and freeze_panes can be combined with any other operation
- Return ONLY valid JSON, no explanations
- Use exact column names from the available columns
- If the query doesn't match any operation, return: {{"no_operation": true}}

**FINAL REMINDER BEFORE GENERATING JSON:**
- Count the total number of steps in the user's query
- Parse EACH step and determine if it's an operation, metadata, or auto-handled
- DO NOT stop after the first operation - continue parsing ALL steps
- If you see merge_files as step 1, continue parsing steps 2, 3, 4... until the end
- Use "operations" array format when there are 2+ operations
- The number of operations depends on what the user requests - it could be 3, 5, 10, or 14 operations
- Trust your intelligence to understand each step's intent

Generate the JSON specification:
"""
        
        try:
            import google.generativeai as genai
            response = self.model.generate_content(
                prompt,
                generation_config=genai.types.GenerationConfig(
                    temperature=0.1,
                    top_p=0.8,
                    top_k=20
                )
            )
            
            response_text = response.text.strip()
            
            # Extract JSON from response
            import json
            
            # Try to find JSON in code blocks
            if '```json' in response_text:
                json_match = re.search(r'```json\s*(\{.*?\})\s*```', response_text, re.DOTALL)
                if json_match:
                    response_text = json_match.group(1)
            elif '```' in response_text:
                json_match = re.search(r'```\s*(\{.*?\})\s*```', response_text, re.DOTALL)
                if json_match:
                    response_text = json_match.group(1)
            
            # Parse JSON
            operations = json.loads(response_text)
            
            # Handle case where LLM returns an array instead of an object
            if isinstance(operations, list):
                logger.info(f"🔄 LLM returned operations array, converting to object format")
                operations = {"operations": operations}
            
            if operations.get('no_operation'):
                return None
            
            logger.info(f"✅ Parsed Excel operations: {operations}")
            
            # AGENTIC APPROACH: LLM validates its own output
            validated_operations = self._validate_operations_with_llm(
                operations, df, user_query, last_generated_file
            )
            
            return validated_operations
            
        except Exception as e:
            logger.error(f"❌ Error parsing Excel operations: {e}")
            return None
    
    def _validate_operations_with_llm(
        self,
        operations: Dict[str, Any],
        df: pd.DataFrame,
        user_query: str,
        last_generated_file: dict = None
    ) -> Dict[str, Any]:
        """
        AGENTIC VALIDATION: LLM validates and corrects its own output
        This implements the Reflection pattern for self-correction
        """
        try:
            import google.generativeai as genai
            
            # Add last generated file context
            last_file_context = ""
            if last_generated_file:
                last_file_context = f"""
**Last Generated File Context**:
- Filename: {last_generated_file.get('filename')}
- Operations already applied: {json.dumps(last_generated_file.get('operations', {}), indent=2)}

**CRITICAL CHECK FOR MERGING**:
If user query contains words like "add", "also", "add on that", "additionally", check:
- Did you MERGE the new operations with the operations from Last Generated File?
- For highlights: If previous file had green highlighting and user says "add red", your output should have BOTH green AND red in an array
- Example: If last file had {{"highlight_cells": {{"color": "green"}}}}, and user says "add red", output should be:
  {{"highlight_cells": [{{"color": "green"}}, {{"color": "red"}}]}}
"""
            
            # Create validation prompt
            validation_prompt = f"""
You are a quality assurance agent reviewing Excel operations that you just generated.

**Original User Query**: "{user_query}"

**Operations You Generated**:
```json
{json.dumps(operations, indent=2)}
```

{last_file_context}

**Available DataFrame Columns**: {list(df.columns)}
**DataFrame Sample**:
{df.head(3).to_string()}

**Your Task - Critical Review**:
Carefully review the operations you generated and check for:

1. **Column Name Accuracy**: Are all column names exactly as they appear in the DataFrame?
2. **Value Format Correctness**: Are percentage values converted correctly (e.g., "0%" → 0.0)?
3. **Operation Logic**: Will these operations achieve the user's goal?
4. **Format Consistency**: Is the JSON structure correct for the operation type?
5. **Completeness**: Did you parse all user requests or miss any?
6. **MERGING CHECK**: If user said "add"/"also", did you include ALL previous operations from Last Generated File?

**IMPORTANT**: Only remove operations if they are IMPOSSIBLE to implement. If an operation is complex but feasible, keep it and try your best to implement it correctly. For example:
- Date comparisons: Use string matching or approximate logic if exact date comparison is complex
- Conditional formatting: Use reasonable default conditions if user didn't specify exact values
- Complex highlighting: Implement with best-effort logic

DO NOT remove operations just because they are "not well-defined" - make reasonable assumptions and implement them.

**CRITICAL - Format Requirements**:
- For multiple operations of the SAME type (e.g., multiple highlight_rows), use array format:
  {{"highlight_rows": [{{"column": "X", ...}}, {{"column": "Y", ...}}]}}
  
- For multiple operations of DIFFERENT types, use individual keys:
  {{"highlight_rows": [...], "freeze_panes": {{}}, "sort": {{}}}}
  
- For single operations, use direct format:
  {{"highlight_rows": {{"column": "X", ...}}}}

**Response Format**:
If you find ANY issues, respond with:
```json
{{
  "status": "CORRECTED",
  "issues_found": ["issue 1", "issue 2"],
  "corrected_operations": {{...corrected operations...}}
}}
```

If everything is perfect, respond with:
```json
{{
  "status": "VALID"
}}
```

Perform your critical review now:
"""
            
            response = self.model.generate_content(
                validation_prompt,
                generation_config=genai.types.GenerationConfig(
                    temperature=0.1,
                    top_p=0.8,
                    top_k=20
                )
            )
            
            response_text = response.text.strip()
            
            # Extract JSON from response
            if '```json' in response_text:
                json_match = re.search(r'```json\s*(\{.*?\})\s*```', response_text, re.DOTALL)
                if json_match:
                    response_text = json_match.group(1)
            elif '```' in response_text:
                json_match = re.search(r'```\s*(\{.*?\})\s*```', response_text, re.DOTALL)
                if json_match:
                    response_text = json_match.group(1)
            
            validation_result = json.loads(response_text)
            
            if validation_result.get("status") == "CORRECTED":
                logger.info(f"🔄 LLM self-corrected operations. Issues found: {validation_result.get('issues_found')}")
                return validation_result["corrected_operations"]
            else:
                logger.info("✅ LLM validated operations as correct")
                return operations
                
        except Exception as e:
            logger.warning(f"⚠️ LLM validation failed: {e}, using original operations")
            return operations
    
    def correct_operations_from_error(
        self,
        operations: Dict[str, Any],
        error_message: str,
        df: pd.DataFrame,
        user_query: str
    ) -> Dict[str, Any]:
        """
        AGENTIC ERROR CORRECTION: LLM fixes operations based on execution errors
        This implements self-healing with error feedback
        """
        try:
            import google.generativeai as genai
            
            correction_prompt = f"""
You are an error correction agent. An operation you generated failed during execution.

**Original User Query**: "{user_query}"

**Operations That Failed**:
```json
{json.dumps(operations, indent=2)}
```

**Error Message**:
{error_message}

**Available DataFrame Columns**: {list(df.columns)}
**DataFrame Sample**:
{df.head(3).to_string()}

**Your Task**:
Analyze the error and generate CORRECTED operations that will work.

Common issues to check:
- Column names must match exactly (case-sensitive)
- Value types must be correct (numbers, strings, etc.)
- Array vs single object format
- Missing required fields

**Response Format**:
```json
{{
  "analysis": "Brief explanation of what went wrong",
  "corrected_operations": {{...corrected operations...}}
}}
```

Generate the corrected operations:
"""
            
            response = self.model.generate_content(
                correction_prompt,
                generation_config=genai.types.GenerationConfig(
                    temperature=0.2,
                    top_p=0.8,
                    top_k=20
                )
            )
            
            response_text = response.text.strip()
            
            # Extract JSON
            if '```json' in response_text:
                json_match = re.search(r'```json\s*(\{.*?\})\s*```', response_text, re.DOTALL)
                if json_match:
                    response_text = json_match.group(1)
            elif '```' in response_text:
                json_match = re.search(r'```\s*(\{.*?\})\s*```', response_text, re.DOTALL)
                if json_match:
                    response_text = json_match.group(1)
            
            correction_result = json.loads(response_text)
            
            logger.info(f"🔧 LLM corrected operations. Analysis: {correction_result.get('analysis')}")
            return correction_result["corrected_operations"]
            
        except Exception as e:
            logger.error(f"❌ LLM error correction failed: {e}")
            return operations
    
    def determine_file_source_with_llm(
        self,
        user_query: str,
        last_generated_file: Optional[Dict[str, Any]],
        conversation_history: List[Dict[str, str]] = None
    ) -> str:
        """
        AGENTIC FILE SOURCE DETERMINATION
        
        Uses LLM to intelligently decide whether to use:
        - "last_generated": Build on the last generated file (sequential operation)
        - "original_files": Start fresh from originally uploaded files
        - "auto": Let the system decide based on operation type
        
        This is the KEY to making the system truly agentic and non-linear.
        """
        try:
            import google.generativeai as genai
            
            # Build context about last generated file
            last_file_context = "No previous file generated in this session"
            if last_generated_file:
                last_file_context = f"""
Last Generated File:
- Filename: {last_generated_file.get('filename')}
- Type: {last_generated_file.get('type')}
- Sheets: {last_generated_file.get('sheets', [])}
- Previous Operation: {last_generated_file.get('operation', 'unknown')}
- Timestamp: {last_generated_file.get('timestamp', 'unknown')}
"""
            
            # Build conversation context
            conv_context = "No previous conversation"
            if conversation_history and len(conversation_history) > 0:
                recent_conv = conversation_history[-3:]  # Last 3 messages
                conv_context = "\n".join([
                    f"{'User' if msg.get('role') == 'user' else 'Assistant'}: {msg.get('content', '')[:100]}"
                    for msg in recent_conv
                ])
            
            prompt = f"""
You are an intelligent file source selector for an Excel operations system.

**Current User Query**: "{user_query}"

**Session Context**:
{last_file_context}

**Recent Conversation**:
{conv_context}

**Your Task**:
Determine whether this operation should use:
1. **"last_generated"**: Build upon the last generated file (sequential operation)
2. **"original_files"**: Start fresh from the originally uploaded files (new operation chain)

**Decision Guidelines**:

USE "last_generated" when:
- User says "also", "additionally", "and then", "next", "after that"
- User refers to "the file", "this file", "current file", "that spreadsheet"
- Operation naturally builds on previous work (e.g., "now highlight", "then delete")
- User mentions previous operations (e.g., "the merged file", "the highlighted data")
- Sequential context is clear from conversation

USE "original_files" when:
- User says "start over", "new file", "from scratch", "original data"
- User explicitly mentions the original file names
- Operation is completely independent (e.g., first operation in session)
- User wants to undo previous work
- No last generated file exists

**CRITICAL AGENTIC PRINCIPLE**:
The order doesn't matter! User can do:
- Merge → Highlight → Delete → Sort
- Delete → Merge → Sort → Highlight  
- Sort → Delete → Highlight → Merge
- ANY random order!

Your job is to understand INTENT, not enforce order.

**Examples**:

Query: "Delete the last row from both sheets"
Context: Last file was a merge with 2 sheets
Decision: "last_generated" (refers to "both sheets" from merged file)

Query: "Merge Output_Applicants and Output_Learners"
Context: No last file
Decision: "original_files" (first operation, mentions original files)

Query: "Now highlight rows where Status is On Break"
Context: Last file was a merge
Decision: "last_generated" ("Now" indicates sequential, build on merge)

Query: "Sort by Salary descending"
Context: Last file was highlighted data
Decision: "last_generated" (continue working on current data)

Query: "Create a new file with just the Applicants data"
Context: Last file was a merge
Decision: "original_files" ("new file" indicates fresh start)

**Response Format** (JSON only):
```json
{{
  "file_source": "last_generated" or "original_files",
  "reasoning": "Brief explanation of why",
  "confidence": "high" or "medium" or "low"
}}
```

Analyze the query and generate your decision:
"""
            
            response = self.model.generate_content(
                prompt,
                generation_config=genai.types.GenerationConfig(
                    temperature=0.1,
                    top_p=0.8,
                    top_k=20
                )
            )
            
            response_text = response.text.strip()
            
            # Extract JSON
            if '```json' in response_text:
                json_match = re.search(r'```json\s*(\{.*?\})\s*```', response_text, re.DOTALL)
                if json_match:
                    response_text = json_match.group(1)
            elif '```' in response_text:
                json_match = re.search(r'```\s*(\{.*?\})\s*```', response_text, re.DOTALL)
                if json_match:
                    response_text = json_match.group(1)
            
            decision = json.loads(response_text)
            
            file_source = decision.get('file_source', 'auto')
            reasoning = decision.get('reasoning', 'No reasoning provided')
            confidence = decision.get('confidence', 'medium')
            
            logger.info(f"🤖 LLM File Source Decision: {file_source} (confidence: {confidence})")
            logger.info(f"   Reasoning: {reasoning}")
            
            return file_source
            
        except Exception as e:
            logger.error(f"❌ Error in LLM file source determination: {e}")
            # Fallback: if last file exists, use it; otherwise use original
            return "last_generated" if last_generated_file else "original_files"
