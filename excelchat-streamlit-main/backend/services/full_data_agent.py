"""
Full LLM-Based Agentic System for Excel Processing
Implements complete DataFrame loading, intelligent profiling, and code generation
"""

import os
import json
import logging
import pandas as pd
import numpy as np
import traceback
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime
import google.generativeai as genai
import matplotlib.pyplot as plt
import seaborn as sns
import io
import base64
import warnings

# Suppress NumPy warnings that trigger Flask auto-reload
warnings.filterwarnings('ignore', category=RuntimeWarning, module='numpy')
warnings.filterwarnings('ignore', message='Mean of empty slice')

# Import Excel operations module
from .excel_operations import ExcelOperations, ExcelOperationIntentParser
from .multi_sheet_handler import handle_multi_sheet_operation

# Import configuration
import sys
from pathlib import Path
backend_path = Path(__file__).resolve().parent.parent
if str(backend_path) not in sys.path:
    sys.path.insert(0, str(backend_path))
from config.settings import settings, error_messages

logger = logging.getLogger(__name__)

class FullDataAgent:
    """
    Complete agentic system that loads entire Excel files and provides
    full LLM-based analysis with code generation and execution
    """
    
    def __init__(self, api_key: str, uploads_dir: str = None, outputs_dir: str = None):
        self.api_key = api_key
        genai.configure(api_key=api_key)
        
        # Directory setup - use settings if not provided
        self.uploads_dir = Path(uploads_dir) if uploads_dir else settings.UPLOADS_DIR
        self.outputs_dir = Path(outputs_dir) if outputs_dir else settings.OUTPUTS_DIR
        self.uploads_dir.mkdir(exist_ok=True, parents=True)
        self.outputs_dir.mkdir(exist_ok=True, parents=True)
        
        # Initialize Gemini model
        self.model = None
        self._initialize_model()
        
        # Initialize Excel operations
        self.excel_ops = ExcelOperations(self.outputs_dir)
        self.intent_parser = ExcelOperationIntentParser(self.model)
        
        # Initialize Google Sheets service
        try:
            from .google_sheets_service import GoogleSheetsService
            self.google_sheets = GoogleSheetsService()
            logger.info("✅ Google Sheets service initialized")
        except Exception as e:
            logger.warning(f"⚠️ Google Sheets service not available: {e}")
            self.google_sheets = None
        
        # Data storage - Multiple DataFrames per session
        self.session_data: Dict[str, Dict] = {}  # session_id -> session data
        # Each session contains:
        # {
        #   "files": {file_id: {"df": DataFrame, "profile": dict, "filename": str}},
        #   "conversation_history": [],
        #   "active_file": file_id or None,
        #   "combined_df": DataFrame (if multiple files)
        # }
        
        # Configuration from settings
        self.MAX_FILE_SIZE = settings.get_max_file_size_bytes()
        self.MAX_ROWS_THRESHOLD = settings.MAX_ROWS_THRESHOLD
        
    def _initialize_model(self):
        """Initialize Gemini model with fallbacks"""
        model_candidates = [
            "models/gemini-2.5-flash",
            "models/gemini-2.0-flash", 
            "models/gemini-1.5-flash",
            "models/gemini-pro"
        ]
        
        for candidate in model_candidates:
            try:
                self.model = genai.GenerativeModel(candidate)
                logger.info(f"✅ Full Data Agent initialized with: {candidate}")
                
                # Test the model
                test_response = self.model.generate_content("Hello, respond with 'OK'")
                logger.info(f"✅ Model test successful: {test_response.text[:20]}")
                return  # Successfully initialized, exit the method
            except Exception as e:
                logger.warning(f"⚠️ Failed to initialize '{candidate}': {e}")
                continue  # Try next candidate
        
        # If we get here, no model was successfully initialized
        raise Exception("❌ Could not initialize any Gemini model")
    
    def _detect_header_row(self, file_path: Path) -> int:
        """
        Detect the correct header row in an Excel file
        Returns the row index (0-based) where the actual headers are
        """
        try:
            # Read first 20 rows without header to analyze
            df_preview = pd.read_excel(file_path, header=None, nrows=20)
            
            # Check each row to find the one with the most non-null, non-numeric values
            best_header_row = 0
            best_score = 0
            
            for idx in range(min(10, len(df_preview))):
                row = df_preview.iloc[idx]
                
                # Count non-null values
                non_null_count = row.notna().sum()
                
                # Count string values (likely column names)
                string_count = sum(1 for val in row if isinstance(val, str) and val.strip())
                
                # Score: prefer rows with many non-null string values
                score = string_count * 2 + non_null_count
                
                if score > best_score:
                    best_score = score
                    best_header_row = idx
            
            # If we found a row with mostly strings in the first 10 rows, use it
            if best_score > 0 and best_header_row > 0:
                logger.info(f"📋 Detected header row at index {best_header_row}")
                return best_header_row
            
            return 0  # Default to first row
            
        except Exception as e:
            logger.warning(f"⚠️ Error detecting header row: {e}, using default (row 0)")
            return 0
    
    def load_excel_file(self, file_path: str, session_id: str, file_id: str = None) -> Dict[str, Any]:
        """
        Load Excel file into memory with support for multiple files per session
        
        Args:
            file_path: Path to the Excel file
            session_id: Session identifier
            file_id: Optional file identifier (auto-generated if None)
        
        Returns:
            Dict with success status and file information
        """
        try:
            file_path = Path(file_path)
            
            # Generate file_id if not provided
            if file_id is None:
                file_id = f"file_{len(self.get_session_files(session_id)) + 1}_{file_path.stem}"
            
            # Initialize session if it doesn't exist
            if session_id not in self.session_data:
                self.session_data[session_id] = {
                    "files": {},
                    "conversation_history": [],
                    "active_file": None,
                    "combined_df": None,
                    "last_generated_file": None,  # Track last generated Excel file
                    "generated_files_history": []  # Track all generated files
                }
            
            # 1. File size validation
            file_size = file_path.stat().st_size
            if file_size > self.MAX_FILE_SIZE:
                return {
                    "success": False,
                    "error": f"File too large: {file_size/1024/1024:.1f}MB > {self.MAX_FILE_SIZE/1024/1024}MB limit"
                }
            
            logger.info(f"📁 Loading Excel file: {file_path.name} ({file_size/1024/1024:.1f}MB) as {file_id}")
            
            # 2. Load entire DataFrame (with multi-sheet support for Excel)
            if file_path.suffix.lower() == '.csv':
                df = pd.read_csv(file_path)
                logger.info(f"📊 Loaded CSV: {len(df)} rows × {len(df.columns)} columns")
                
            elif file_path.suffix.lower() in ['.xlsx', '.xls']:
                # Check for multiple sheets
                xl_file = pd.ExcelFile(file_path)
                sheet_names = xl_file.sheet_names
                
                if len(sheet_names) > 1:
                    # Multi-sheet Excel file - load each sheet as a separate "file"
                    logger.info(f"📊 Detected multi-sheet Excel with {len(sheet_names)} sheets: {sheet_names}")
                    
                    loaded_sheets = []
                    for idx, sheet_name in enumerate(sheet_names):
                        df_sheet = pd.read_excel(xl_file, sheet_name=sheet_name)
                        
                        # Log sheet info
                        if len(df_sheet) == 0:
                            logger.warning(f"⚠️ Sheet '{sheet_name}' is empty, skipping optimization")
                        else:
                            logger.info(f"✅ Loaded sheet '{sheet_name}': {len(df_sheet)} rows × {len(df_sheet.columns)} columns")
                        
                        # Create unique file_id for each sheet
                        sheet_file_id = f"{file_id}_sheet_{idx}_{sheet_name.replace(' ', '_')}"
                        
                        # Optimize data types
                        df_sheet_optimized = self._optimize_dtypes(df_sheet)
                        memory_saved_sheet = df_sheet.memory_usage(deep=True).sum() - df_sheet_optimized.memory_usage(deep=True).sum()
                        
                        # Create data profile
                        data_profile_sheet = self._create_comprehensive_profile(df_sheet_optimized)
                        
                        # Store sheet data
                        self.session_data[session_id]["files"][sheet_file_id] = {
                            "df": df_sheet_optimized,
                            "profile": data_profile_sheet,
                            "filename": f"{file_path.stem}_{sheet_name}",
                            "sheet_name": sheet_name,
                            "sheet_index": idx,
                            "original_file": file_path.name,
                            "original_file_path": str(file_path),
                            "is_sheet": True,
                            "file_size_mb": file_size / 1024 / 1024,
                            "memory_usage_mb": df_sheet_optimized.memory_usage(deep=True).sum() / 1024 / 1024,
                            "memory_saved_mb": memory_saved_sheet / 1024 / 1024
                        }
                        
                        loaded_sheets.append({
                            "sheet_name": sheet_name,
                            "file_id": sheet_file_id,
                            "rows": len(df_sheet_optimized),
                            "columns": len(df_sheet_optimized.columns)
                        })
                        
                        logger.info(f"✅ Loaded sheet '{sheet_name}': {len(df_sheet_optimized)} rows × {len(df_sheet_optimized.columns)} columns")
                    
                    # Set first sheet as active
                    if self.session_data[session_id]["active_file"] is None:
                        self.session_data[session_id]["active_file"] = loaded_sheets[0]["file_id"]
                    
                    # Update combined DataFrame
                    self._update_combined_dataframe(session_id)
                    
                    return {
                        "success": True,
                        "file_id": file_id,
                        "is_multi_sheet": True,
                        "sheets": loaded_sheets,
                        "total_rows": sum(s["rows"] for s in loaded_sheets),
                        "message": f"Loaded {len(sheet_names)} sheets from {file_path.name}"
                    }
                else:
                    # Single sheet Excel file - detect header row
                    header_row = self._detect_header_row(file_path)
                    df = pd.read_excel(xl_file, sheet_name=sheet_names[0], header=header_row)
                    logger.info(f"📊 Loaded single-sheet Excel: {len(df)} rows × {len(df.columns)} columns (header at row {header_row})")
                    
            elif file_path.suffix.lower() == '.ods':
                df = pd.read_excel(file_path, engine='odf')
                logger.info(f"📊 Loaded ODS: {len(df)} rows × {len(df.columns)} columns")
                
            elif file_path.suffix.lower() == '.gsheet':
                # Try CSV first, then Excel
                try:
                    df = pd.read_csv(file_path)
                except:
                    df = pd.read_excel(file_path)
                logger.info(f"📊 Loaded Google Sheet: {len(df)} rows × {len(df.columns)} columns")
            else:
                return {"success": False, "error": f"Unsupported file type: {file_path.suffix}"}
            
            # Check row count
            if len(df) > self.MAX_ROWS_THRESHOLD:
                logger.warning(f"⚠️ Large dataset: {len(df)} rows > {self.MAX_ROWS_THRESHOLD} threshold")
            
            # 3. Optimize data types to reduce memory usage
            df_optimized = self._optimize_dtypes(df)
            memory_saved = df.memory_usage(deep=True).sum() - df_optimized.memory_usage(deep=True).sum()
            logger.info(f"💾 Memory optimization: Saved {memory_saved/1024/1024:.1f}MB")
            
            # 4. Create comprehensive data profile
            data_profile = self._create_comprehensive_profile(df_optimized)
            
            # 5. Store file data in session
            self.session_data[session_id]["files"][file_id] = {
                "df": df_optimized,
                "profile": data_profile,
                "filename": file_path.name,
                "file_path": str(file_path),  # Store source file path for format preservation
                "file_size_mb": file_size / 1024 / 1024,
                "memory_usage_mb": df_optimized.memory_usage(deep=True).sum() / 1024 / 1024,
                "memory_saved_mb": memory_saved / 1024 / 1024
            }
            
            # Set as active file if it's the first one
            if self.session_data[session_id]["active_file"] is None:
                self.session_data[session_id]["active_file"] = file_id
            
            # Update combined DataFrame if multiple files
            self._update_combined_dataframe(session_id)
            
            logger.info(f"✅ Excel file loaded successfully: {file_id} for session: {session_id}")
            
            return {
                "success": True,
                "session_id": session_id,
                "file_id": file_id,
                "filename": file_path.name,
                "rows": len(df_optimized),
                "columns": len(df_optimized.columns),
                "memory_usage_mb": df_optimized.memory_usage(deep=True).sum() / 1024 / 1024,
                "file_size_mb": file_size / 1024 / 1024,
                "memory_saved_mb": memory_saved / 1024 / 1024,
                "total_files": len(self.session_data[session_id]["files"]),
                "profile": data_profile
            }
            
        except Exception as e:
            logger.error(f"❌ Error loading Excel file: {e}", exc_info=True)
            return {"success": False, "error": str(e)}
    
    def load_google_sheet(
        self, 
        spreadsheet_url: str, 
        session_id: str, 
        sheet_name: str = None,
        file_id: str = None
    ) -> Dict[str, Any]:
        """
        Load Google Sheet directly from URL
        
        Args:
            spreadsheet_url: Google Sheets URL
            session_id: Session identifier
            sheet_name: Optional worksheet name
            file_id: Optional file identifier
        
        Returns:
            Dict with success status and file information
        """
        try:
            if not self.google_sheets:
                return {
                    "success": False,
                    "error": "Google Sheets integration not configured. Please check credentials."
                }
            
            logger.info(f"📊 Loading Google Sheet: {spreadsheet_url[:50]}...")
            
            # Generate file_id if not provided (needed for multi-sheet processing)
            if file_id is None:
                file_id = f"gsheet_{len(self.get_session_files(session_id)) + 1}"
            
            # Initialize session if needed
            if session_id not in self.session_data:
                self.session_data[session_id] = {
                    "files": {},
                    "conversation_history": [],
                    "active_file": None,
                    "combined_df": None,
                    "last_generated_file": None,
                    "generated_files_history": []
                }
            
            # Check if we should load all sheets or just one
            if sheet_name is None:
                # Load all sheets to check if it's multi-sheet
                all_sheets = self.google_sheets.read_all_sheets_to_dataframes(spreadsheet_url)
                
                if len(all_sheets) > 1:
                    # Multi-sheet Google Sheet - load each sheet as a separate "file"
                    logger.info(f"📊 Detected multi-sheet Google Sheet with {len(all_sheets)} sheets: {list(all_sheets.keys())}")
                    
                    loaded_sheets = []
                    for idx, (sheet_name, df_sheet) in enumerate(all_sheets.items()):
                        # Log sheet info
                        if len(df_sheet) == 0:
                            logger.warning(f"⚠️ Sheet '{sheet_name}' is empty, skipping optimization")
                        else:
                            logger.info(f"✅ Loaded sheet '{sheet_name}': {len(df_sheet)} rows × {len(df_sheet.columns)} columns")
                        
                        # Create unique file_id for each sheet
                        sheet_file_id = f"{file_id}_sheet_{idx}_{sheet_name.replace(' ', '_')}"
                        
                        # Optimize data types
                        df_sheet_optimized = self._optimize_dtypes(df_sheet)
                        
                        # Create data profile
                        data_profile_sheet = self._create_comprehensive_profile(df_sheet_optimized)
                        
                        # Store sheet data
                        self.session_data[session_id]["files"][sheet_file_id] = {
                            "df": df_sheet_optimized,
                            "profile": data_profile_sheet,
                            "filename": f"GoogleSheet_{sheet_name}",
                            "source_type": "google_sheets",
                            "spreadsheet_url": spreadsheet_url,
                            "sheet_name": sheet_name,
                            "sheet_index": idx,
                            "original_file": "Google Sheet",
                            "original_file_path": spreadsheet_url,
                            "is_sheet": True,
                            "file_size_mb": 0,  # Google Sheets don't have file size
                            "memory_usage_mb": df_sheet_optimized.memory_usage(deep=True).sum() / 1024 / 1024,
                            "memory_saved_mb": 0
                        }
                        
                        loaded_sheets.append({
                            "sheet_name": sheet_name,
                            "file_id": sheet_file_id,
                            "rows": int(len(df_sheet_optimized)),        # Convert to Python int
                            "columns": int(len(df_sheet_optimized.columns))  # Convert to Python int
                        })
                    
                    # Set active file to first sheet
                    if loaded_sheets:
                        self.session_data[session_id]["active_file"] = loaded_sheets[0]["file_id"]
                    
                    # Update combined DataFrame for multi-sheet searching
                    self._update_combined_dataframe(session_id)
                    
                    return {
                        "success": True,
                        "file_id": file_id,
                        "sheets": loaded_sheets,
                        "is_multi_sheet": True,
                        "message": f"Successfully loaded {len(loaded_sheets)} sheets from Google Sheet"
                    }
                else:
                    # Single sheet - use the first (and only) sheet
                    sheet_name, df = next(iter(all_sheets.items()))
                    logger.info(f"📊 Single-sheet Google Sheet: {sheet_name}")
            else:
                # Load specific sheet
                df = self.google_sheets.read_sheet_to_dataframe(
                    spreadsheet_url, 
                    sheet_name=sheet_name
                )
            
            # Optimize data types
            df_optimized = self._optimize_dtypes(df)
            
            # Create data profile
            data_profile = self._create_comprehensive_profile(df_optimized)
            
            # Store in session
            self.session_data[session_id]["files"][file_id] = {
                "df": df_optimized,
                "profile": data_profile,
                "filename": f"GoogleSheet_{sheet_name or 'Sheet1'}",
                "source_type": "google_sheets",
                "spreadsheet_url": spreadsheet_url,
                "sheet_name": sheet_name,
                "memory_usage_mb": df_optimized.memory_usage(deep=True).sum() / 1024 / 1024
            }
            
            # Set as active file if first one
            if self.session_data[session_id]["active_file"] is None:
                self.session_data[session_id]["active_file"] = file_id
            
            # Update combined DataFrame
            self._update_combined_dataframe(session_id)
            
            logger.info(f"✅ Google Sheet loaded: {file_id} ({len(df_optimized)} rows)")
            
            return {
                "success": True,
                "session_id": session_id,
                "file_id": file_id,
                "filename": f"GoogleSheet_{sheet_name or 'Sheet1'}",
                "rows": int(len(df_optimized)),           # Convert to Python int
                "columns": int(len(df_optimized.columns)), # Convert to Python int
                "memory_usage_mb": float(df_optimized.memory_usage(deep=True).sum() / 1024 / 1024),
                "total_files": len(self.session_data[session_id]["files"]),
                "profile": data_profile
            }
            
        except Exception as e:
            logger.error(f"❌ Error loading Google Sheet: {e}", exc_info=True)
            return {"success": False, "error": str(e)}
    
    def load_google_sheet_all_tabs(
        self,
        spreadsheet_url: str,
        session_id: str
    ) -> Dict[str, Any]:
        """
        Load all sheets/tabs from a Google Spreadsheet
        
        Args:
            spreadsheet_url: Google Sheets URL
            session_id: Session identifier
        
        Returns:
            Dict with success status and information about loaded sheets
        """
        try:
            if not self.google_sheets:
                return {
                    "success": False,
                    "error": "Google Sheets integration not configured. Please check credentials."
                }
            
            logger.info(f"📊 Loading all sheets from Google Spreadsheet: {spreadsheet_url[:50]}...")
            
            # Read all sheets
            all_sheets = self.google_sheets.read_all_sheets_to_dataframes(spreadsheet_url)
            
            if not all_sheets:
                return {
                    "success": False,
                    "error": "No sheets found or all sheets are empty"
                }
            
            # Initialize session if needed
            if session_id not in self.session_data:
                self.session_data[session_id] = {
                    "files": {},
                    "conversation_history": [],
                    "active_file": None,
                    "combined_df": None,
                    "last_generated_file": None,
                    "generated_files_history": []
                }
            
            loaded_sheets = []
            total_rows = 0
            
            # Load each sheet as a separate file
            for sheet_name, df in all_sheets.items():
                file_id = f"gsheet_{session_id}_{sheet_name.replace(' ', '_')}"
                
                # Optimize data types
                df_optimized = self._optimize_dtypes(df)
                
                # Create data profile
                data_profile = self._create_comprehensive_profile(df_optimized)
                
                # Store in session
                self.session_data[session_id]["files"][file_id] = {
                    "df": df_optimized,
                    "profile": data_profile,
                    "filename": f"GoogleSheet_{sheet_name}",
                    "source_type": "google_sheets",
                    "spreadsheet_url": spreadsheet_url,
                    "sheet_name": sheet_name,
                    "memory_usage_mb": df_optimized.memory_usage(deep=True).sum() / 1024 / 1024
                }
                
                loaded_sheets.append({
                    "sheet_name": sheet_name,
                    "file_id": file_id,
                    "rows": len(df_optimized),
                    "columns": len(df_optimized.columns)
                })
                
                total_rows += len(df_optimized)
                
                logger.info(f"  ✅ Loaded sheet '{sheet_name}': {len(df_optimized)} rows × {len(df_optimized.columns)} columns")
            
            # Set first sheet as active if no active file
            if self.session_data[session_id]["active_file"] is None and loaded_sheets:
                self.session_data[session_id]["active_file"] = loaded_sheets[0]["file_id"]
            
            # Update combined DataFrame
            self._update_combined_dataframe(session_id)
            
            logger.info(f"✅ Loaded {len(loaded_sheets)} sheets with {total_rows} total rows")
            
            return {
                "success": True,
                "session_id": session_id,
                "sheets_loaded": len(loaded_sheets),
                "sheets": loaded_sheets,
                "total_rows": total_rows,
                "total_files": len(self.session_data[session_id]["files"])
            }
            
        except Exception as e:
            logger.error(f"❌ Error loading all sheets from Google Spreadsheet: {e}", exc_info=True)
            return {"success": False, "error": str(e)}
    
    def get_session_files(self, session_id: str) -> Dict[str, Dict]:
        """Get all files for a session"""
        if session_id not in self.session_data:
            return {}
        return self.session_data[session_id]["files"]
    
    def _update_combined_dataframe(self, session_id: str):
        """Update combined DataFrame when multiple files are loaded"""
        if session_id not in self.session_data:
            return
        
        files = self.session_data[session_id]["files"]
        if len(files) <= 1:
            # Single file or no files - no need for combined DataFrame
            self.session_data[session_id]["combined_df"] = None
            return
        
        # Combine all DataFrames
        dataframes = []
        for file_id, file_data in files.items():
            df = file_data["df"].copy()
            df["_source_file"] = file_data["filename"]  # Add source file column
            df["_file_id"] = file_id  # Add file ID column
            dataframes.append(df)
        
        try:
            # Attempt to concatenate - this works if columns are similar
            combined_df = pd.concat(dataframes, ignore_index=True, sort=False)
            self.session_data[session_id]["combined_df"] = combined_df
            logger.info(f"📊 Combined {len(files)} files into single DataFrame: {len(combined_df)} rows")
        except Exception as e:
            logger.warning(f"⚠️ Could not combine files - different structures: {e}")
            self.session_data[session_id]["combined_df"] = None
    
    def set_active_file(self, session_id: str, file_id: str) -> bool:
        """Set the active file for analysis"""
        if session_id not in self.session_data:
            return False
        if file_id not in self.session_data[session_id]["files"]:
            return False
        
        self.session_data[session_id]["active_file"] = file_id
        logger.info(f"📌 Set active file: {file_id} for session: {session_id}")
        return True
    
    def get_dataframe_by_sheet_name(self, session_id: str, sheet_name: str) -> pd.DataFrame:
        """Get DataFrame for a specific sheet by name"""
        if session_id not in self.session_data:
            return None
        
        files = self.session_data[session_id]["files"]
        
        # Search for sheet by name
        for file_id, file_data in files.items():
            if file_data.get("is_sheet", False):
                if file_data.get("sheet_name", "").lower() == sheet_name.lower():
                    logger.info(f"📄 Found sheet '{sheet_name}' with {len(file_data['df'])} rows")
                    return file_data["df"]
        
        logger.warning(f"⚠️ Sheet '{sheet_name}' not found")
        return None
    
    def get_active_dataframe(self, session_id: str) -> pd.DataFrame:
        """Get the DataFrame to use for analysis (active file, combined, or single file)"""
        if session_id not in self.session_data:
            return None
        
        session = self.session_data[session_id]
        files = session["files"]
        
        if len(files) == 0:
            return None
        elif len(files) == 1:
            # Single file - return it
            file_id = list(files.keys())[0]
            return files[file_id]["df"]
        else:
            # Multiple files - return combined if available, otherwise active file
            if session["combined_df"] is not None:
                return session["combined_df"]
            elif session["active_file"] and session["active_file"] in files:
                return files[session["active_file"]]["df"]
            else:
                # Fallback to first file
                file_id = list(files.keys())[0]
                return files[file_id]["df"]
    
    def _build_multi_file_context(self, session_id: str, df: pd.DataFrame, user_query: str) -> str:
        """Build context for LLM with multi-file awareness and conversation history"""
        session = self.session_data[session_id]
        files = session["files"]
        history = session.get("conversation_history", [])
        
        context_parts = []
        
        # Conversation history for contextual memory
        if history:
            context_parts.append("=== CONVERSATION HISTORY ===")
            context_parts.append("Previous questions and answers in this session:")
            
            # Include last 5 conversations for context (to avoid token limits)
            recent_history = history[-5:] if len(history) > 5 else history
            for idx, conv in enumerate(recent_history, 1):
                user_q = conv.get("user_query", "")
                # Get the execution output or result
                exec_result = conv.get("execution_result", {})
                output = exec_result.get("execution_output", "") if isinstance(exec_result, dict) else str(exec_result)
                
                context_parts.append(f"\n{idx}. User asked: \"{user_q}\"")
                if output:
                    # Truncate long outputs
                    output_preview = output[:200] + "..." if len(output) > 200 else output
                    context_parts.append(f"   Answer: {output_preview}")
            
            context_parts.append("\n⚠️ IMPORTANT: Use this conversation history to understand context. If the user refers to something mentioned before (like a BUD ID, name, or value), use that information from the history above.")
            context_parts.append("")
        
        # Multi-file/Multi-sheet information
        if len(files) > 1:
            context_parts.append("=== MULTI-FILE/MULTI-SHEET SESSION ===")
            context_parts.append(f"Total files/sheets loaded: {len(files)}")
            
            for file_id, file_data in files.items():
                # Check if this is a sheet from a multi-sheet Excel
                if file_data.get("is_sheet", False):
                    original_file = file_data.get('original_file', 'merged file')
                    context_parts.append(f"\nSheet: {file_data['sheet_name']} from {original_file} (ID: {file_id})")
                    context_parts.append(f"- Sheet Index: {file_data['sheet_index']}")
                else:
                    context_parts.append(f"\nFile: {file_data['filename']} (ID: {file_id})")
                
                context_parts.append(f"- Rows: {len(file_data['df'])}")
                context_parts.append(f"- Columns: {len(file_data['df'].columns)}")
                context_parts.append(f"- Column names: {list(file_data['df'].columns)}")
                
            context_parts.append(f"\n⚠️ CRITICAL: When asked about specific sheets or files, use the sheet/file name to identify which data to work with. If user mentions a sheet name (e.g., 'Learners sheet', 'Applicants tab'), work with that specific sheet's data.")
            
            if session["combined_df"] is not None:
                context_parts.append(f"\n📊 COMBINED DATASET (currently active):")
                context_parts.append(f"- Total rows: {len(df)}")
                context_parts.append(f"- Total columns: {len(df.columns)}")
                context_parts.append(f"- Includes '_source_file' and '_file_id' columns for tracking")
            else:
                active_file = session["active_file"]
                if active_file:
                    context_parts.append(f"\n📌 ACTIVE FILE: {files[active_file]['filename']}")
        else:
            # Single file
            file_id = list(files.keys())[0]
            file_data = files[file_id]
            context_parts.append(f"=== SINGLE FILE SESSION ===")
            context_parts.append(f"File: {file_data['filename']}")
        
        # Current DataFrame information
        context_parts.append(f"\n=== CURRENT DATAFRAME ===")
        context_parts.append(f"Shape: {df.shape[0]} rows × {df.shape[1]} columns")
        context_parts.append(f"Columns: {list(df.columns)}")
        
        # Data types
        context_parts.append(f"\nData Types:")
        for col, dtype in df.dtypes.items():
            context_parts.append(f"- {col}: {dtype}")
        
        # Sample data
        context_parts.append(f"\nFirst 3 rows:")
        context_parts.append(df.head(3).to_string())
        
        # Missing values
        missing = df.isnull().sum()
        if missing.sum() > 0:
            context_parts.append(f"\nMissing values:")
            for col, count in missing.items():
                if count > 0:
                    context_parts.append(f"- {col}: {count}")
        
        return "\n".join(context_parts)
    
    def clear_session(self, session_id: str):
        """Clear all data for a session"""
        if session_id in self.session_data:
            del self.session_data[session_id]
            logger.info(f"🗑️ Cleared session data: {session_id}")
    
    def _optimize_dtypes(self, df: pd.DataFrame) -> pd.DataFrame:
        """Optimize DataFrame data types to reduce memory usage"""
        df_optimized = df.copy()
        
        # Check if DataFrame is empty to avoid division by zero
        if len(df_optimized) == 0:
            return df_optimized
        
        for col in df_optimized.columns:
            col_type = df_optimized[col].dtype
            
            # Optimize numeric types
            if col_type in ['int64']:
                if df_optimized[col].min() >= -128 and df_optimized[col].max() <= 127:
                    df_optimized[col] = df_optimized[col].astype('int8')
                elif df_optimized[col].min() >= -32768 and df_optimized[col].max() <= 32767:
                    df_optimized[col] = df_optimized[col].astype('int16')
                elif df_optimized[col].min() >= -2147483648 and df_optimized[col].max() <= 2147483647:
                    df_optimized[col] = df_optimized[col].astype('int32')
            
            elif col_type in ['float64']:
                df_optimized[col] = pd.to_numeric(df_optimized[col], downcast='float')
            
            # Convert repeated strings to category
            elif col_type == 'object':
                if df_optimized[col].nunique() / len(df_optimized) < 0.5:  # Less than 50% unique
                    df_optimized[col] = df_optimized[col].astype('category')
        
        return df_optimized
    
    def _create_comprehensive_profile(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Create detailed data profile for LLM understanding"""
        
        profile = {
            "basic_info": {
                "rows": len(df),
                "columns": len(df.columns),
                "memory_usage_mb": df.memory_usage(deep=True).sum() / 1024 / 1024,
                "column_names": list(df.columns),
                "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()}
            },
            "data_quality": {
                "missing_values": df.isnull().sum().to_dict(),
                "missing_percentage": (df.isnull().sum() / len(df) * 100).to_dict(),
                "duplicate_rows": df.duplicated().sum(),
                "unique_values_per_column": {col: df[col].nunique() for col in df.columns}
            },
            "sample_data": {
                "head_5": df.head(5).to_dict('records'),
                "tail_5": df.tail(5).to_dict('records'),
                "random_sample_10": df.sample(min(10, len(df))).to_dict('records') if len(df) > 0 else []
            },
            "statistical_summary": {},
            "categorical_analysis": {},
            "numerical_analysis": {},
            "temporal_analysis": {},
            "relationships": {}
        }
        
        # Statistical summary
        try:
            desc = df.describe(include='all')
            profile["statistical_summary"] = self._clean_nan_values(desc.to_dict())
        except Exception as e:
            logger.warning(f"Could not generate statistical summary: {e}")
        
        # Analyze categorical columns
        categorical_cols = df.select_dtypes(include=['object', 'category']).columns
        for col in categorical_cols:
            try:
                value_counts = df[col].value_counts().head(10)
                profile["categorical_analysis"][col] = {
                    "unique_count": df[col].nunique(),
                    "most_frequent": value_counts.to_dict(),
                    "sample_values": df[col].dropna().unique()[:20].tolist()
                }
            except Exception as e:
                logger.warning(f"Error analyzing categorical column {col}: {e}")
        
        # Analyze numerical columns
        numerical_cols = df.select_dtypes(include=['number']).columns
        for col in numerical_cols:
            try:
                profile["numerical_analysis"][col] = {
                    "min": float(df[col].min()) if not pd.isna(df[col].min()) else None,
                    "max": float(df[col].max()) if not pd.isna(df[col].max()) else None,
                    "mean": float(df[col].mean()) if not pd.isna(df[col].mean()) else None,
                    "median": float(df[col].median()) if not pd.isna(df[col].median()) else None,
                    "std": float(df[col].std()) if not pd.isna(df[col].std()) else None,
                    "quartiles": df[col].quantile([0.25, 0.5, 0.75]).to_dict(),
                    "outliers_count": len(df[df[col] > df[col].quantile(0.95)]) if len(df) > 0 else 0
                }
            except Exception as e:
                logger.warning(f"Error analyzing numerical column {col}: {e}")
        
        # Analyze temporal columns
        datetime_cols = df.select_dtypes(include=['datetime64']).columns
        for col in datetime_cols:
            try:
                profile["temporal_analysis"][col] = {
                    "min_date": str(df[col].min()),
                    "max_date": str(df[col].max()),
                    "date_range_days": (df[col].max() - df[col].min()).days if not df[col].isna().all() else 0
                }
            except Exception as e:
                logger.warning(f"Error analyzing temporal column {col}: {e}")
        
        # Analyze relationships (correlations)
        if len(numerical_cols) > 1:
            try:
                corr_matrix = df[numerical_cols].corr()
                high_correlations = []
                for i in range(len(corr_matrix.columns)):
                    for j in range(i+1, len(corr_matrix.columns)):
                        corr_val = corr_matrix.iloc[i, j]
                        if abs(corr_val) > 0.7 and not pd.isna(corr_val):
                            high_correlations.append({
                                "column1": corr_matrix.columns[i],
                                "column2": corr_matrix.columns[j],
                                "correlation": float(corr_val)
                            })
                profile["relationships"]["high_correlations"] = high_correlations
            except Exception as e:
                logger.warning(f"Error analyzing correlations: {e}")
        
        # Clean all NaN values from the profile before returning
        return self._clean_nan_values(profile)
    
    def _clean_nan_values(self, obj):
        """Recursively clean NaN values from nested dictionaries and lists for JSON serialization"""
        import math
        
        if isinstance(obj, dict):
            return {k: self._clean_nan_values(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [self._clean_nan_values(item) for item in obj]
        elif isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
            return None
        elif pd.isna(obj):
            return None
        else:
            return obj
    
    def process_query(self, session_id: str, user_query: str) -> Dict[str, Any]:
        """
        Process user query using full LLM-based agentic approach
        
        Flow:
        1. Get full DataFrame from memory
        2. Check if query requires Excel file operations (highlight, filter, etc.)
        3. If Excel operations: Generate and return downloadable Excel file
        4. Otherwise: Create comprehensive context for LLM
        5. LLM generates Python code
        6. Execute code on full DataFrame
        7. Return results (data, plots, insights, or Excel files)
        """
        try:
            if session_id not in self.session_data:
                return {"success": False, "error": "No data loaded for this session"}
            
            # Get the appropriate DataFrame for analysis
            df = self.get_active_dataframe(session_id)
            if df is None:
                return {"success": False, "error": "No data available for analysis"}
            
            # Get context about the data structure
            session = self.session_data[session_id]
            files = session["files"]
            
            logger.info(f"🤖 Processing query for session {session_id}: '{user_query[:50]}...'")
            
            # Get conversation history for context
            conversation_history = session.get("conversation_history", [])
            
            # Get last generated file for sequential operations
            last_generated_file = session.get("last_generated_file")
            
            # Check if this is an Excel operation request (with conversation history and last file context)
            is_excel_op, operations = self.intent_parser.parse_intent(
                user_query, 
                df, 
                conversation_history,
                last_generated_file
            )
            
            if is_excel_op and operations:
                logger.info(f"📊 Detected Excel operation request with operations: {operations}")
                return self._handle_excel_operation(session_id, df, user_query, operations)
            
            # Build comprehensive context for LLM with multi-file awareness
            context = self._build_multi_file_context(session_id, df, user_query)
            
            # Generate code using LLM
            code_response = self._generate_code_with_llm(context, user_query)
            
            if not code_response["success"]:
                return code_response
            
            # Execute generated code on full DataFrame
            execution_result = self._execute_code_safely(df, code_response["code"], session_id)
            
            # Store conversation history in session data
            if session_id in self.session_data:
                self.session_data[session_id]["conversation_history"].append({
                    "user_query": user_query,
                    "generated_code": code_response["code"],
                    "execution_result": execution_result,
                    "timestamp": datetime.now().isoformat()
                })
            
            return {
                "success": True,
                "session_id": session_id,
                "query": user_query,
                "generated_code": code_response["code"],
                "explanation": code_response.get("explanation", ""),
                "result": execution_result,
                "data_shape": f"{len(df)} rows × {len(df.columns)} columns"
            }
            
        except Exception as e:
            logger.error(f"❌ Error processing query: {e}", exc_info=True)
            return {"success": False, "error": str(e)}
    
    def _build_llm_context(self, df: pd.DataFrame, profile: Dict, user_query: str) -> str:
        """Build comprehensive context for LLM including full data understanding"""
        
        context_parts = [
            "# FULL EXCEL DATA ANALYSIS CONTEXT",
            "",
            "## Dataset Overview:",
            f"- Shape: {len(df)} rows × {len(df.columns)} columns",
            f"- Memory Usage: {df.memory_usage(deep=True).sum() / 1024 / 1024:.1f}MB",
            f"- Columns: {list(df.columns)}",
            "",
            "## Column Details:"
        ]
        
        # Add detailed column information
        for col in df.columns:
            col_info = f"- **{col}** ({df[col].dtype})"
            if col in profile["numerical_analysis"]:
                stats = profile["numerical_analysis"][col]
                col_info += f" | Range: {stats.get('min', 'N/A')} to {stats.get('max', 'N/A')} | Mean: {stats.get('mean', 'N/A'):.2f}"
            elif col in profile["categorical_analysis"]:
                stats = profile["categorical_analysis"][col]
                col_info += f" | {stats['unique_count']} unique values | Top: {list(stats['most_frequent'].keys())[:3]}"
            
            # Add missing value info
            missing_pct = profile["data_quality"]["missing_percentage"].get(col, 0)
            if missing_pct > 0:
                col_info += f" | Missing: {missing_pct:.1f}%"
            
            context_parts.append(col_info)
        
        context_parts.extend([
            "",
            "## Sample Data (First 5 rows):"
        ])
        
        # Add sample data in a readable format
        sample_df = df.head(5)
        for i, row in sample_df.iterrows():
            row_str = " | ".join([f"{col}: {row[col]}" for col in df.columns[:6]])  # Limit to first 6 columns
            if len(df.columns) > 6:
                row_str += " | ..."
            context_parts.append(f"Row {i+1}: {row_str}")
        
        # Add correlations if any
        if profile["relationships"].get("high_correlations"):
            context_parts.extend([
                "",
                "## Key Relationships:",
            ])
            for corr in profile["relationships"]["high_correlations"][:5]:  # Top 5 correlations
                context_parts.append(f"- {corr['column1']} ↔ {corr['column2']}: {corr['correlation']:.2f}")
        
        context_parts.extend([
            "",
            "## Data Quality Issues:",
            f"- Duplicate rows: {profile['data_quality']['duplicate_rows']}",
            f"- Columns with missing values: {len([col for col, pct in profile['data_quality']['missing_percentage'].items() if pct > 0])}"
        ])
        
        context_parts.extend([
            "",
            f"## User Query: {user_query}",
            "",
            "## Instructions:",
            "1. You have access to the COMPLETE DataFrame 'df' with ALL data in memory",
            "2. Analyze the query and provide COMPREHENSIVE insights, not just basic answers",
            "3. Generate Python code that goes beyond simple operations - provide deep analysis",
            "4. Include statistical insights, patterns, trends, and business implications",
            "5. For visualizations, create meaningful charts that tell a story",
            "6. Always provide context and interpretation of results",
            "7. Use dynamic response generation to explain findings and insights",
            "8. Handle edge cases and provide robust analysis",
            "",
            "## Available DataFrame: 'df'",
            f"df.shape = {df.shape}",
            f"df.columns = {list(df.columns)}",
            "",
            "## Analysis Approach:",
            "- Don't just answer the question - provide valuable insights",
            "- Explain WHY the results matter",
            "- Include relevant statistics and context",
            "- Identify patterns, trends, and anomalies",
            "- Provide actionable recommendations when appropriate",
            "",
            "Generate comprehensive Python code with detailed analysis:"
        ])
        
        return "\n".join(context_parts)
    
    def _classify_query_intent_with_llm(self, user_query: str) -> str:
        """Use LLM to intelligently classify user query intent"""
        
        intent_prompt = f"""
You are an intelligent query classifier. Analyze the user's query and determine if they want a SIMPLE direct answer or COMPLEX comprehensive analysis.

User Query: "{user_query}"

Classification Guidelines:

**SIMPLE queries** are requests for:
- Basic information (counts, dimensions, simple facts)
- Direct answers to straightforward questions
- Quick lookups or basic data retrieval
- Simple mathematical operations
- Basic descriptive information
- Row/column counts, data shape, basic statistics
- Simple data retrieval without visualization
- Simple chart requests (single chart, basic visualization)

**COMPLEX queries** are requests for:
- Deep analysis, insights, or patterns
- Multiple visualizations with detailed analysis
- Comparisons and correlations with interpretation
- Business intelligence and recommendations
- Statistical analysis and interpretation
- Trend analysis and forecasting
- Strategic insights and decision support
- Comprehensive reports with multiple sections

Examples:
- "How many rows are there?" → SIMPLE
- "I need the row count" → SIMPLE
- "What are the column names?" → SIMPLE  
- "Show me the data size" → SIMPLE
- "Give me a chart based on status" → SIMPLE
- "Create a simple visualization" → SIMPLE
- "Show me a basic graph" → SIMPLE
- "Analyze sales trends over time with detailed insights" → COMPLEX
- "What insights can you provide about customer behavior?" → COMPLEX
- "Compare the performance of different regions with recommendations" → COMPLEX

Respond with ONLY one word: "simple" or "complex"
"""
        
        try:
            response = self.model.generate_content(
                intent_prompt,
                generation_config=genai.types.GenerationConfig(
                    temperature=0.1,  # Low temperature for consistent classification
                    top_p=0.8,
                    top_k=10
                )
            )
            
            classification = response.text.strip().lower()
            
            # Validate response
            if classification in ["simple", "complex"]:
                logger.info(f"🎯 LLM classified query '{user_query[:30]}...' as: {classification}")
                return classification
            else:
                # Fallback to complex if unclear
                logger.warning(f"⚠️ LLM classification unclear: '{classification}', defaulting to complex")
                return "complex"
                
        except Exception as e:
            logger.error(f"❌ Error in LLM intent classification: {e}")
            # Fallback to complex analysis for safety
            return "complex"
    
    def _generate_code_with_llm(self, context: str, user_query: str) -> Dict[str, Any]:
        """Generate Python code using LLM based on full data context"""
        
        # Use LLM to intelligently classify query intent
        intent_classification = self._classify_query_intent_with_llm(user_query)
        
        if intent_classification == "simple":
            # Generate concise response for simple queries
            prompt = f"""
{context}

Based on the dataset context above, generate Python code to answer this SIMPLE query: "{user_query}"

**SPECIAL INSTRUCTION FOR CHART REQUESTS**: If the user asks for a chart/graph/visualization, you MUST create the actual chart, not just explore the data.

**FOR CHART REQUESTS**: Generate complete working code that creates the requested chart type. Use this template and adapt it for the requested data:

```python
import json
import pandas as pd

try:
    # Find the appropriate column based on user request
    # For blood group: look for 'blood', 'group'
    # For gender: look for 'gender', 'sex'
    # For other requests: look for relevant keywords
    
    target_col = None
    search_terms = []  # Add appropriate search terms based on user query
    
    # Example for blood group chart:
    # search_terms = ['blood', 'group']
    # Example for gender chart:
    # search_terms = ['gender', 'sex']
    
    for col in df.columns:
        for term in search_terms:
            if term in col.lower():
                target_col = col
                break
        if target_col:
            break
    
    if target_col is None:
        # Generate dynamic response for no column found
        response = self._generate_dynamic_search_response(
            search_term="chart column",
            found_records=0,
            user_query=user_query,
            success=False
        )
        print(f"{{response}} Available columns: {{df.columns.tolist()}}")
    else:
        # Clean and get data counts
        column_data = df[target_col].astype(str).str.strip().str.upper()
        column_data = column_data.replace('NAN', 'UNKNOWN')
        column_data = column_data.replace('', 'UNKNOWN')
        
        data_counts = column_data.value_counts()
        
        # Prepare chart data for frontend (Recharts format)
        chart_data = []
        colors = ['#8884d8', '#82ca9d', '#ffc658', '#ff7300', '#00ff00', '#ff00ff', '#00ffff', '#ffff00', 
                  '#8dd1e1', '#d084d0', '#87d068', '#ffa940', '#ff7875', '#40a9ff', '#b37feb']
        
        for i, (label, value) in enumerate(data_counts.items()):
            chart_data.append({{
                'name': str(label),
                'value': int(value),
                'fill': colors[i % len(colors)]
            }})
        
        # Create chart configuration based on user request
        # Determine chart type from user query
        chart_type = 'pie'  # default
        if 'line' in user_query.lower() or 'line chart' in user_query.lower() or 'line graph' in user_query.lower():
            chart_type = 'line'
        elif 'bar' in user_query.lower() or 'bar chart' in user_query.lower():
            chart_type = 'bar'
        
        if chart_type == 'line':
            # For line charts, convert data to x,y format
            chart_data = []
            for i, (label, value) in enumerate(data_counts.items()):
                chart_data.append({{
                    'name': str(label),
                    'value': int(value)
                }})
            
            chart_config = {{
                'type': 'line',
                'title': f'{{target_col}} Distribution',
                'data': chart_data,
                'dataKey': 'value',
                'nameKey': 'name'
            }}
        else:
            # Default pie chart format
            chart_config = {{
                'type': chart_type,
                'title': f'{{target_col}} Distribution',
                'data': chart_data,
                'dataKey': 'value',
                'nameKey': 'name'
            }}
        
        # Output chart data as JSON for frontend
        print(f"[CHART_DATA]{{json.dumps(chart_config)}}")
        
        # Generate dynamic response for successful chart creation
        chart_response = self._generate_dynamic_chart_response(
            chart_type=chart_type,
            target_column=target_col,
            data_counts=data_counts.to_dict(),
            user_query=user_query
        )
        print(chart_response)

except Exception as e:
    print(f"Error: {{e}}")
    import traceback
    traceback.print_exc()
```

**IMPORTANT**: For chart requests, use the template above and:
1. Set appropriate search_terms based on the user's request (e.g., ['gender', 'sex'] for gender charts, ['blood', 'group'] for blood group charts)
2. The template will automatically find and use the correct column

## Requirements for Simple Queries:
1. **Be Professional**: Generate natural, conversational responses like a data analyst would
2. **Be Direct**: Answer the specific question asked, don't over-analyze
3. **No Markdown**: Avoid using ** or ## formatting - write naturally
4. **Use DataFrame 'df'**: Access the complete dataset for accurate results
5. **Libraries Available**: pandas as pd, numpy as np, matplotlib.pyplot as plt, seaborn as sns, os, datetime
6. **Single Answer**: Print only ONE final answer, avoid repetition or multiple formats
7. **Charts for Web**: If creating charts, ALWAYS save with plt.savefig() and plt.close()

## CRITICAL RULES:
1. **NEVER create sample data** - The DataFrame 'df' is already loaded with the user's actual data
2. **NEVER overwrite 'df'** - Use the existing DataFrame that contains the real uploaded data
3. **Use 'df' directly** - It contains the complete dataset from the uploaded file
4. **Multi-file queries**: When asked about "both files" or individual file statistics, you MUST provide information for ALL files listed in the context above. For example, if asked "row and column count for both files", respond with details for each file separately
5. **USE CONVERSATION HISTORY**: If the context includes conversation history, use it to understand what the user is referring to. For example, if they previously asked about "BUD000421995" and now ask "what's the trainer name", you should look up the trainer for BUD000421995 without them having to repeat it

## Code Structure for Simple Queries:
```python
try:
    # CRITICAL: Use the existing 'df' - DO NOT create sample data
    # The DataFrame 'df' already contains the user's uploaded data
    
    # Your analysis code here (example for row/column count):
    num_rows, num_columns = df.shape
    
    # For multi-file queries about individual file stats, extract from context and respond about ALL files:
    # Example: "The Applicants file has 53 rows and 9 columns, while the Learners file has 159 rows and 23 columns."
    
    # For charts (if requested), use this EXACT pattern:
    # import matplotlib.pyplot as plt
    # import os
    # plt.figure(figsize=(10, 6))
    # df['Status'].value_counts().plot(kind='bar', color='skyblue')
    # plt.title('Status Distribution')
    # plt.xlabel('Status')
    # plt.ylabel('Count')
    # plt.xticks(rotation=45)
    # plt.tight_layout()
    # chart_path = os.path.join(outputs_dir, "chart_status_distribution.png")
    # plt.savefig(chart_path, dpi=300, bbox_inches='tight')
    # plt.close()
    # if os.path.exists(chart_path):
    #     print(f"Chart saved to {{chart_path}}")
    # else:
    #     print(f"ERROR: Chart not saved to {{chart_path}}")
    
    # Generate your response dynamically based on the analysis
    # Print the actual results, insights, or answer to the user's question
    # Be natural and conversational - explain what you found
    
except Exception as e:
    print(f"Error: {{e}}")
```

**CRITICAL INSTRUCTIONS**:
1. **ALWAYS generate Python code** - Even for simple questions, you must provide executable Python code
2. **For chart requests**: Use the complete chart creation template above with [CHART_DATA] output
3. **For data lookups**: Generate code that searches the DataFrame and uses dynamic response generation
4. **For conversational queries**: Generate code that uses contextual, LLM-based responses

**EXAMPLE FOR NAME LOOKUPS**:
```python
try:
    # Search for student by name (case-insensitive)
    name_to_find = "MIDHUN"  # Replace with actual name from query
    
    # Search in all text columns for the name
    found_rows = []
    for col in df.columns:
        if df[col].dtype == 'object':  # Text columns
            mask = df[col].astype(str).str.upper().str.contains(name_to_find, na=False)
            if mask.any():
                found_rows.extend(df[mask].index.tolist())
    
    if found_rows:
        # Remove duplicates and get unique rows
        found_rows = list(set(found_rows))
        result_df = df.iloc[found_rows]
        
        # Generate dynamic response for successful search
        search_response = self._generate_dynamic_search_response(
            search_term=name_to_find,
            found_records=len(found_rows),
            user_query=user_query,
            success=True
        )
        print(search_response)
        
        for idx, row in result_df.iterrows():
            print(f"Row {{idx}}: {{row.to_dict()}}")
    else:
        # Generate dynamic response for no results found
        no_results_response = self._generate_dynamic_search_response(
            search_term=name_to_find,
            found_records=0,
            user_query=user_query,
            success=False
        )
        print(no_results_response)
        print(f"Available columns: {{df.columns.tolist()}}")
        
except Exception as e:
    print(f"Error searching for name: {{e}}")
```

Generate Python code that analyzes the data and provides results. Always wrap your response in ```python code blocks:
"""
        else:
            # Generate comprehensive response for complex queries
            prompt = f"""
{context}

Based on the comprehensive dataset context above, generate intelligent Python code to answer: "{user_query}"

## Code Generation Requirements:
1. **Data Access**: Use DataFrame 'df' containing the COMPLETE dataset - NEVER create sample data
2. **Libraries Available**: pandas as pd, numpy as np, matplotlib.pyplot as plt, seaborn as sns, os, datetime
3. **Analysis Depth**: Go beyond basic operations - provide comprehensive insights
4. **Statistical Rigor**: Include relevant statistics, distributions, correlations
5. **Visual Storytelling**: Create meaningful visualizations that reveal patterns
6. **Business Context**: Interpret results and explain their significance
7. **Dynamic Insights**: Use LLM-generated contextual responses to explain findings
8. **Web-Ready Charts**: ALWAYS save charts with plt.savefig() and plt.close() - NO plt.show()

## CRITICAL RULES:
- **NEVER create sample data** - The DataFrame 'df' is already loaded with the user's actual data
- **NEVER overwrite 'df'** - Use the existing DataFrame that contains the real uploaded data
- **Use 'df' directly** - It contains the complete dataset from the uploaded file
- **USE CONVERSATION HISTORY**: If the context includes conversation history, use it to understand what the user is referring to. Extract relevant information (IDs, names, values) from previous questions to provide contextual answers

## Code Structure:
```python
try:
    # 1. Data Analysis & Exploration
    # Perform your analysis here
    
    # 2. Core Analysis (answer the specific question)
    # Your main analysis code here
    
    # 3. Additional Insights & Context
    # Provide deeper insights, patterns, trends
    
    # 4. Statistical Summary
    # Include relevant statistics
    
    # 5. Visualizations (if applicable) - USE JSON CHARTS FOR INTERACTIVE DISPLAY
    # For charts, use the JSON format for interactive frontend display:
    # 
    # import json
    # 
    # # Example for pie chart:
    # column_data = df['COLUMN_NAME'].astype(str).str.strip().str.upper()
    # data_counts = column_data.value_counts().head(15)  # Limit categories
    # chart_data = []
    # colors = ['#8884d8', '#82ca9d', '#ffc658', '#ff7300', '#00ff00', '#ff00ff', '#00ffff', '#ffff00']
    # for i, (label, value) in enumerate(data_counts.items()):
    #     chart_data.append({{'name': str(label), 'value': int(value), 'fill': colors[i % len(colors)]}})
    # chart_config = {{'type': 'pie', 'title': 'COLUMN_NAME Distribution', 'data': chart_data, 'dataKey': 'value', 'nameKey': 'name'}}
    # print(f"[CHART_DATA]{{json.dumps(chart_config)}}")
    #
    # IMPORTANT: Replace 'COLUMN_NAME' with actual column name and choose appropriate chart type
    
    # 6. Print your comprehensive analysis results
    # Structure your output naturally based on what you found
    # Include: findings, insights, statistics, recommendations
    # Be conversational and professional
    
except Exception as e:
    print(f"Analysis Error: {{e}}")
```

## Response Guidelines:
- **Be Comprehensive**: Don't just answer - provide valuable insights
- **Explain Context**: Why do these results matter?
- **Identify Patterns**: Look for trends, outliers, correlations
- **Business Value**: What actions can be taken based on these insights?
- **Statistical Depth**: Include confidence intervals, significance tests when relevant
- **Visual Excellence**: Create publication-quality charts with proper labels

Generate intelligent, comprehensive Python code:
"""
        
        try:
            response = self.model.generate_content(
                prompt,
                generation_config=genai.types.GenerationConfig(
                    temperature=0.1,  # Low temperature for code generation
                    top_p=0.9,
                    top_k=40
                )
            )
            
            response_text = response.text
            
            # Extract code from response
            code_blocks = []
            lines = response_text.split('\n')
            in_code_block = False
            current_code = []
            
            for line in lines:
                if line.strip().startswith('```python'):
                    in_code_block = True
                    current_code = []
                elif line.strip() == '```' and in_code_block:
                    in_code_block = False
                    if current_code:
                        code_blocks.append('\n'.join(current_code))
                elif in_code_block:
                    current_code.append(line)
            
            if not code_blocks:
                return {"success": False, "error": "No Python code found in LLM response"}
            
            # Use the first (or combined) code block
            generated_code = '\n'.join(code_blocks)
            
            # Extract explanation (text after code blocks)
            explanation_parts = []
            skip_until_explanation = True
            for line in lines:
                if not skip_until_explanation:
                    explanation_parts.append(line)
                elif line.strip() == '```' and '```python' not in line:
                    skip_until_explanation = False
            
            explanation = '\n'.join(explanation_parts).strip()
            
            return {
                "success": True,
                "code": generated_code,
                "explanation": explanation,
                "raw_response": response_text
            }
            
        except Exception as e:
            logger.error(f"❌ Error generating code with LLM: {e}")
            return {"success": False, "error": f"LLM code generation failed: {str(e)}"}
    
    def _execute_code_safely(self, df: pd.DataFrame, code: str, session_id: str) -> Dict[str, Any]:
        """Execute generated Python code safely on the full DataFrame"""
        
        try:
            # Create safe execution environment with necessary imports pre-loaded
            import os
            import sys
            from io import StringIO
            
            # Configure matplotlib for web display (non-interactive backend)
            import matplotlib
            matplotlib.use('Agg')  # Set backend before importing pyplot
            plt.switch_backend('Agg')  # Use non-interactive backend
            plt.ioff()  # Turn off interactive mode
            
            safe_globals = {
                'df': df,
                'pd': pd,
                'np': np,
                'plt': plt,
                'sns': sns,
                'matplotlib': matplotlib,
                'datetime': datetime,
                'os': os,
                'outputs_dir': str(self.outputs_dir),
                'StringIO': StringIO,
                '__builtins__': {
                    'len': len,
                    'str': str,
                    'int': int,
                    'float': float,
                    'list': list,
                    'dict': dict,
                    'tuple': tuple,
                    'set': set,
                    'print': print,
                    'range': range,
                    'enumerate': enumerate,
                    'zip': zip,
                    'min': min,
                    'max': max,
                    'sum': sum,
                    'abs': abs,
                    'round': round,
                    'sorted': sorted,
                    'reversed': reversed,
                    'any': any,
                    'all': all,
                    'isinstance': isinstance,
                    'type': type,
                    'hasattr': hasattr,
                    'getattr': getattr,
                    'setattr': setattr,
                    '__import__': __import__,  # Allow imports
                    'ImportError': ImportError,
                    'ValueError': ValueError,
                    'TypeError': TypeError,
                    'KeyError': KeyError,
                    'IndexError': IndexError,
                    'Exception': Exception
                }
            }
            
            safe_locals = {}
            
            # Capture output by redirecting stdout
            output_buffer = io.StringIO()
            original_stdout = sys.stdout
            sys.stdout = output_buffer
            
            try:
                # Execute code
                logger.info(f"🔧 Executing generated code for session {session_id}")
                logger.info(f"📝 Generated code:\n{code}")
                
                # Ensure outputs directory exists
                self.outputs_dir.mkdir(exist_ok=True)
                
                # Count PNG files before execution
                png_files_before = len(list(self.outputs_dir.glob("*.png")))
                logger.info(f"📈 PNG files before execution: {png_files_before}")
                
                exec(code, safe_globals, safe_locals)
                
                # Count PNG files after execution
                png_files_after = len(list(self.outputs_dir.glob("*.png")))
                logger.info(f"📈 PNG files after execution: {png_files_after}")
                
                if png_files_after == png_files_before:
                    logger.warning("⚠️ No new PNG files created - chart generation may have failed silently")
                
            except Exception as exec_error:
                logger.error(f"❌ Code execution error: {exec_error}")
                import traceback
                error_traceback = traceback.format_exc()
                logger.error(f"❌ Full traceback: {error_traceback}")
                # Add error to output
                output_buffer.write(f"\nExecution Error: {exec_error}\nTraceback: {error_traceback}")
            finally:
                # Restore stdout
                sys.stdout = original_stdout
            
            # Collect results
            result = {
                "success": True,
                "execution_output": output_buffer.getvalue(),
                "variables": {},
                "charts": [],
                "dataframes": {}
            }
            
            # Extract meaningful results from locals
            for var_name, var_value in safe_locals.items():
                if not var_name.startswith('_'):
                    if isinstance(var_value, pd.DataFrame):
                        # Store DataFrame info and sample
                        result["dataframes"][var_name] = {
                            "shape": var_value.shape,
                            "columns": list(var_value.columns),
                            "sample": var_value.head(10).to_dict('records') if len(var_value) > 0 else [],
                            "dtypes": {col: str(dtype) for col, dtype in var_value.dtypes.items()}
                        }
                    elif isinstance(var_value, (int, float, str, bool, list, dict)):
                        result["variables"][var_name] = var_value
                    elif hasattr(var_value, '__str__') and len(str(var_value)) < 1000:
                        result["variables"][var_name] = str(var_value)
            
            # Check for chart data in output instead of PNG files
            output_text = result["execution_output"]
            chart_data_list = []
            
            # Extract chart data from output
            import re
            
            def extract_json_from_line(line):
                """Extract complete JSON object from a line containing [CHART_DATA]"""
                if '[CHART_DATA]' not in line:
                    return None
                
                json_start = line.find('[CHART_DATA]') + len('[CHART_DATA]')
                json_part = line[json_start:].strip()
                
                if not json_part.startswith('{'):
                    return None
                
                # Count braces to find the complete JSON object
                brace_count = 0
                end_pos = 0
                
                for i, char in enumerate(json_part):
                    if char == '{':
                        brace_count += 1
                    elif char == '}':
                        brace_count -= 1
                        if brace_count == 0:
                            end_pos = i + 1
                            break
                
                if end_pos > 0:
                    return json_part[:end_pos]
                return None
            
            chart_matches = []
            lines = output_text.split('\n')
            for line in lines:
                if '[CHART_DATA]' in line:
                    json_data = extract_json_from_line(line)
                    if json_data:
                        chart_matches.append(json_data)
            
            for match in chart_matches:
                try:
                    import json
                    chart_config = json.loads(match)
                    chart_data_list.append(chart_config)
                    logger.info(f"📊 Found chart data: {chart_config.get('type', 'unknown')} - {chart_config.get('title', 'untitled')}")
                except json.JSONDecodeError as e:
                    logger.warning(f"⚠️ Failed to parse chart data: {e}")
            
            result["charts"] = chart_data_list
            
            # Clean up chart data markers from output text
            cleaned_output = re.sub(r'\[CHART_DATA\][^\n]*', '', output_text)
            result["execution_output"] = cleaned_output.strip()
            
            if chart_data_list:
                logger.info(f"📊 Found {len(chart_data_list)} chart(s) in execution output")
            else:
                logger.info("📊 No chart data found in execution output")
            
            if result["charts"]:
                logger.info(f"📊 Chart data: {[c.get('title', 'untitled') for c in result['charts']]}")
            
            logger.info(f"✅ Code execution successful for session {session_id}")
            return result
            
        except Exception as e:
            logger.error(f"❌ Code execution failed: {e}")
            return {
                "success": False,
                "error": str(e),
                "traceback": traceback.format_exc()
            }
    
    def _load_all_sheets_from_last_file(self, session_id: str) -> Optional[List[Dict]]:
        """
        Load all sheets from last generated file
        Returns list of dicts with sheet_name and df, or None if no last file
        """
        try:
            last_file = self.session_data[session_id].get("last_generated_file")
            
            if not last_file or not last_file.get("path"):
                logger.info("📋 No last generated file found")
                return None
            
            file_path = Path(last_file["path"])
            if not file_path.exists():
                logger.warning(f"⚠️ Last generated file not found: {file_path}")
                return None
            
            logger.info(f"📂 Loading all sheets from: {file_path.name}")
            
            # Load all sheets from Excel file
            xl_file = pd.ExcelFile(file_path)
            sheets = []
            
            for sheet_name in xl_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                sheets.append({
                    "sheet_name": sheet_name,
                    "df": df
                })
                logger.info(f"  📄 Loaded sheet '{sheet_name}': {len(df)} rows")
            
            return sheets
            
        except Exception as e:
            logger.error(f"❌ Error loading sheets from last file: {e}")
            return None
    
    def _apply_formatting_to_multi_sheet_file(
        self,
        session_id: str,
        sheets: List[Dict],
        operations: Dict[str, Any],
        target_sheet: str,
        user_query: str
    ) -> Dict[str, Any]:
        """
        Apply formatting operations (highlighting) to a specific sheet in a multi-sheet file
        """
        try:
            from datetime import datetime
            import openpyxl
            from openpyxl.styles import PatternFill
            
            timestamp = int(datetime.now().timestamp())
            filename = f"excel_formatted_{timestamp}.xlsx"
            output_path = Path(self.outputs_dir) / filename
            
            logger.info(f"🎨 Applying formatting to '{target_sheet}' sheet in multi-sheet file")
            
            # Step 1: Write all sheets to Excel file
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet in sheets:
                    sheet["df"].to_excel(
                        writer,
                        sheet_name=sheet["sheet_name"],
                        index=False
                    )
                    logger.info(f"  ✅ Saved sheet '{sheet['sheet_name']}': {len(sheet['df'])} rows")
            
            # Step 2: Apply formatting to target sheet using openpyxl
            wb = openpyxl.load_workbook(output_path)
            ws = wb[target_sheet]
            
            # Get target sheet DataFrame for condition checking
            target_df = None
            for sheet in sheets:
                if sheet["sheet_name"].lower() == target_sheet.lower():
                    target_df = sheet["df"]
                    break
            
            if target_df is None:
                raise ValueError(f"Target sheet '{target_sheet}' not found")
            
            # Apply subtotals if present
            if 'subtotals' in operations:
                from openpyxl.styles import Font
                subtotal_spec = operations['subtotals']
                group_by = subtotal_spec.get('group_by')
                aggregate_column = subtotal_spec.get('aggregate_column')
                function = subtotal_spec.get('function', 'count').upper()
                
                if group_by and group_by in target_df.columns and aggregate_column and aggregate_column in target_df.columns:
                    logger.info(f"  📊 Applying subtotals: group_by={group_by}, aggregate={aggregate_column}, function={function}")
                    
                    # Get column indices
                    group_col_idx = list(target_df.columns).index(group_by) + 1
                    agg_col_idx = list(target_df.columns).index(aggregate_column) + 1
                    
                    # Track groups
                    groups = []
                    current_group = None
                    group_start_row = 2
                    
                    for row_idx in range(2, ws.max_row + 1):
                        group_value = ws.cell(row=row_idx, column=group_col_idx).value
                        
                        if current_group is not None and group_value != current_group:
                            groups.append({
                                'name': current_group,
                                'start_row': group_start_row,
                                'end_row': row_idx - 1,
                                'insert_at': row_idx
                            })
                            group_start_row = row_idx
                        
                        current_group = group_value
                    
                    if current_group is not None:
                        groups.append({
                            'name': current_group,
                            'start_row': group_start_row,
                            'end_row': ws.max_row,
                            'insert_at': ws.max_row + 1
                        })
                    
                    logger.info(f"  📊 Found {len(groups)} groups for subtotals")
                    
                    # Insert subtotal rows (in reverse)
                    for group in reversed(groups):
                        insert_row = group['insert_at']
                        group_name = group['name']
                        start_row = group['start_row']
                        end_row = group['end_row']
                        
                        ws.insert_rows(insert_row)
                        ws.cell(row=insert_row, column=group_col_idx, value=f"{group_name} Count")
                        
                        # Calculate value
                        if function == 'COUNT':
                            calculated_value = sum(1 for r in range(start_row, end_row + 1) 
                                                   if ws.cell(row=r, column=agg_col_idx).value is not None)
                        elif function == 'SUM':
                            calculated_value = sum(float(ws.cell(row=r, column=agg_col_idx).value or 0) 
                                                   for r in range(start_row, end_row + 1))
                        else:  # AVERAGE
                            values = [float(ws.cell(row=r, column=agg_col_idx).value or 0) 
                                     for r in range(start_row, end_row + 1)]
                            calculated_value = sum(values) / len(values) if values else 0
                        
                        # Set numeric value
                        cell = ws.cell(row=insert_row, column=agg_col_idx)
                        cell.value = calculated_value
                        cell.number_format = '0' if function == 'COUNT' else '0.00'
                        
                        # Format subtotal row
                        for col in range(1, ws.max_column + 1):
                            cell = ws.cell(row=insert_row, column=col)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color='FFE0E0E0', end_color='FFE0E0E0', fill_type='solid')
                        
                        logger.info(f"  ➕ Subtotal for '{group_name}': {calculated_value} items")
            
            # Apply highlight_rows if present
            if 'highlight_rows' in operations:
                highlight_specs = operations['highlight_rows']
                if not isinstance(highlight_specs, list):
                    highlight_specs = [highlight_specs]
                
                for spec in highlight_specs:
                    column = spec.get('column')
                    condition = spec.get('condition', {})
                    color = spec.get('color', 'yellow')
                    
                    if column not in target_df.columns:
                        logger.warning(f"  ⚠️ Column '{column}' not found in {target_sheet}")
                        continue
                    
                    # Map colors to RGB
                    color_map = {
                        'yellow': 'FFFF00',
                        'red': 'FF0000',
                        'green': '00FF00',
                        'blue': '0000FF',
                        'orange': 'FFA500',
                        'purple': '800080',
                        'pink': 'FFC0CB',
                        'cyan': '00FFFF',
                        'light_green': '90EE90',
                        'light_blue': 'ADD8E6',
                        'light_yellow': 'FFFFE0',
                        'light_red': 'FFB6C1',
                        'gray': '808080'
                    }
                    
                    fill_color = color_map.get(color, 'FFFF00')
                    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    
                    # Get column index
                    col_idx = list(target_df.columns).index(column) + 1
                    
                    # Apply highlighting to matching rows
                    operator = condition.get('operator', '==')
                    value = condition.get('value')
                    
                    for row_idx, cell_value in enumerate(target_df[column], start=2):  # Start at row 2 (after header)
                        should_highlight = False
                        
                        if operator == '==':
                            should_highlight = str(cell_value) == str(value)
                        elif operator == '!=':
                            should_highlight = str(cell_value) != str(value)
                        elif operator == 'contains':
                            should_highlight = str(value).lower() in str(cell_value).lower()
                        elif operator == 'in':
                            should_highlight = str(cell_value) in [str(v) for v in value]
                        
                        if should_highlight:
                            # Apply fill to entire row
                            for col in range(1, len(target_df.columns) + 1):
                                ws.cell(row=row_idx, column=col).fill = fill
                            logger.info(f"  🎨 Highlighted row {row_idx} in {color}")
            
            # Save workbook
            wb.save(output_path)
            wb.close()
            
            logger.info(f"✅ Formatting applied successfully: {filename}")
            
            # Update tracking
            self.session_data[session_id]["last_generated_file"] = {
                "filename": filename,
                "path": str(output_path),
                "type": "updated_multi_sheet",
                "sheets": [s["sheet_name"] for s in sheets],
                "operations": operations,
                "timestamp": datetime.now().isoformat(),
                "operation": "formatting"
            }
            
            self.session_data[session_id]["generated_files_history"].append({
                "filename": filename,
                "path": str(output_path),
                "type": "formatting",
                "timestamp": datetime.now().isoformat()
            })
            
            download_url = f"/api/download/{filename}"
            total_rows = sum(len(s["df"]) for s in sheets)
            summary = f"✅ Applied operations in '{target_sheet}' sheet. File has {len(sheets)} sheets ({', '.join([s['sheet_name'] for s in sheets])}) - {total_rows} total rows."
            
            return {
                "success": True,
                "session_id": session_id,
                "operation_type": "excel_file",
                "excel_file": filename,
                "download_url": download_url,
                "generated_code": "",
                "explanation": summary,
                "result": {
                    "success": True,
                    "execution_output": summary,
                    "excel_files": [download_url],
                    "plots": [],
                    "dataframes": {},
                    "variables": {}
                },
                "data_shape": f"{len(sheets)} sheets, {total_rows} total rows"
            }
            
        except Exception as e:
            logger.error(f"❌ Error applying formatting to multi-sheet file: {e}", exc_info=True)
            return {
                "success": False,
                "error": f"Failed to apply formatting: {str(e)}"
            }
    
    def _save_updated_multi_sheet_file(
        self,
        session_id: str,
        sheets: List[Dict],
        operations: Dict[str, Any],
        user_query: str
    ) -> Dict[str, Any]:
        """
        Save updated multi-sheet file with all sheets preserved
        """
        try:
            from datetime import datetime
            
            timestamp = int(datetime.now().timestamp())
            filename = f"excel_updated_{timestamp}.xlsx"
            output_path = Path(self.outputs_dir) / filename
            
            logger.info(f"💾 Saving updated file with {len(sheets)} sheets: {filename}")
            
            # Check if source file has formatting (from last_generated_file)
            source_file = None
            last_file = self.session_data[session_id].get("last_generated_file")
            if last_file and last_file.get("path"):
                source_path = Path(last_file["path"])
                # Preserve formatting from any multi-sheet file (merged, formatted, or updated)
                if source_path.exists() and last_file.get("type") in ["merged", "updated_multi_sheet"]:
                    source_file = source_path
                    logger.info(f"  📋 Source file exists, will preserve formatting: {source_path.name}")
            
            if source_file:
                # Preserve formatting by loading source file with openpyxl
                import openpyxl
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                wb = openpyxl.load_workbook(source_file)
                
                # Update each sheet's data while preserving formatting
                for sheet_info in sheets:
                    sheet_name = sheet_info["sheet_name"]
                    df = sheet_info["df"]
                    
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        
                        # Strategy: Update cells in-place to preserve formatting
                        # First, clear data from rows that will be removed
                        new_row_count = len(df)
                        old_row_count = ws.max_row - 1  # Exclude header
                        
                        if new_row_count < old_row_count:
                            # Delete extra rows from the end
                            rows_to_delete = old_row_count - new_row_count
                            ws.delete_rows(new_row_count + 2, rows_to_delete)
                        
                        # Update existing cells with new data (preserves formatting)
                        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
                            for c_idx, value in enumerate(row, start=1):
                                cell = ws.cell(row=r_idx, column=c_idx)
                                cell.value = value
                                # Keep existing formatting (fill, font, etc.)
                        
                        logger.info(f"  ✅ Updated sheet '{sheet_name}': {len(df)} rows (formatting preserved)")
                    else:
                        # Sheet doesn't exist in source, add it normally
                        ws = wb.create_sheet(sheet_name)
                        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                            for c_idx, value in enumerate(row, start=1):
                                ws.cell(row=r_idx, column=c_idx, value=value)
                        logger.info(f"  ✅ Added new sheet '{sheet_name}': {len(df)} rows")
                
                wb.save(output_path)
                wb.close()
            else:
                # No formatting to preserve, use pandas (faster)
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    for sheet in sheets:
                        sheet["df"].to_excel(
                            writer,
                            sheet_name=sheet["sheet_name"],
                            index=False
                        )
                        logger.info(f"  ✅ Saved sheet '{sheet['sheet_name']}': {len(sheet['df'])} rows")
            
            # Update last_generated_file tracking
            self.session_data[session_id]["last_generated_file"] = {
                "filename": filename,
                "path": str(output_path),
                "type": "updated_multi_sheet",
                "sheets": [s["sheet_name"] for s in sheets],
                "timestamp": datetime.now().isoformat(),
                "operation": "update"
            }
            
            # Add to generated files history
            self.session_data[session_id]["generated_files_history"].append({
                "filename": filename,
                "path": str(output_path),
                "type": "updated_multi_sheet",
                "timestamp": datetime.now().isoformat()
            })
            
            logger.info(f"📌 Tracked updated file: {filename}")
            
            # Generate summary
            total_rows = sum(len(s["df"]) for s in sheets)
            sheet_names = ", ".join(s["sheet_name"] for s in sheets)
            
            # Determine operation description
            operation_desc = []
            if operations.get('remove_last_row'):
                operation_desc.append("Removed last row")
            if 'delete_rows' in operations:
                operation_desc.append("Deleted rows")
            if 'filter' in operations:
                operation_desc.append("Filtered data")
            if 'sort' in operations:
                operation_desc.append("Sorted data")
            
            op_text = " and ".join(operation_desc) if operation_desc else "Applied operations"
            
            target_sheet = operations.get('target_sheet')
            if target_sheet:
                summary = f"✅ {op_text} in '{target_sheet}' sheet. File has {len(sheets)} sheets ({sheet_names}) - {total_rows} total rows."
            else:
                summary = f"✅ {op_text} to {len(sheets)} sheets ({sheet_names}) - {total_rows} total rows."
            
            download_url = f"/api/download/{filename}"
            
            # Store in conversation history
            self.session_data[session_id]["conversation_history"].append({
                "user_query": user_query,
                "operation_type": "multi_sheet_update",
                "sheets_updated": len(sheets),
                "excel_file": filename,
                "timestamp": datetime.now().isoformat()
            })
            
            return {
                "success": True,
                "session_id": session_id,
                "query": user_query,
                "operation_type": "excel_file",
                "excel_file": filename,
                "download_url": download_url,
                "generated_code": "",
                "explanation": summary,
                "result": {
                    "success": True,
                    "execution_output": summary,
                    "excel_files": [download_url],
                    "plots": [],
                    "dataframes": {},
                    "variables": {}
                },
                "data_shape": f"{len(sheets)} sheets, {total_rows} total rows"
            }
            
        except Exception as e:
            logger.error(f"❌ Error saving updated multi-sheet file: {e}", exc_info=True)
            return {
                "success": False,
                "error": f"Failed to save updated file: {str(e)}"
            }
    
    def _handle_excel_operation(
        self,
        session_id: str,
        df: pd.DataFrame,
        user_query: str,
        operations: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Handle Excel file operations (highlighting, filtering, merge files, etc.)
        and return downloadable Excel file OR Google Sheets URL
        """
        try:
            # ⚡ PRIORITY CHECK: Is this a Google Sheets source?
            # Check FIRST before any other logic to ensure Google Sheets operations are routed correctly
            is_google_sheet = False
            google_sheet_url = None
            
            if session_id in self.session_data:
                files = self.session_data[session_id].get("files", {})
                
                # Check all files for Google Sheets source
                for file_id, file_info in files.items():
                    if file_info.get("source_type") == "google_sheets":
                        is_google_sheet = True
                        google_sheet_url = file_info.get("spreadsheet_url")
                        logger.info(f"📊 Detected Google Sheets source at start: {google_sheet_url}")
                        break  # Found a Google Sheet, use it
            
            # If Google Sheets source detected, route to Google Sheets handler
            if is_google_sheet and google_sheet_url:
                logger.info(f"🔗 Routing to Google Sheets handler (detected at start)")
                
                # Handle multi-operations for Google Sheets
                if 'operations' in operations and isinstance(operations['operations'], list):
                    logger.info(f"🔄 Multi-operation request for Google Sheets: {len(operations['operations'])} operations")
                    # Merge all operations into a single dict
                    merged_ops = {}
                    for op in operations['operations']:
                        merged_ops.update(op)
                    operations = merged_ops
                
                # Apply operations to Google Sheet (pass session data for storage like Excel)
                result = self.excel_ops.apply_operations_to_google_sheet(
                    df=df,
                    operations=operations,
                    original_url=google_sheet_url,
                    session_data=self.session_data,
                    session_id=session_id
                )
                
                if result.get('success'):
                    # Return Google Sheets response
                    return {
                        "success": True,
                        "type": "google_sheet_operation",
                        "sheet_url": result['sheet_url'],
                        "original_url": result['original_url'],
                        "operations_applied": result.get('operations_applied', []),
                        "rows": result.get('rows', len(df)),
                        "columns": result.get('columns', len(df.columns)),
                        "message": self._generate_operation_success_message(
                            operations_applied=result.get('operations_applied', []),
                            user_query=user_query,
                            rows=result.get('rows', len(df)),
                            columns=result.get('columns', len(df.columns)),
                            sheet_type="Google Sheet"
                        )
                    }
                else:
                    return {
                        "success": False,
                        "error": result.get('error', 'Failed to apply operations to Google Sheet')
                    }
            
            # Not Google Sheets - continue with Excel operations
            logger.info(f"📄 Not Google Sheets source - proceeding with Excel operations")
            
            # Handle case where operations is a list (direct operations array)
            if isinstance(operations, list):
                logger.info(f"🔄 Direct operations list detected: {len(operations)} operations")
                # Convert list to expected format
                operations_dict = {"operations": operations}
                return self._handle_multi_operation_request(session_id, user_query, operations_dict, df)
            
            # Check if this is a merge_files operation
            if operations.get('merge_files'):
                return self._handle_merge_files_operation(session_id, user_query, operations)
            
            # Check if this is a rename_tabs operation
            if operations.get('rename_tabs'):
                return self._handle_rename_tabs_operation(session_id, user_query, operations)
            
            # Check if this is a multi-operation request (operations array)
            if 'operations' in operations and isinstance(operations['operations'], list):
                logger.info(f"🔄 Multi-operation request detected: {len(operations['operations'])} operations")
                return self._handle_multi_operation_request(session_id, user_query, operations, df)
            
            # AGENTIC FILE SOURCE DETERMINATION
            # Let the LLM decide whether to use last generated file or original files
            last_file = self.session_data[session_id].get("last_generated_file")
            conversation_history = self.session_data[session_id].get("conversation_history", [])
            
            file_source_decision = self.intent_parser.determine_file_source_with_llm(
                user_query=user_query,
                last_generated_file=last_file,
                conversation_history=conversation_history
            )
            
            logger.info(f"🎯 File Source Decision: {file_source_decision}")
            
            # Check if user wants to apply to all sheets OR target a specific sheet
            apply_to_all_sheets = operations.get('apply_to_all_sheets', False)
            target_sheet = operations.get('target_sheet')
            
            # AGENTIC FILE SELECTION: Use LLM decision to determine file source
            if file_source_decision == "last_generated" and last_file:
                # LLM decided to use last generated file
                logger.info(f"📋 Using last_generated_file: {last_file.get('filename')} (type: {last_file.get('type')}, sheets: {last_file.get('sheets')})")
                
                # Check if it's a multi-sheet file
                is_multi_sheet = False
                if last_file.get("sheets") and len(last_file["sheets"]) > 1:
                    is_multi_sheet = True
                elif last_file.get("path"):
                    file_path = Path(last_file["path"])
                    if file_path.exists():
                        try:
                            xl_file = pd.ExcelFile(file_path)
                            if len(xl_file.sheet_names) > 1:
                                is_multi_sheet = True
                                logger.info(f"  ✅ Detected multi-sheet file with {len(xl_file.sheet_names)} sheets")
                        except:
                            pass
                
                if is_multi_sheet:
                    logger.info(f"🔄 Sequential operation detected - updating last generated file: {last_file['filename']}")
                    
                    # Load all sheets from last generated file
                    all_sheets = self._load_all_sheets_from_last_file(session_id)
                    
                    if all_sheets:
                        # Import the operations function
                        from .multi_sheet_handler import _apply_operations_to_dataframe
                        
                        # Check if this is a formatting operation (highlighting, conditional format)
                        is_formatting_op = any(key in operations for key in ['highlight_rows', 'highlight_cells', 'conditional_format'])
                        
                        if is_formatting_op:
                            # For formatting operations, use dedicated multi-sheet formatting method
                            logger.info(f"  🎨 Detected formatting operation - using dedicated multi-sheet formatter")
                            
                            if target_sheet:
                                return self._apply_formatting_to_multi_sheet_file(
                                    session_id,
                                    all_sheets,
                                    operations,
                                    target_sheet,
                                    user_query
                                )
                            else:
                                return {
                                    "success": False,
                                    "error": "Formatting operations require a target sheet. Please specify which sheet to format."
                                }
                        
                        # For data operations (filter, sort, delete), use the existing logic
                        # Apply operations based on scope
                        if apply_to_all_sheets:
                            logger.info(f"  📊 Applying to ALL {len(all_sheets)} sheets")
                            # Apply to all sheets
                            for sheet in all_sheets:
                                sheet["df"] = _apply_operations_to_dataframe(sheet["df"], operations)
                        
                        elif target_sheet:
                            logger.info(f"  📄 Applying to target sheet: '{target_sheet}'")
                            # Apply to target sheet only, preserve others
                            found_target = False
                            for sheet in all_sheets:
                                if sheet["sheet_name"].lower() == target_sheet.lower():
                                    sheet["df"] = _apply_operations_to_dataframe(sheet["df"], operations)
                                    found_target = True
                                    logger.info(f"  ✅ Updated '{sheet['sheet_name']}' sheet")
                            
                            if not found_target:
                                return {
                                    "success": False,
                                    "error": f"Sheet '{target_sheet}' not found in last generated file. Available sheets: {[s['sheet_name'] for s in all_sheets]}"
                                }
                        
                        # Save updated multi-sheet file with all sheets preserved
                        return self._save_updated_multi_sheet_file(
                            session_id,
                            all_sheets,
                            operations,
                            user_query
                        )
                    else:
                        logger.warning("⚠️ Could not load sheets from last file, falling back to session data")
                
                # Fallback: Use session data (original behavior)
                if apply_to_all_sheets:
                    logger.info("📊 Applying to all sheets from session data")
                    return handle_multi_sheet_operation(
                        self.session_data,
                        session_id,
                        operations,
                        self.excel_ops,
                        Path(self.outputs_dir)
                    )
                
                # Handle single-sheet operations on multi-sheet file
                if target_sheet:
                    logger.info(f"📄 Getting sheet '{target_sheet}' from session data")
                    # Get DataFrame for specific sheet
                    df = self.get_dataframe_by_sheet_name(session_id, target_sheet)
                    if df is None:
                        return {
                            "success": False,
                            "error": f"Sheet '{target_sheet}' not found. Please check the sheet name."
                        }
                    logger.info(f"📄 Using sheet '{target_sheet}' for operations")
                
                # If not multi-sheet but LLM decided to use last generated file (single-sheet)
                elif not is_multi_sheet and last_file.get("path"):
                    last_file_path = Path(last_file["path"])
                    if last_file_path.exists():
                        logger.info(f"🔄 SEQUENTIAL OPERATION: Loading last generated file: {last_file_path.name}")
                        logger.info(f"  📋 Previous operations: {last_file.get('operations', {})}")
                        try:
                            # Use header detection for last generated files to handle files with title rows
                            header_row = self._detect_header_row(last_file_path)
                            df = pd.read_excel(last_file_path, header=header_row)
                            logger.info(f"  ✅ Loaded: {len(df)} rows × {len(df.columns)} columns (header at row {header_row})")
                            # Log first few rows to verify sort order
                            if 'sort' in str(last_file.get('operations', {})):
                                logger.info(f"  📊 Data preview (first 3 rows to verify sort order):")
                                for idx, row in df.head(3).iterrows():
                                    # Only log the first few columns to avoid clutter
                                    preview = {k: v for k, v in list(dict(row).items())[:3]}
                                    logger.info(f"    Row {idx}: {preview}...")
                        except Exception as e:
                            logger.warning(f"  ⚠️ Could not load last file: {e}, using session DataFrame")
                            df = self.get_active_dataframe(session_id)
                    else:
                        logger.warning(f"  ⚠️ Last file path doesn't exist, using session DataFrame")
                        df = self.get_active_dataframe(session_id)
            
            elif file_source_decision == "original_files":
                # LLM decided to use original uploaded files
                logger.info(f"📁 Using original uploaded files (session DataFrame)")
                
                # CRITICAL FIX: Check if this is a multi-sheet operation BEFORE getting combined DataFrame
                files = self.session_data[session_id]["files"]
                has_multi_sheets = any(file_data.get("is_sheet", False) for file_data in files.values())
                
                if has_multi_sheets and (target_sheet or apply_to_all_sheets):
                    logger.info(f"🔄 MULTI-SHEET OPERATION DETECTED - Preserving sheet structure")
                    logger.info(f"  📊 Target sheet: {target_sheet}")
                    logger.info(f"  📊 Apply to all: {apply_to_all_sheets}")
                    
                    # Use multi-sheet handler to preserve original structure
                    from .multi_sheet_handler import handle_multi_sheet_operation
                    return handle_multi_sheet_operation(
                        self.session_data,
                        session_id,
                        operations,
                        self.excel_ops,
                        Path(self.outputs_dir)
                    )
                else:
                    # Single sheet operation - use combined DataFrame
                    df = self.get_active_dataframe(session_id)
            
            # Generate filename based on operation type or custom name
            import time
            timestamp = int(time.time())
            
            # Check if user specified a custom filename
            if 'custom_filename' in operations and operations['custom_filename']:
                custom_name = operations['custom_filename']
                # Clean the custom name (remove special chars, keep alphanumeric and underscores)
                custom_name_clean = ''.join(c for c in custom_name if c.isalnum() or c in ('_', '-')).strip()
                if custom_name_clean:
                    filename = f"{custom_name_clean}_{timestamp}.xlsx"
                    logger.info(f"📝 Using custom filename: {filename}")
                else:
                    # If cleaning resulted in empty string, use default
                    filename = f"excel_data_{timestamp}.xlsx"
            else:
                # Auto-generate filename based on operation type
                op_type = "data"
                filter_desc = ""
                
                # Check for filtering to create more descriptive filename
                if 'filter' in operations:
                    filter_ops = operations['filter']
                    if filter_ops:
                        # Get first filter column and value for filename
                        first_col = list(filter_ops.keys())[0]
                        first_cond = filter_ops[first_col]
                        if isinstance(first_cond, dict):
                            val = str(first_cond.get('value', ''))[:20]  # Limit length
                        else:
                            val = str(first_cond)[:20]
                        # Clean value for filename (remove special chars)
                        val_clean = ''.join(c for c in val if c.isalnum() or c in (' ', '_')).replace(' ', '_')
                        filter_desc = f"_{first_col}_{val_clean}"
                        op_type = "filtered"
                
                if 'highlight_rows' in operations or 'highlight_cells' in operations:
                    op_type = "highlighted"
                elif 'conditional_format' in operations:
                    op_type = "formatted"
                elif 'sort' in operations and op_type == "data":
                    op_type = "sorted"
                
                filename = f"excel_{op_type}{filter_desc}_{timestamp}.xlsx"
            
            # Determine source file path for format preservation
            source_file_path = None
            if session_id in self.session_data:
                # PRIORITY 1: Use last generated file if it exists (for sequential operations)
                last_file = self.session_data[session_id].get("last_generated_file")
                if last_file and last_file.get("path"):
                    last_file_path = Path(last_file["path"])
                    if last_file_path.exists():
                        source_file_path = str(last_file_path)
                        logger.info(f"📋 Using last generated file as source: {last_file_path.name}")
                
                # PRIORITY 2: Use session files if no last generated file
                if not source_file_path:
                    files = self.session_data[session_id].get("files", {})
                    if files:
                        # Use the first file's path (or active file if set)
                        active_file = self.session_data[session_id].get("active_file")
                        if active_file and active_file in files:
                            source_file_path = files[active_file].get("file_path")
                        else:
                            # Use first file
                            first_file = list(files.values())[0]
                            source_file_path = first_file.get("file_path")
            
            # Create Excel file with operations (Google Sheets already handled at start)
            excel_path = self.excel_ops.create_excel_with_operations(
                    df=df,
                    operations=operations,
                    filename=filename,
                    source_file_path=source_file_path  # Pass source file for format preservation
                )
            
            # Get the filename from path
            excel_filename = Path(excel_path).name
            
            # Generate download URL
            download_url = f"/api/download/{excel_filename}"
            
            # Create a descriptive message about what was done
            operation_summary = self._generate_operation_summary(operations, df)
            
            # Store in conversation history
            if session_id in self.session_data:
                self.session_data[session_id]["conversation_history"].append({
                    "user_query": user_query,
                    "operation_type": "excel_file_generation",
                    "operations": operations,
                    "excel_file": excel_filename,
                    "timestamp": datetime.now().isoformat()
                })
                
                # Track last generated file for sequential operations
                self.session_data[session_id]["last_generated_file"] = {
                    "filename": excel_filename,
                    "path": excel_path,
                    "type": "excel_operation",
                    "operations": operations,
                    "timestamp": datetime.now().isoformat(),
                    "operation": "excel_operations"
                }
                
                # Add to generated files history
                self.session_data[session_id]["generated_files_history"].append({
                    "filename": excel_filename,
                    "path": excel_path,
                    "type": "excel_operation",
                    "timestamp": datetime.now().isoformat()
                })
                
                logger.info(f"📌 Tracked last generated file: {excel_filename}")
            
            logger.info(f"✅ Excel file created: {excel_filename}")
            
            return {
                "success": True,
                "session_id": session_id,
                "query": user_query,
                "operation_type": "excel_file",
                "excel_file": excel_filename,
                "download_url": download_url,
                "generated_code": "",  # No code execution for Excel operations
                "explanation": operation_summary,
                "result": {
                    "success": True,
                    "execution_output": operation_summary,
                    "excel_files": [download_url],
                    "plots": [],
                    "dataframes": {},
                    "variables": {}
                },
                "data_shape": f"{len(df)} rows × {len(df.columns)} columns"
            }
            
        except Exception as e:
            logger.error(f"❌ Error handling Excel operation: {e}", exc_info=True)
            return {
                "success": False,
                "error": f"Failed to create Excel file: {str(e)}"
            }
    
    def _handle_merge_files_operation(self, session_id: str, user_query: str, operations: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Handle merging multiple uploaded files into a single Excel file with multiple sheets
        Supports custom sheet names from user query
        
        AGENTIC ENHANCEMENT: Can merge with last generated file if user refers to it
        """
        try:
            if session_id not in self.session_data:
                return {"success": False, "error": "No session data found"}
            
            session = self.session_data[session_id]
            files = session.get("files", {})
            last_generated_file = session.get("last_generated_file")
            conversation_history = session.get("conversation_history", [])
            
            # AGENTIC: Use LLM to determine if user wants to merge with last generated file
            merge_with_last_file = False
            if last_generated_file and len(files) == 1:
                # User uploaded 1 new file and there's a last generated file
                # Let LLM decide if user wants to merge with it
                logger.info(f"🤖 Asking LLM: Should we merge with last generated file?")
                
                file_source_decision = self.intent_parser.determine_file_source_with_llm(
                    user_query=user_query,
                    last_generated_file=last_generated_file,
                    conversation_history=conversation_history
                )
                
                if file_source_decision == "last_generated":
                    merge_with_last_file = True
                    logger.info(f"🤖 LLM Decision: Merge with last generated file ({last_generated_file.get('filename')})")
                else:
                    logger.info(f"🤖 LLM Decision: Merge uploaded files only")
            
            # Check if we have enough files to merge
            total_files_to_merge = len(files)
            if merge_with_last_file:
                total_files_to_merge += 1  # Include last generated file
            
            if total_files_to_merge < 2:
                return {
                    "success": False,
                    "error": "You need to upload at least 2 files to merge them. Currently only 1 file is uploaded."
                }
            
            # Get custom sheet names if provided
            custom_sheet_names = None
            sheet_name_mapping = {}  # Map filename to sheet name
            
            if operations and 'sheet_names' in operations:
                custom_sheet_names = operations['sheet_names']
                if custom_sheet_names and len(custom_sheet_names) != len(files):
                    logger.warning(f"⚠️ Sheet names count ({len(custom_sheet_names)}) doesn't match files count ({len(files)}). Using filenames.")
                    custom_sheet_names = None
                elif custom_sheet_names:
                    # Create intelligent mapping of filenames to sheet names
                    available_sheet_names = custom_sheet_names.copy()
                    
                    for file_id, file_data in files.items():
                        original_filename = file_data["filename"]
                        matched = False
                        
                        # Try to match sheet name to filename
                        for sheet_name in available_sheet_names:
                            # Extract base name from filename (remove Output_ prefix and extension)
                            base_filename = original_filename.lower().replace('output_', '').replace('.xlsx', '').replace('.xls', '').replace('.csv', '').replace('_', '').replace(' ', '')
                            base_sheet_name = sheet_name.lower().replace('_', '').replace(' ', '')
                            
                            if base_sheet_name in base_filename or base_filename in base_sheet_name:
                                sheet_name_mapping[file_id] = sheet_name
                                available_sheet_names.remove(sheet_name)
                                matched = True
                                logger.info(f"📝 Matched '{original_filename}' → sheet '{sheet_name}'")
                                break
                        
                        if not matched and available_sheet_names:
                            # Fallback: use first available sheet name
                            sheet_name_mapping[file_id] = available_sheet_names.pop(0)
                            logger.info(f"📝 Using sheet name '{sheet_name_mapping[file_id]}' for '{original_filename}' (no match found)")
            
            # Generate filename
            import time
            timestamp = int(time.time())
            filename = f"excel_merged_{total_files_to_merge}_files_{timestamp}.xlsx"
            excel_path = self.outputs_dir / filename
            
            # Prepare list of files/sheets to merge
            sheets_to_merge = []
            
            # Add last generated file first (if merging with it)
            if merge_with_last_file:
                last_file_path = Path(last_generated_file["path"])
                if last_file_path.exists():
                    try:
                        # Load all sheets from last generated file
                        xl_file = pd.ExcelFile(last_file_path)
                        for sheet_name in xl_file.sheet_names:
                            df_sheet = pd.read_excel(last_file_path, sheet_name=sheet_name)
                            sheets_to_merge.append({
                                "df": df_sheet,
                                "sheet_name": sheet_name,
                                "source": "last_generated"
                            })
                            logger.info(f"📋 Including sheet '{sheet_name}' from last generated file ({len(df_sheet)} rows)")
                    except Exception as e:
                        logger.warning(f"⚠️ Could not load last generated file: {e}")
            
            # Add uploaded files
            for file_id, file_data in files.items():
                df = file_data["df"]
                original_filename = file_data["filename"]
                
                # Use mapped sheet name if available, otherwise use filename
                if file_id in sheet_name_mapping:
                    sheet_name = sheet_name_mapping[file_id][:31]
                else:
                    sheet_name = original_filename[:31]
                
                # Clean sheet name (remove invalid characters)
                invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
                for char in invalid_chars:
                    sheet_name = sheet_name.replace(char, '_')
                
                # Remove file extension from sheet name if present
                if sheet_name.endswith('.xlsx'):
                    sheet_name = sheet_name[:-5]
                elif sheet_name.endswith('.xls'):
                    sheet_name = sheet_name[:-4]
                elif sheet_name.endswith('.csv'):
                    sheet_name = sheet_name[:-4]
                
                sheets_to_merge.append({
                    "df": df,
                    "sheet_name": sheet_name,
                    "source": "uploaded"
                })
            
            # Create Excel writer and write all sheets
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                for sheet_data in sheets_to_merge:
                    df = sheet_data["df"]
                    sheet_name = sheet_data["sheet_name"]
                    source = sheet_data["source"]
                    
                    # Write DataFrame to sheet
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    logger.info(f"📄 Added sheet '{sheet_name}' from {source} ({len(df)} rows)")
            
            logger.info(f"✅ Merged {total_files_to_merge} files/sheets into: {filename}")
            
            # Store sheet data in session for subsequent operations
            # Clear existing files and add sheets
            sheet_files = {}
            for idx, sheet_data in enumerate(sheets_to_merge):
                df = sheet_data["df"]
                sheet_name = sheet_data["sheet_name"]
                source = sheet_data["source"]
                
                # Store as sheet
                sheet_file_id = f"sheet_{sheet_name}_{idx}"
                sheet_files[sheet_file_id] = {
                    "df": df.copy(),
                    "filename": f"{sheet_name}.xlsx",
                    "sheet_name": sheet_name,
                    "is_sheet": True,
                    "sheet_index": idx,
                    "file_path": str(excel_path),
                    "source": source
                }
            
            # Update session with sheet data
            self.session_data[session_id]["files"] = sheet_files
            logger.info(f"📊 Stored {len(sheet_files)} sheets in session for subsequent operations")
            
            # Generate download URL
            download_url = f"/api/download/{filename}"
            
            # Generate LLM summary for merge operation
            file_list = [f"{sheet_data['sheet_name']} ({len(sheet_data['df'])} rows)" for sheet_data in sheets_to_merge]
            total_rows = sum(len(sheet_data['df']) for sheet_data in sheets_to_merge)
            
            merge_context = ""
            if merge_with_last_file:
                merge_context = f"\nNote: Merged with last generated file ({last_generated_file.get('filename')})"
            
            prompt = f"""You are a helpful data assistant. Generate a brief, specific message about the file merge operation that was just completed.

Sheets merged:
{chr(10).join([f"- {f}" for f in file_list])}
{merge_context}

Total sheets: {len(sheets_to_merge)}
Total rows across all sheets: {total_rows}

Instructions:
1. Focus on the merge operation details
2. Mention how many files were merged
3. Mention that each file is in a separate tab/sheet
4. Keep it brief (1-2 sentences)
5. DO NOT use markdown formatting like ** or ###
6. Use plain text only
7. You can use emojis like ✅ 📊 📄
8. DO NOT start with generic phrases like "Okay", "Done", "Ready", "Your file is ready"
9. Start directly with the merge details

Example format:
- "✅ Merged {len(files)} files into one workbook with separate tabs - {total_rows} total rows."
- "📊 Combined {len(files)} files into separate sheets - {total_rows} rows across all data."

Generate the message:"""
            
            try:
                import google.generativeai as genai
                response = self.model.generate_content(
                    prompt,
                    generation_config=genai.types.GenerationConfig(
                        temperature=0.7,
                        top_p=0.9,
                        top_k=40
                    )
                )
                summary = response.text.strip()
            except Exception as e:
                logger.error(f"❌ Error generating merge summary: {e}")
                summary = f"✅ Merged {len(files)} files into one Excel workbook! Each file is in a separate tab. Total: {total_rows} rows."
            
            # Store in conversation history
            self.session_data[session_id]["conversation_history"].append({
                "user_query": user_query,
                "operation_type": "merge_files",
                "files_merged": len(files),
                "excel_file": filename,
                "timestamp": datetime.now().isoformat()
            })
            
            # Track last generated file for sequential operations
            self.session_data[session_id]["last_generated_file"] = {
                "filename": filename,
                "path": str(excel_path),
                "type": "merged",
                "sheets": [sheet_data["sheet_name"] for sheet_data in sheet_files.values()],
                "timestamp": datetime.now().isoformat(),
                "operation": "merge_files"
            }
            
            # Add to generated files history
            self.session_data[session_id]["generated_files_history"].append({
                "filename": filename,
                "path": str(excel_path),
                "type": "merged",
                "timestamp": datetime.now().isoformat()
            })
            
            logger.info(f"📌 Tracked last generated file: {filename}")
            
            return {
                "success": True,
                "session_id": session_id,
                "query": user_query,
                "operation_type": "excel_file",
                "excel_file": filename,
                "download_url": download_url,
                "generated_code": "",
                "explanation": summary,
                "result": {
                    "success": True,
                    "execution_output": summary,
                    "excel_files": [download_url],
                    "plots": [],
                    "dataframes": {},
                    "variables": {}
                },
                "data_shape": f"{len(files)} files merged, {total_rows} total rows"
            }
            
        except Exception as e:
            logger.error(f"❌ Error merging files: {e}", exc_info=True)
            return {
                "success": False,
                "error": f"Failed to merge files: {str(e)}"
            }
    
    def _handle_rename_tabs_operation(
        self,
        session_id: str,
        user_query: str,
        operations: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Handle renaming tabs in the last generated Excel file or uploaded file
        """
        try:
            from openpyxl import load_workbook
            
            # Get last generated file
            last_file = self.session_data[session_id].get("last_generated_file")
            
            # If no generated file, try to find an uploaded Excel file with multiple sheets
            if not last_file:
                logger.info("📂 No generated file found, checking for uploaded Excel files...")
                
                # Look for uploaded files with multiple sheets
                files = self.session_data[session_id].get("files", {})
                uploaded_excel_file = None
                
                for file_id, file_data in files.items():
                    # Check if this is a sheet from a multi-sheet Excel file
                    if file_data.get("is_sheet", False) and file_data.get("original_file_path"):
                        file_path = Path(file_data["original_file_path"])
                        if file_path.suffix.lower() in ['.xlsx', '.xls'] and file_path.exists():
                            uploaded_excel_file = {
                                "path": str(file_path),
                                "filename": file_data["original_file"]
                            }
                            logger.info(f"✅ Found uploaded multi-sheet Excel file: {file_data['original_file']}")
                            break
                    # Check if this is a single-sheet uploaded Excel file
                    elif file_data.get("file_path") and not file_data.get("is_sheet", False):
                        file_path = Path(file_data["file_path"])
                        if file_path.suffix.lower() in ['.xlsx', '.xls'] and file_path.exists():
                            uploaded_excel_file = {
                                "path": str(file_path),
                                "filename": file_data["filename"]
                            }
                            logger.info(f"✅ Found uploaded Excel file: {file_data['filename']}")
                            break
                
                if not uploaded_excel_file:
                    return {
                        "success": False,
                        "error": "No Excel file found. Please upload an Excel file or generate one first."
                    }
                
                last_file = uploaded_excel_file
            
            source_path = Path(last_file["path"])
            
            if not source_path.exists():
                return {
                    "success": False,
                    "error": f"Source file not found: {source_path.name}"
                }
            
            logger.info(f"📝 Renaming tabs in file: {source_path.name}")
            
            # Load the workbook
            wb = load_workbook(source_path)
            rename_map = operations.get('rename_tabs', {})
            
            # Get sheet names in order
            sheet_names = wb.sheetnames
            
            # Apply renames
            renamed_count = 0
            for sheet_key, new_name in rename_map.items():
                # Extract sheet index from key (e.g., "sheet_0" -> 0)
                if sheet_key.startswith('sheet_'):
                    try:
                        sheet_index = int(sheet_key.split('_')[1])
                        if sheet_index < len(sheet_names):
                            old_name = sheet_names[sheet_index]
                            ws = wb[old_name]
                            ws.title = new_name
                            logger.info(f"  ✅ Renamed '{old_name}' → '{new_name}'")
                            renamed_count += 1
                    except (ValueError, IndexError, KeyError) as e:
                        logger.warning(f"  ⚠️ Could not rename {sheet_key}: {e}")
            
            # Save to new file
            timestamp = int(datetime.now().timestamp())
            filename = f"excel_renamed_tabs_{timestamp}.xlsx"
            excel_path = Path(self.outputs_dir) / filename
            
            wb.save(excel_path)
            wb.close()
            
            logger.info(f"✅ Renamed {renamed_count} tabs in: {filename}")
            
            # Update session with renamed sheets
            # Update last_generated_file tracking
            self.session_data[session_id]["last_generated_file"] = {
                "filename": filename,
                "path": str(excel_path),
                "type": "renamed_tabs",
                "sheets": list(rename_map.values()),
                "timestamp": datetime.now().isoformat(),
                "operation": "rename_tabs"
            }
            
            # Add to generated files history
            self.session_data[session_id]["generated_files_history"].append({
                "filename": filename,
                "path": str(excel_path),
                "type": "renamed_tabs",
                "timestamp": datetime.now().isoformat()
            })
            
            logger.info(f"📌 Tracked last generated file: {filename}")
            
            download_url = f"/api/download/{filename}"
            summary = f"✅ Renamed {renamed_count} tab(s) in the Excel file. New tab names: {', '.join(rename_map.values())}"
            
            # Store in conversation history
            self.session_data[session_id]["conversation_history"].append({
                "user_query": user_query,
                "operation_type": "rename_tabs",
                "tabs_renamed": renamed_count,
                "excel_file": filename,
                "timestamp": datetime.now().isoformat()
            })
            
            return {
                "success": True,
                "session_id": session_id,
                "query": user_query,
                "operation_type": "excel_file",
                "excel_file": filename,
                "download_url": download_url,
                "generated_code": "",
                "explanation": summary,
                "result": {
                    "success": True,
                    "execution_output": summary,
                    "excel_files": [download_url],
                    "plots": [],
                    "dataframes": {},
                    "variables": {}
                },
                "data_shape": f"{renamed_count} tabs renamed"
            }
            
        except Exception as e:
            logger.error(f"❌ Error renaming tabs: {e}", exc_info=True)
            return {
                "success": False,
                "error": f"Failed to rename tabs: {str(e)}"
            }
    
    def _handle_multi_operation_request(
        self,
        session_id: str,
        user_query: str,
        operations_spec: Dict[str, Any],
        df: pd.DataFrame
    ) -> Dict[str, Any]:
        """
        Handle multiple operations in sequence
        Each operation can target different sheets or all sheets
        """
        try:
            operations_list = operations_spec['operations']
            logger.info(f"🔄 Processing {len(operations_list)} operations sequentially")
            
            # Import required for multi-sheet handling
            from .multi_sheet_handler import handle_multi_sheet_operation
            import time
            
            # Check if first operation is merge_files
            if operations_list and operations_list[0].get('merge_files'):
                logger.info("🔀 First operation is merge_files - executing merge first")
                merge_op = operations_list[0]
                
                # Execute merge operation
                merge_result = self._handle_merge_files_operation(session_id, user_query, merge_op)
                
                if not merge_result.get('success'):
                    return merge_result
                
                # If there are more operations after merge, apply them to the merged file
                if len(operations_list) > 1:
                    logger.info(f"📋 Applying {len(operations_list) - 1} additional operations to merged file")
                    
                    # Create a new operations spec with remaining operations
                    remaining_ops = {'operations': operations_list[1:]}
                    
                    # Get the merged file's DataFrame (combined from all sheets)
                    # We need to reload the sheets data after merge
                    files = self.session_data[session_id]["files"]
                    sheets_data = []
                    
                    for file_id, file_data in files.items():
                        if file_data.get("is_sheet", False):
                            sheets_data.append({
                                "sheet_name": file_data["sheet_name"],
                                "df": file_data["df"].copy(),
                                "sheet_index": file_data["sheet_index"],
                                "file_id": file_id
                            })
                    
                    sheets_data.sort(key=lambda x: x["sheet_index"])
                    
                    if sheets_data:
                        logger.info(f"📊 Merged file has {len(sheets_data)} sheets - applying remaining operations")
                        # Continue with remaining operations on the multi-sheet file
                        operations_list = remaining_ops['operations']
                    else:
                        # No sheets found, just return merge result
                        return merge_result
                else:
                    # Only merge operation, return merge result
                    return merge_result
            
            # Get all sheets from session
            files = self.session_data[session_id]["files"]
            sheets_data = []
            
            for file_id, file_data in files.items():
                if file_data.get("is_sheet", False):
                    sheets_data.append({
                        "sheet_name": file_data["sheet_name"],
                        "df": file_data["df"].copy(),
                        "sheet_index": file_data["sheet_index"],
                        "file_id": file_id
                    })
            
            # Sort by sheet index
            sheets_data.sort(key=lambda x: x["sheet_index"])
            
            if not sheets_data:
                # No multi-sheet file, use single DataFrame
                logger.info("📄 No multi-sheet file detected, processing as single file")
                # Process data operations on single DataFrame (filter, sort, delete, etc.)
                result_df = df.copy()
                for idx, op in enumerate(operations_list, 1):
                    logger.info(f"  Operation {idx}/{len(operations_list)}: {list(op.keys())}")
                    result_df = self._apply_single_operation_to_df(result_df, op)
                
                # Create Excel file with final result
                # AGENTIC APPROACH: LLM already generated correct format with validation
                # No manual transformation needed - use operations_spec directly
                timestamp = int(time.time())
                filename = f"excel_multi_op_{timestamp}.xlsx"
                
                # Try to create Excel file with retry on error
                max_retries = 2
                for attempt in range(max_retries):
                    try:
                        excel_path = self.excel_ops.create_excel_with_operations(
                            df=result_df,
                            operations=operations_spec,  # LLM-validated operations
                            filename=filename
                        )
                        break  # Success!
                    except Exception as e:
                        if attempt < max_retries - 1:
                            # AGENTIC ERROR CORRECTION: Ask LLM to fix the operations
                            logger.warning(f"⚠️ Excel creation failed (attempt {attempt + 1}/{max_retries}): {e}")
                            logger.info("🔧 Asking LLM to correct operations based on error...")
                            
                            corrected_ops = self.intent_parser.correct_operations_from_error(
                                operations=operations_spec,
                                error_message=str(e),
                                df=result_df,
                                user_query=user_query
                            )
                            operations_spec = corrected_ops
                        else:
                            # Final attempt failed
                            raise
                
                download_url = f"/api/download/{Path(excel_path).name}"
                summary = f"✅ Applied {len(operations_list)} operations - {len(result_df)} rows × {len(result_df.columns)} columns"
                
                return {
                    "success": True,
                    "session_id": session_id,
                    "operation_type": "excel_file",
                    "excel_file": Path(excel_path).name,
                    "download_url": download_url,
                    "generated_code": "",
                    "explanation": summary,
                    "result": {
                        "success": True,
                        "execution_output": summary,
                        "excel_files": [download_url],
                        "plots": [],
                        "dataframes": {},
                        "variables": {}
                    },
                    "data_shape": f"{len(result_df)} rows × {len(result_df.columns)} columns"
                }
            
            # Multi-sheet file - process each operation
            logger.info(f"📊 Multi-sheet file detected: {len(sheets_data)} sheets")
            
            # Process operations sequentially
            for idx, operation in enumerate(operations_list, 1):
                logger.info(f"🔄 Processing operation {idx}/{len(operations_list)}: {list(operation.keys())}")
                
                if operation.get('apply_to_all_sheets'):
                    # Apply to all sheets
                    logger.info(f"  → Applying to ALL sheets")
                    for sheet_info in sheets_data:
                        sheet_info["df"] = self._apply_single_operation_to_df(
                            sheet_info["df"],
                            operation
                        )
                
                elif 'target_sheet' in operation:
                    # Apply to specific sheet
                    target_sheet = operation['target_sheet']
                    logger.info(f"  → Applying to sheet '{target_sheet}'")
                    
                    for sheet_info in sheets_data:
                        if sheet_info["sheet_name"].lower() == target_sheet.lower():
                            sheet_info["df"] = self._apply_single_operation_to_df(
                                sheet_info["df"],
                                operation
                            )
                            logger.info(f"  ✅ Applied to '{sheet_info['sheet_name']}'")
                            break
                else:
                    # No target specified - apply to all sheets
                    logger.info(f"  → No target specified, applying to ALL sheets")
                    for sheet_info in sheets_data:
                        sheet_info["df"] = self._apply_single_operation_to_df(
                            sheet_info["df"],
                            operation
                        )
            
            # Create multi-sheet Excel file with all processed sheets
            timestamp = int(time.time())
            filename = f"excel_multi_op_{timestamp}.xlsx"
            output_path = Path(self.outputs_dir) / filename
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_info in sheets_data:
                    sheet_info["df"].to_excel(
                        writer,
                        sheet_name=sheet_info["sheet_name"],
                        index=False
                    )
            
            # Apply freeze panes if any operation requested it
            self._apply_freeze_panes_from_operations(output_path, operations_list, sheets_data)
            
            # Apply all formatting operations (auto filters, subtotals, highlighting, etc.)
            self._apply_formatting_operations_to_multi_sheet(output_path, operations_list, sheets_data)
            
            download_url = f"/api/download/{filename}"
            total_rows = sum(len(s["df"]) for s in sheets_data)
            
            # Generate dynamic summary using LLM based on actual operations performed
            summary = self._generate_dynamic_operations_summary(
                operations_list=operations_list,
                user_query=user_query,
                sheets_count=len(sheets_data),
                total_rows=total_rows
            )
            
            logger.info(f"✅ Multi-operation complete: {filename}")
            
            # Track last generated file for sequential operations
            if session_id in self.session_data:
                self.session_data[session_id]["last_generated_file"] = {
                    "filename": filename,
                    "path": str(output_path),
                    "type": "multi_operation",
                    "sheets": [s["sheet_name"] for s in sheets_data],
                    "operations": operations_list,
                    "timestamp": datetime.now().isoformat(),
                    "operation": "multi_operations"
                }
                
                # Add to generated files history
                self.session_data[session_id]["generated_files_history"].append({
                    "filename": filename,
                    "path": str(output_path),
                    "type": "multi_operation",
                    "timestamp": datetime.now().isoformat()
                })
                
                logger.info(f"📌 Tracked last generated file: {filename}")
            
            return {
                "success": True,
                "session_id": session_id,
                "operation_type": "excel_file",
                "excel_file": filename,
                "download_url": download_url,
                "generated_code": "",
                "explanation": summary,
                "result": {
                    "success": True,
                    "execution_output": summary,
                    "excel_files": [download_url],
                    "plots": [],
                    "dataframes": {},
                    "variables": {}
                },
                "data_shape": f"{len(sheets_data)} sheets, {total_rows} total rows"
            }
            
        except Exception as e:
            logger.error(f"❌ Error in multi-operation request: {e}", exc_info=True)
            return {
                "success": False,
                "error": f"Failed to process multi-operation request: {str(e)}"
            }
    
    def _apply_single_operation_to_df(self, df: pd.DataFrame, operation: Dict[str, Any]) -> pd.DataFrame:
        """Apply a single operation to a DataFrame"""
        df_result = df.copy()
        
        # Remove apply_to_all_sheets and target_sheet keys as they're for routing only
        op = {k: v for k, v in operation.items() if k not in ['apply_to_all_sheets', 'target_sheet']}
        
        # Remove last row
        if op.get('remove_last_row'):
            if len(df_result) > 0:
                df_result = df_result.iloc[:-1]
                logger.info(f"    ✂️ Removed last row: {len(df)} → {len(df_result)} rows")
        
        # Delete rows
        if 'delete_rows' in op:
            delete_spec = op['delete_rows']
            rows_before = len(df_result)
            
            if 'row_numbers' in delete_spec:
                row_numbers = delete_spec['row_numbers']
                indices_to_drop = [r - 1 for r in row_numbers if 0 <= r - 1 < len(df_result)]
                df_result = df_result.drop(df_result.index[indices_to_drop])
                
            elif 'column' in delete_spec and 'condition' in delete_spec:
                column = delete_spec['column']
                condition = delete_spec['condition']
                
                if column in df_result.columns:
                    operator = condition.get('operator', '==')
                    value = condition.get('value')
                    
                    if operator == '==':
                        mask = df_result[column] == value
                    elif operator == '!=':
                        mask = df_result[column] != value
                    elif operator == 'in':
                        mask = df_result[column].isin(value if isinstance(value, list) else [value])
                    elif operator == '>':
                        mask = df_result[column] > value
                    elif operator == '<':
                        mask = df_result[column] < value
                    elif operator == '>=':
                        mask = df_result[column] >= value
                    elif operator == '<=':
                        mask = df_result[column] <= value
                    else:
                        mask = pd.Series([False] * len(df_result))
                    
                    df_result = df_result[~mask]
            
            logger.info(f"    ✂️ Deleted rows: {rows_before} → {len(df_result)} rows")
        
        # Filter
        if 'filter' in op:
            filter_ops = op['filter']
            rows_before = len(df_result)
            
            for column, condition in filter_ops.items():
                if column in df_result.columns:
                    if isinstance(condition, dict):
                        operator = condition.get('operator', '==')
                        value = condition.get('value')
                        
                        if operator == '==':
                            df_result = df_result[df_result[column] == value]
                        elif operator == '!=':
                            df_result = df_result[df_result[column] != value]
                        elif operator == 'in':
                            df_result = df_result[df_result[column].isin(value if isinstance(value, list) else [value])]
                        elif operator == '>':
                            df_result = df_result[df_result[column] > value]
                        elif operator == '<':
                            df_result = df_result[df_result[column] < value]
                        elif operator == '>=':
                            df_result = df_result[df_result[column] >= value]
                        elif operator == '<=':
                            df_result = df_result[df_result[column] <= value]
            
            logger.info(f"    🔍 Filtered: {rows_before} → {len(df_result)} rows")
        
        # Sort
        if 'sort' in op:
            sort_spec = op['sort']
            by_columns = sort_spec.get('by', [])
            ascending = sort_spec.get('ascending', True)
            
            if by_columns:
                valid_columns = [col for col in by_columns if col in df_result.columns]
                if valid_columns:
                    df_result = df_result.sort_values(by=valid_columns, ascending=ascending)
                    logger.info(f"    📊 Sorted by {valid_columns}")
        
        return df_result
    
    def _apply_freeze_panes_from_operations(self, file_path: Path, operations_list: list, sheets: list):
        """Apply freeze panes from any operation that requested it"""
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            
            # Check if any operation has freeze_panes
            freeze_operations = [op for op in operations_list if 'freeze_panes' in op]
            
            if not freeze_operations:
                return
            
            wb = load_workbook(file_path)
            
            for freeze_op in freeze_operations:
                freeze_spec = freeze_op['freeze_panes']
                freeze_row = freeze_spec.get('row', 1)
                freeze_col = freeze_spec.get('col', 0)
                
                # Convert to 1-indexed for openpyxl
                freeze_row_idx = freeze_row + 1 if freeze_row > 0 else 1
                freeze_col_idx = freeze_col + 1 if freeze_col > 0 else 1
                
                # Determine which sheets to apply to
                if freeze_op.get('apply_to_all_sheets'):
                    # Apply to all sheets
                    for sheet_info in sheets:
                        sheet_name = sheet_info["sheet_name"]
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            col_letter = get_column_letter(freeze_col_idx)
                            freeze_cell = f"{col_letter}{freeze_row_idx}"
                            ws.freeze_panes = freeze_cell
                            logger.info(f"    ❄️ Froze panes at {freeze_cell} in '{sheet_name}'")
                
                elif 'target_sheet' in freeze_op:
                    # Apply to specific sheet
                    target_sheet = freeze_op['target_sheet']
                    for sheet_info in sheets:
                        if sheet_info["sheet_name"].lower() == target_sheet.lower():
                            sheet_name = sheet_info["sheet_name"]
                            if sheet_name in wb.sheetnames:
                                ws = wb[sheet_name]
                                col_letter = get_column_letter(freeze_col_idx)
                                freeze_cell = f"{col_letter}{freeze_row_idx}"
                                ws.freeze_panes = freeze_cell
                                logger.info(f"    ❄️ Froze panes at {freeze_cell} in '{sheet_name}'")
                            break
            
            wb.save(file_path)
            wb.close()
            
        except Exception as e:
            logger.error(f"⚠️ Error applying freeze panes: {e}")
    
    def _apply_formatting_operations_to_multi_sheet(self, file_path: Path, operations_list: list, sheets: list):
        """Apply all formatting operations (auto filters, subtotals, highlighting, etc.) to multi-sheet file"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter
            
            wb = load_workbook(file_path)
            
            # Apply auto filters to all sheets by default
            for sheet_info in sheets:
                sheet_name = sheet_info["sheet_name"]
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if ws.max_row > 1 and ws.max_column > 0:
                        ws.auto_filter.ref = ws.dimensions
                        logger.info(f"    🔽 Applied auto filter to '{sheet_name}'")
            
            # Process each operation for formatting
            for operation in operations_list:
                # Determine target sheets
                target_sheets = []
                
                if operation.get('apply_to_all_sheets'):
                    target_sheets = [s["sheet_name"] for s in sheets]
                elif 'target_sheet' in operation:
                    target_sheet = operation['target_sheet']
                    for sheet_info in sheets:
                        if sheet_info["sheet_name"].lower() == target_sheet.lower():
                            target_sheets.append(sheet_info["sheet_name"])
                            break
                
                # Apply subtotals
                if 'subtotals' in operation and target_sheets:
                    for sheet_name in target_sheets:
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            # Find the sheet's DataFrame
                            sheet_df = next((s["df"] for s in sheets if s["sheet_name"] == sheet_name), None)
                            if sheet_df is not None:
                                self._apply_subtotals_to_sheet(ws, sheet_df, operation['subtotals'])
                                logger.info(f"    ➕ Applied subtotals to '{sheet_name}'")
                
                # Apply highlighting (highlight_rows)
                if 'highlight_rows' in operation and target_sheets:
                    highlights = operation['highlight_rows']
                    if not isinstance(highlights, list):
                        highlights = [highlights]
                    
                    for sheet_name in target_sheets:
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            sheet_df = next((s["df"] for s in sheets if s["sheet_name"] == sheet_name), None)
                            if sheet_df is not None:
                                for hl_spec in highlights:
                                    self._apply_highlighting_to_sheet(ws, sheet_df, hl_spec)
                                logger.info(f"    🎨 Applied highlighting to '{sheet_name}'")
                
                # Apply cell highlighting (highlight_cells)
                if 'highlight_cells' in operation and target_sheets:
                    highlights = operation['highlight_cells']
                    if not isinstance(highlights, list):
                        highlights = [highlights]
                    
                    for sheet_name in target_sheets:
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            sheet_df = next((s["df"] for s in sheets if s["sheet_name"] == sheet_name), None)
                            if sheet_df is not None:
                                for hl_spec in highlights:
                                    self._apply_cell_highlighting_to_sheet(ws, sheet_df, hl_spec)
                                logger.info(f"    🎨 Applied cell highlighting to '{sheet_name}'")
                
                # Apply conditional formatting (conditional_format)
                if 'conditional_format' in operation:
                    cf_spec = operation['conditional_format']
                    column = cf_spec.get('column')
                    rules = cf_spec.get('rules', [])
                    
                    # Apply to all sheets or target sheets
                    sheets_to_apply = target_sheets if target_sheets else [s["sheet_name"] for s in sheets]
                    
                    for sheet_name in sheets_to_apply:
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            sheet_df = next((s["df"] for s in sheets if s["sheet_name"] == sheet_name), None)
                            if sheet_df is not None and column in sheet_df.columns:
                                # Apply each rule as cell highlighting
                                for rule in rules:
                                    hl_spec = {
                                        'column': column,
                                        'condition': rule.get('condition', {}),
                                        'color': rule.get('color', 'yellow')
                                    }
                                    self._apply_cell_highlighting_to_sheet(ws, sheet_df, hl_spec)
                                logger.info(f"    🎨 Applied {len(rules)} conditional format rules to '{sheet_name}'")
            
            wb.save(file_path)
            wb.close()
            logger.info(f"✅ Applied all formatting operations to multi-sheet file")
            
        except Exception as e:
            logger.error(f"⚠️ Error applying formatting operations: {e}", exc_info=True)
    
    def _apply_subtotals_to_sheet(self, ws, df: pd.DataFrame, subtotal_spec: Dict[str, Any]):
        """Apply subtotals to a specific sheet"""
        try:
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter
            
            group_by = subtotal_spec.get('group_by')
            aggregate_column = subtotal_spec.get('aggregate_column')
            function = subtotal_spec.get('function', 'count').upper()
            
            if not group_by or group_by not in df.columns:
                return
            if not aggregate_column or aggregate_column not in df.columns:
                return
            
            # Get column indices
            group_col_idx = df.columns.get_loc(group_by) + 1
            agg_col_idx = df.columns.get_loc(aggregate_column) + 1
            
            # Track group changes
            current_group = None
            subtotal_rows = []
            
            for row_idx in range(2, ws.max_row + 1):
                group_value = ws.cell(row=row_idx, column=group_col_idx).value
                
                if current_group is not None and group_value != current_group:
                    subtotal_rows.append((row_idx, current_group))
                
                current_group = group_value
            
            if current_group is not None:
                subtotal_rows.append((ws.max_row + 1, current_group))
            
            # Insert subtotal rows (in reverse)
            for insert_row, group_name in reversed(subtotal_rows):
                ws.insert_rows(insert_row)
                ws.cell(row=insert_row, column=group_col_idx, value=f"{group_name} Count")
                
                # Find range
                start_row = insert_row - 1
                while start_row > 1 and ws.cell(row=start_row, column=group_col_idx).value == group_name:
                    start_row -= 1
                start_row += 1
                end_row = insert_row - 1
                
                # Calculate the numeric value
                if function == 'COUNT':
                    calculated_value = sum(1 for r in range(start_row, end_row + 1) 
                                          if ws.cell(row=r, column=agg_col_idx).value is not None)
                elif function == 'SUM':
                    calculated_value = sum(float(ws.cell(row=r, column=agg_col_idx).value or 0) 
                                          for r in range(start_row, end_row + 1))
                else:  # AVERAGE
                    values = [float(ws.cell(row=r, column=agg_col_idx).value or 0) 
                             for r in range(start_row, end_row + 1)]
                    calculated_value = sum(values) / len(values) if values else 0
                
                # Set the numeric value (not formula)
                cell = ws.cell(row=insert_row, column=agg_col_idx)
                cell.value = calculated_value
                cell.number_format = '0' if function == 'COUNT' else '0.00'
                
                # Format
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=insert_row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='FFE0E0E0', end_color='FFE0E0E0', fill_type='solid')
                
        except Exception as e:
            logger.error(f"⚠️ Subtotal application failed: {e}")
    
    def _resolve_column_name(self, column_ref: str, df: pd.DataFrame) -> str:
        """
        Resolve column reference to actual column name
        Handles both "Column V" and actual names like "RAG Rating"
        """
        import re
        
        # If it's already a valid column name, return it
        if column_ref in df.columns:
            return column_ref
        
        # Check if it's a column letter reference (e.g., "Column V")
        match = re.match(r'Column ([A-Z]+)', column_ref, re.IGNORECASE)
        if match:
            letter = match.group(1).upper()
            # Convert letter to index (A=1, B=2, ..., V=22)
            col_idx = 0
            for char in letter:
                col_idx = col_idx * 26 + (ord(char) - ord('A') + 1)
            
            # Get column name at that index
            if 0 < col_idx <= len(df.columns):
                actual_name = df.columns[col_idx - 1]
                logger.info(f"Resolved '{column_ref}' to '{actual_name}'")
                return actual_name
        
        # Return original if no mapping found
        logger.warning(f"Could not resolve column '{column_ref}'")
        return column_ref
    
    def _apply_highlighting_to_sheet(self, ws, df: pd.DataFrame, hl_spec: Dict[str, Any]):
        """Apply row highlighting to a specific sheet"""
        try:
            from openpyxl.styles import PatternFill
            
            column_ref = hl_spec.get('column')
            condition = hl_spec.get('condition', {})
            color = hl_spec.get('color', 'yellow')
            
            # Resolve column name (handles "Column V" → "RAG Rating")
            column = self._resolve_column_name(column_ref, df)
            
            if not column or column not in df.columns:
                logger.warning(f"Column '{column}' not found in DataFrame. Available: {list(df.columns)}")
                return
            
            # Color mapping - match Excel's conditional formatting colors
            color_map = {
                'red': 'FFC7CE', 'green': 'C6EFCE', 'yellow': 'FFFF99',  # Excel conditional format colors
                'amber': 'FFEB9C',  # Excel amber/orange for RAG rating
                'blue': '0000FF', 'orange': 'FFA500', 'purple': '800080',
                'pink': 'FFC0CB', 'cyan': '00FFFF', 'light_green': '90EE90',
                'light_blue': 'ADD8E6', 'light_yellow': 'FFFFE0',
                'light_red': 'FFCCCB', 'gray': 'D3D3D3'
            }
            fill_color = color_map.get(color.lower(), 'FFFF99')
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            col_idx = df.columns.get_loc(column) + 1
            
            # Apply highlighting
            for row_idx in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                
                if self._check_condition(cell_value, condition):
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col).fill = fill
                        
        except Exception as e:
            logger.error(f"⚠️ Highlighting failed: {e}")
    
    def _apply_cell_highlighting_to_sheet(self, ws, df: pd.DataFrame, hl_spec: Dict[str, Any]):
        """Apply cell highlighting to a specific sheet"""
        try:
            from openpyxl.styles import PatternFill
            
            column_ref = hl_spec.get('column')
            condition = hl_spec.get('condition', {})
            color = hl_spec.get('color', 'yellow')
            
            # Resolve column name (handles "Column V" → "RAG Rating")
            column = self._resolve_column_name(column_ref, df)
            
            if not column or column not in df.columns:
                logger.warning(f"Column '{column}' not found in DataFrame. Available: {list(df.columns)}")
                return
            
            # Color mapping - match Excel's conditional formatting colors
            color_map = {
                'red': 'FFC7CE', 'green': 'C6EFCE', 'yellow': 'FFFF99',  # Excel conditional format colors
                'amber': 'FFEB9C',  # Excel amber/orange for RAG rating
                'blue': '0000FF', 'orange': 'FFA500', 'purple': '800080',
                'pink': 'FFC0CB', 'cyan': '00FFFF', 'light_green': '90EE90',
                'light_blue': 'ADD8E6', 'light_yellow': 'FFFFE0',
                'light_red': 'FFCCCB', 'gray': 'D3D3D3'
            }
            fill_color = color_map.get(color.lower(), 'FFFF99')
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            col_idx = df.columns.get_loc(column) + 1
            
            # Apply highlighting to cells only
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value
                
                if self._check_condition(cell_value, condition):
                    cell.fill = fill
                        
        except Exception as e:
            logger.error(f"⚠️ Cell highlighting failed: {e}")
    
    def _check_condition(self, value, condition):
        """Check if a value meets a condition"""
        try:
            if isinstance(condition, dict):
                operator = condition.get('operator', '==')
                target = condition.get('value')
                
                if operator == '==':
                    return value == target
                elif operator == '!=':
                    return value != target
                elif operator == '>':
                    return value > target
                elif operator == '<':
                    return value < target
                elif operator == '>=':
                    return value >= target
                elif operator == '<=':
                    return value <= target
                elif operator == 'contains':
                    return target in str(value) if value else False
                elif operator == 'in':
                    return value in target
                elif operator == 'regex':
                    # Support regex pattern matching
                    import re
                    if value is None:
                        return False
                    pattern = re.compile(target)
                    return bool(pattern.search(str(value)))
            else:
                return value == condition
        except Exception as e:
            logger.warning(f"Condition check failed: {e}")
            return False
    
    def _generate_operation_summary(self, operations: Dict[str, Any], df: pd.DataFrame) -> str:
        """Generate a human-readable summary using LLM"""
        logger.info(f"🔍 GENERATING SUMMARY - Operations: {operations}")
        logger.info(f"🔍 GENERATING SUMMARY - DataFrame shape: {df.shape}")
        
        # Build operation description for LLM
        operation_details = []
        
        # Filtering
        if 'filter' in operations:
            filter_ops = operations['filter']
            logger.info(f"🔍 FILTER OPERATIONS: {filter_ops}")
            
            for column, condition in filter_ops.items():
                if isinstance(condition, dict):
                    op = condition.get('operator', '==')
                    val = condition.get('value')
                    logger.info(f"🔍 DYNAMIC VALUES - Column: {column}, Operator: {op}, Value: {val}")
                    
                    if op == '==':
                        operation_details.append(f"Filtered {column} to show only '{val}'")
                    elif op == '!=':
                        operation_details.append(f"Filtered {column} to exclude '{val}'")
                    elif op == '>':
                        operation_details.append(f"Filtered {column} greater than {val}")
                    elif op == '<':
                        operation_details.append(f"Filtered {column} less than {val}")
                    elif op == '>=':
                        operation_details.append(f"Filtered {column} at least {val}")
                    elif op == '<=':
                        operation_details.append(f"Filtered {column} at most {val}")
                    else:
                        operation_details.append(f"Filtered {column} {op} {val}")
                else:
                    operation_details.append(f"Filtered {column} to '{condition}'")
        
        # Sorting
        if 'sort' in operations:
            sort_ops = operations['sort']
            by_cols = sort_ops.get('by', [])
            ascending = sort_ops.get('ascending', True)
            direction = "ascending" if ascending else "descending"
            operation_details.append(f"Sorted by {', '.join(by_cols)} ({direction})")
        
        # Highlighting - handle both single dict and list of dicts
        if 'highlight_rows' in operations:
            hl = operations['highlight_rows']
            # Handle list of multiple highlights
            if isinstance(hl, list):
                for highlight in hl:
                    column = highlight.get('column')
                    color = highlight.get('color', 'yellow')
                    operation_details.append(f"Highlighted rows in {column} with {color} color")
            else:
                column = hl.get('column')
                color = hl.get('color', 'yellow')
                operation_details.append(f"Highlighted rows in {column} with {color} color")
        
        if 'highlight_cells' in operations:
            hl = operations['highlight_cells']
            # Handle list of multiple highlights
            if isinstance(hl, list):
                for highlight in hl:
                    column = highlight.get('column')
                    color = highlight.get('color', 'yellow')
                    operation_details.append(f"Highlighted cells in {column} with {color} color")
            else:
                column = hl.get('column')
                color = hl.get('color', 'yellow')
                operation_details.append(f"Highlighted cells in {column} with {color} color")
        
        # Conditional formatting
        if 'conditional_format' in operations:
            cf = operations['conditional_format']
            column = cf.get('column')
            rules = cf.get('rules', [])
            operation_details.append(f"Applied {len(rules)} conditional formatting rules to {column}")
        
        # Duplicates
        if 'highlight_duplicates' in operations:
            dup = operations['highlight_duplicates']
            columns = dup.get('columns', [])
            color = dup.get('color', 'light_red')
            operation_details.append(f"Highlighted duplicates in {len(columns)} columns with {color} color")
        
        # Nulls
        if 'highlight_nulls' in operations:
            null = operations['highlight_nulls']
            columns = null.get('columns', [])
            color = null.get('color', 'gray')
            operation_details.append(f"Highlighted missing values in {len(columns)} columns with {color} color")
        
        # Top/Bottom N
        if 'top_n' in operations:
            n = operations['top_n']
            operation_details.append(f"Included top {n} records")
        
        if 'bottom_n' in operations:
            n = operations['bottom_n']
            operation_details.append(f"Included bottom {n} records")
        
        # Target sheet
        target_sheet_info = ""
        if 'target_sheet' in operations:
            target_sheet = operations['target_sheet']
            target_sheet_info = f" from sheet '{target_sheet}'"
            operation_details.append(f"Working with sheet: {target_sheet}")
        
        # Remove last row
        if operations.get('remove_last_row'):
            operation_details.append(f"Removed last row (filter description){target_sheet_info}")
        
        # Delete specific rows
        if 'delete_rows' in operations:
            delete_spec = operations['delete_rows']
            if 'row_numbers' in delete_spec:
                row_nums = delete_spec['row_numbers']
                operation_details.append(f"Deleted {len(row_nums)} specific rows")
            elif 'column' in delete_spec:
                column = delete_spec['column']
                operation_details.append(f"Deleted rows matching condition in {column}")
        
        # Freeze panes
        if 'freeze_panes' in operations:
            freeze_spec = operations['freeze_panes']
            freeze_row = freeze_spec.get('row', 1)
            freeze_col = freeze_spec.get('col', 0)
            if freeze_row > 0 and freeze_col > 0:
                operation_details.append(f"Froze top {freeze_row} row(s) and {freeze_col} column(s)")
            elif freeze_row > 0:
                operation_details.append(f"Froze top {freeze_row} row(s)")
            elif freeze_col > 0:
                operation_details.append(f"Froze {freeze_col} column(s)")
        
        # Get file info
        row_count = len(df)
        col_count = len(df.columns)
        logger.info(f"🔍 DYNAMIC ROW COUNT: {row_count}")
        logger.info(f"🔍 DYNAMIC COLUMN COUNT: {col_count}")
        
        # If no operations were performed (only custom_filename), return minimal message
        if not operation_details:
            logger.info("🔍 No operations performed - only custom filename")
            return f"📄 {row_count} rows × {col_count} columns"
        
        # Use LLM to generate natural summary
        operations_text = "\n".join([f"- {detail}" for detail in operation_details])
        
        prompt = f"""You are a helpful data assistant. Generate a brief, specific message about the Excel file operations that were just completed.

Operations performed:
{operations_text}

File details:
- {row_count} rows
- {col_count} columns

Instructions:
1. Focus on WHAT WAS DONE, not generic greetings
2. Be specific about the operations performed
3. Include the row and column count naturally
4. Keep it brief (1-2 sentences)
5. DO NOT use markdown formatting like ** or ###
6. Use plain text only
7. You can use emojis like ✅ 📊 📄
8. DO NOT start with generic phrases like "Okay", "Done", "Ready", "Created your Excel file"
9. Start directly with the operation details

Example format:
- "✅ Filtered data to show only Status = Invited - 53 rows and 9 columns."
- "📊 Applied highlighting to Active status rows - 120 rows across 8 columns."
- "✅ Sorted by Salary (descending) and filtered Department = Sales - 45 rows, 12 columns."

Generate the message:"""

        try:
            import google.generativeai as genai
            response = self.model.generate_content(
                prompt,
                generation_config=genai.types.GenerationConfig(
                    temperature=0.7,  # Higher for more varied responses
                    top_p=0.9,
                    top_k=40
                )
            )
            
            llm_summary = response.text.strip()
            logger.info(f"🔍 LLM GENERATED SUMMARY: {llm_summary}")
            return llm_summary
            
        except Exception as e:
            logger.error(f"❌ Error generating LLM summary: {e}")
            # Fallback to simple summary
            ops_summary = ", ".join(operation_details) if operation_details else "processed your data"
            return f"✅ Excel file created! {ops_summary}. File contains {row_count} rows and {col_count} columns."
    
    def get_session_info(self, session_id: str) -> Dict[str, Any]:
        """Get information about a loaded session"""
        if session_id not in self.session_data:
            return {"success": False, "error": "Session not found"}
        
        session = self.session_data[session_id]
        files = session["files"]
        history = session["conversation_history"]
        
        # Get active DataFrame info
        df = self.get_active_dataframe(session_id)
        if df is None:
            return {"success": False, "error": "No active data in session"}
        
        return {
            "success": True,
            "session_id": session_id,
            "total_files": len(files),
            "active_file": session.get("active_file"),
            "data_shape": f"{len(df)} rows × {len(df.columns)} columns",
            "memory_usage_mb": df.memory_usage(deep=True).sum() / 1024 / 1024,
            "columns": list(df.columns),
            "conversation_count": len(history),
            "files_info": {
                file_id: {
                    "filename": file_data["filename"],
                    "rows": len(file_data["df"]),
                    "columns": len(file_data["df"].columns)
                }
                for file_id, file_data in files.items()
            }
        }
    
    def _generate_operation_success_message(
        self, 
        operations_applied: list, 
        user_query: str, 
        rows: int, 
        columns: int, 
        sheet_type: str = "Google Sheet"
    ) -> str:
        """Generate contextual success message using LLM based on operations applied"""
        try:
            # Create context for LLM
            operations_text = ', '.join(operations_applied) if operations_applied else "data processing"
            
            prompt = f"""Generate a natural, contextual success message for a user who requested data operations.

Context:
- User query: "{user_query}"
- Operations applied: {operations_text}
- Sheet type: {sheet_type}
- Data size: {rows} rows, {columns} columns

Requirements:
- Be conversational and natural (not robotic)
- Reference the specific operations that were completed
- Keep it concise (1-2 sentences max)
- Use appropriate emojis sparingly
- Don't repeat the user's exact words
- Focus on what was accomplished

Examples of good messages:
- "✅ Successfully merged your data into separate tabs as requested!"
- "🎯 Your sheets have been organized with the new tab names!"
- "✅ Data filtering and highlighting completed - check out the results!"
- "🔄 Merge operation finished! Your files are now combined in separate tabs."

Generate the success message:"""

            import google.generativeai as genai
            response = self.model.generate_content(
                prompt,
                generation_config=genai.types.GenerationConfig(
                    temperature=0.7,  # Higher for more varied, natural responses
                    top_p=0.9,
                    top_k=32,
                    max_output_tokens=100  # Keep it concise
                )
            )
            
            generated_message = response.text.strip()
            
            # Fallback to simple message if LLM fails or returns empty
            if not generated_message or len(generated_message) < 10:
                return f"✅ Successfully applied {operations_text} to {sheet_type}"
            
            return generated_message
            
        except Exception as e:
            logger.warning(f"⚠️ Failed to generate LLM message: {e}")
            # Fallback to simple message
            operations_text = ', '.join(operations_applied) if operations_applied else "operations"
            return f"✅ Successfully applied {operations_text} to {sheet_type}"
    
    def _generate_dynamic_operations_summary(
        self,
        operations_list: list,
        user_query: str,
        sheets_count: int,
        total_rows: int
    ) -> str:
        """Generate dynamic summary of operations performed using LLM"""
        try:
            # Create a concise operations summary for the LLM
            operations_summary = []
            for op in operations_list:
                # Extract key operation details
                op_details = []
                for key, value in op.items():
                    if key in ['merge_files', 'rename_tabs', 'delete_rows', 'sort', 'highlight', 
                              'conditional_formatting', 'subtotals', 'freeze_panes', 'auto_filter', 'fit_columns']:
                        op_details.append(f"{key}: {value}")
                if op_details:
                    operations_summary.append(", ".join(op_details))
            
            prompt = f"""Generate a concise, professional success message for Excel operations completion.

User's original request: "{user_query}"

Operations performed:
{chr(10).join(f"- {op}" for op in operations_summary[:5])}  # Show first 5 operations

Results:
- {sheets_count} sheets processed
- {total_rows} total rows
- Multi-sheet Excel file created

Requirements:
- Start with "✅ Successfully"
- Mention 2-3 key operations performed (use natural language, not technical terms)
- Include sheet count and row count
- End with "Download your multi-sheet Excel file."
- Keep it under 100 words
- Be conversational and informative

Example format: "✅ Successfully merged files into separate tabs, applied sorting and filters, and highlighted key data across 2 sheets with 207 total rows. Download your multi-sheet Excel file."

Generate the success message:"""

            import google.generativeai as genai
            response = self.model.generate_content(
                prompt,
                generation_config=genai.types.GenerationConfig(
                    temperature=0.7,  # Moderate creativity for natural language
                    top_p=0.9,
                    top_k=32,
                    max_output_tokens=150  # Keep it concise
                )
            )
            
            generated_summary = response.text.strip()
            
            # Fallback to simple message if LLM fails or returns empty
            if not generated_summary or len(generated_summary) < 20:
                return f"✅ Successfully processed your request across {sheets_count} sheets with {total_rows} total rows. Download your multi-sheet Excel file."
            
            return generated_summary
            
        except Exception as e:
            logger.warning(f"⚠️ Failed to generate LLM operations summary: {e}")
            # Fallback to simple message
            return f"✅ Successfully processed your request across {sheets_count} sheets with {total_rows} total rows. Download your multi-sheet Excel file."

    def _generate_dynamic_chart_response(
        self,
        chart_type: str,
        target_column: str,
        data_counts: dict,
        user_query: str
    ) -> str:
        """Generate dynamic response for chart creation using LLM"""
        try:
            prompt = f"""Generate a natural, informative response for a chart creation request.

Context:
- User query: "{user_query}"
- Chart type: {chart_type}
- Column analyzed: {target_column}
- Data distribution: {dict(list(data_counts.items())[:5])}  # Top 5 categories
- Total categories: {len(data_counts)}
- Total records: {sum(data_counts.values())}

Requirements:
- Be conversational and informative
- Explain what the chart shows
- Highlight key insights from the data distribution
- Keep it concise (1-2 sentences)
- Use appropriate emojis sparingly
- Don't repeat the user's exact words

Examples:
- "📊 Created a pie chart showing the distribution across {len(data_counts)} categories. The data reveals interesting patterns in your {target_column} field!"
- "🎯 Your {target_column} breakdown is now visualized! The chart shows {sum(data_counts.values())} records distributed across {len(data_counts)} different categories."
- "📈 Generated a comprehensive view of your {target_column} data - the visualization highlights the key patterns in your dataset!"

Generate a natural response:"""

            response = self.llm.invoke(prompt, temperature=0.7, max_tokens=100)
            
            if response and len(response.strip()) > 10:
                return response.strip()
            else:
                # Fallback to informative default
                return f"📊 Created {chart_type} chart showing {target_column} distribution across {len(data_counts)} categories with {sum(data_counts.values())} total records."
                
        except Exception as e:
            logger.error(f"❌ Error generating dynamic chart response: {e}")
            # Fallback to informative default
            return f"📊 Created {chart_type} chart showing {target_column} distribution across {len(data_counts)} categories with {sum(data_counts.values())} total records."
    
    def _generate_dynamic_search_response(
        self,
        search_term: str,
        found_records: int,
        user_query: str,
        success: bool = True
    ) -> str:
        """Generate dynamic response for search operations using LLM"""
        try:
            if success:
                prompt = f"""Generate a natural response for a successful data search.

Context:
- User query: "{user_query}"
- Search term: "{search_term}"
- Records found: {found_records}

Requirements:
- Be conversational and helpful
- Confirm what was found
- Keep it concise (1 sentence)
- Use appropriate emojis sparingly
- Don't repeat the user's exact words

Examples:
- "✅ Found {found_records} record(s) matching '{search_term}' in your dataset!"
- "🎯 Located {found_records} entries for '{search_term}' - here are the details:"
- "📋 Successfully retrieved {found_records} record(s) containing '{search_term}'."

Generate a natural response:"""
            else:
                prompt = f"""Generate a helpful response for when a search finds no results.

Context:
- User query: "{user_query}"
- Search term: "{search_term}"
- Records found: 0

Requirements:
- Be helpful and suggest alternatives
- Keep it concise (1-2 sentences)
- Use appropriate emojis sparingly
- Offer to help with alternatives

Examples:
- "❌ No records found for '{search_term}'. Try checking the spelling or searching for partial matches."
- "🔍 Couldn't locate '{search_term}' in the dataset. Would you like to see what data is available?"
- "⚠️ No matches for '{search_term}'. The search might be case-sensitive or the term might not exist in this dataset."

Generate a helpful response:"""

            response = self.llm.invoke(prompt, temperature=0.7, max_tokens=100)
            
            if response and len(response.strip()) > 10:
                return response.strip()
            else:
                # Fallback to informative default
                if success:
                    return f"✅ Found {found_records} record(s) matching '{search_term}' in your dataset!"
                else:
                    return f"❌ No records found for '{search_term}'. Try checking the spelling or searching for partial matches."
                
        except Exception as e:
            logger.error(f"❌ Error generating dynamic search response: {e}")
            # Fallback to informative default
            if success:
                return f"✅ Found {found_records} record(s) matching '{search_term}' in your dataset!"
            else:
                return f"❌ No records found for '{search_term}'. Try checking the spelling or searching for partial matches."
    
    def clear_session(self, session_id: str) -> bool:
        """Clear session data from memory"""
        try:
            if session_id in self.session_data:
                del self.session_data[session_id]
                logger.info(f"🗑️ Cleared session: {session_id}")
                return True
            else:
                logger.warning(f"⚠️ Session not found: {session_id}")
                return False
        except Exception as e:
            logger.error(f"❌ Error clearing session {session_id}: {e}")
            return False
